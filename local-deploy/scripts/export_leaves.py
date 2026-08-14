"""
Merkoteks LOCAL Backup — Operational Export (Human-readable)
=============================================================
Generates 4 human-readable files for business continuity:
  1. Yillik_Izin_Tum_Kayitlar.xlsx  (2 sheets: TUM_IZINLER, SON_ISLENENLER)
  2. Yillik_Izin_Tum_Kayitlar.csv   (UTF-8 BOM)
  3. SON_ISLENEN_IZINLER.txt         (last 100 records, newest first)
  4. EN_SON_IZIN.txt                 (single most recent record)

Usage inside container:
  python /tmp/export_leaves.py /tmp/backup_out
"""
import sys, os, csv, io
from datetime import datetime, timezone
from pymongo import MongoClient
from openpyxl import Workbook

OUT_DIR = sys.argv[1] if len(sys.argv) > 1 else "/tmp/backup_out"
os.makedirs(OUT_DIR, exist_ok=True)

client = MongoClient(os.environ.get("MONGO_URL", "mongodb://mongodb:27017"))
db = client[os.environ.get("DB_NAME", "merkoteks_hr")]

# Load personnel lookup (id -> personnel record)
pmap = {p["id"]: p for p in db.personnel.find({}, {"_id": 0})}

# Load ALL yıllık izin (leaves collection; special_leaves ayrı collection'da)
leaves = list(db.leaves.find({"$or": [{"deleted": {"$ne": True}}, {"deleted": {"$exists": False}}]}, {"_id": 0}))

# Enrich + normalize rows
rows = []
for L in leaves:
    p = pmap.get(L.get("personnel_id"), {})
    rows.append({
        "Sicil No": p.get("sicil_no", ""),
        "Ad Soyad": p.get("ad_soyad", ""),
        "Departman": p.get("departman", ""),
        "Şirket": p.get("sirket", ""),
        "İzin Başlangıç Tarihi": L.get("start_date", ""),
        "İzin Bitiş Tarihi": L.get("end_date", ""),
        "İşe Başlama Tarihi": p.get("ise_giris", ""),
        "İzin Gün Sayısı": L.get("days", ""),
        "İzin Hakediş / Dönem Yılı": L.get("hakedis_yili", "") or L.get("year", "") or "",
        "Kayıt Oluşturma Tarihi": L.get("created_at", ""),
        "Son Güncelleme Tarihi": L.get("updated_at", "") or L.get("created_at", ""),
        "Kayıt ID": L.get("id", ""),
        "Personel Durumu": "Aktif" if p.get("aktif", True) else "İşten Ayrılmış",
    })

# --- 1) XLSX (2 sheets) ---
wb = Workbook()
ws1 = wb.active
ws1.title = "TUM_IZINLER"
sorted_by_start = sorted(rows, key=lambda r: (r["İzin Başlangıç Tarihi"] or ""))
if rows:
    headers = list(rows[0].keys())
    ws1.append(headers)
    for r in sorted_by_start:
        ws1.append([r[h] for h in headers])
ws2 = wb.create_sheet("SON_ISLENENLER")
sorted_by_created_desc = sorted(rows, key=lambda r: (r["Kayıt Oluşturma Tarihi"] or ""), reverse=True)
if rows:
    ws2.append(headers)
    for r in sorted_by_created_desc:
        ws2.append([r[h] for h in headers])
xlsx_path = os.path.join(OUT_DIR, "Yillik_Izin_Tum_Kayitlar.xlsx")
wb.save(xlsx_path)

# --- 2) CSV (UTF-8 BOM) ---
csv_path = os.path.join(OUT_DIR, "Yillik_Izin_Tum_Kayitlar.csv")
with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
    if rows:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        for r in sorted_by_start:
            w.writerow(r)

# --- 3) SON_ISLENEN_IZINLER.txt (last 100) ---
txt_path = os.path.join(OUT_DIR, "SON_ISLENEN_IZINLER.txt")
now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
with open(txt_path, "w", encoding="utf-8") as f:
    f.write(f"MERKOTEKS PERSONEL SİSTEMİ — Son İşlenen Yıllık İzin Kayıtları\n")
    f.write(f"{'='*70}\n")
    f.write(f"Backup tarihi     : {now}\n")
    f.write(f"Toplam yıllık izin: {len(rows)} kayıt\n")
    f.write(f"{'='*70}\n\n")
    for r in sorted_by_created_desc[:100]:
        f.write(f"Sicil No               : {r['Sicil No']}\n")
        f.write(f"Ad Soyad               : {r['Ad Soyad']}\n")
        f.write(f"İzin Başlangıç Tarihi  : {r['İzin Başlangıç Tarihi']}\n")
        f.write(f"İzin Bitiş Tarihi      : {r['İzin Bitiş Tarihi']}\n")
        f.write(f"İşe Başlama Tarihi     : {r['İşe Başlama Tarihi']}\n")
        f.write(f"İzin Gün Sayısı        : {r['İzin Gün Sayısı']}\n")
        f.write(f"Kayıt Oluşturma Tarihi : {r['Kayıt Oluşturma Tarihi']}\n")
        f.write(f"Kayıt ID               : {r['Kayıt ID']}\n")
        f.write("-" * 70 + "\n")

# --- 4) EN_SON_IZIN.txt (single most recent) ---
last_path = os.path.join(OUT_DIR, "EN_SON_IZIN.txt")
with open(last_path, "w", encoding="utf-8") as f:
    f.write(f"MERKOTEKS — SİSTEMDE EN SON İŞLENEN YILLIK İZİN\n")
    f.write(f"{'='*50}\n")
    f.write(f"Son yedek: {now}\n\n")
    if sorted_by_created_desc:
        r = sorted_by_created_desc[0]
        f.write(f"Son işlenen yıllık izin:\n\n")
        f.write(f"Sicil No           : {r['Sicil No']}\n")
        f.write(f"Ad Soyad           : {r['Ad Soyad']}\n")
        f.write(f"İzin Başlangıç     : {r['İzin Başlangıç Tarihi']}\n")
        f.write(f"İzin Bitiş         : {r['İzin Bitiş Tarihi']}\n")
        f.write(f"İşe Başlama        : {r['İşe Başlama Tarihi']}\n")
        f.write(f"İzin Gün Sayısı    : {r['İzin Gün Sayısı']}\n")
        f.write(f"Kayıt Oluşturma    : {r['Kayıt Oluşturma Tarihi']}\n")
        f.write(f"Kayıt ID           : {r['Kayıt ID']}\n")
    else:
        f.write("(Sistemde hiç yıllık izin kaydı yok)\n")

# --- Output for backup.bat to capture ---
mongo_count = len(rows)
# Excel actual row count (exclude header)
excel_row_count = ws1.max_row - 1 if rows else 0
# CSV row count
with open(csv_path, encoding="utf-8-sig") as f:
    csv_row_count = sum(1 for _ in f) - 1
    csv_row_count = max(0, csv_row_count)

match = mongo_count == excel_row_count == csv_row_count
print(f"MONGO_COUNT={mongo_count}")
print(f"EXCEL_ROW_COUNT={excel_row_count}")
print(f"CSV_ROW_COUNT={csv_row_count}")
print(f"COUNTS_MATCH={'YES' if match else 'NO'}")
print(f"XLSX_PATH={xlsx_path}")
print(f"CSV_PATH={csv_path}")
print(f"TXT_PATH={txt_path}")
print(f"LAST_PATH={last_path}")
sys.exit(0 if match else 2)
