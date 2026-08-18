from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import re
import uuid
import asyncio
import logging
from datetime import datetime, date, timezone, timedelta
from typing import List, Optional, Literal, Dict

import bcrypt
import hmac
import jwt
import httpx
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict

# -----------------------------------------------------------------------------
# Config / DB
# -----------------------------------------------------------------------------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGO = "HS256"
JWT_EXP_MIN = 60 * 24  # 24h for internal desktop app

app = FastAPI(title="Merkoteks Personel ve İzin Sistemi")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("merkoteks")

# -----------------------------------------------------------------------------
# Object Storage (Iter 59) — Emergent Object Storage varsayılan; STORAGE_MODE=local
# olduğunda yerel disk kullanılır (on-premise/LAN kurulum için).
# -----------------------------------------------------------------------------
_STORAGE_MODE = (os.environ.get("STORAGE_MODE") or "emergent").lower()
_UPLOAD_PATH = os.environ.get("UPLOAD_PATH", "/data/uploads")
if _STORAGE_MODE == "local":
    os.makedirs(_UPLOAD_PATH, exist_ok=True)
    log.info("Storage MODE=local, path=%s", _UPLOAD_PATH)

_STORAGE_BASE = (os.environ.get("INTEGRATION_PROXY_URL") or "").strip() \
    or "https://integrations.emergentagent.com"
_STORAGE_URL = _STORAGE_BASE.rstrip("/") + "/objstore/api/v1/storage"
_EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY", "")
_APP_NAME = "merkoteks_hr"
_storage_key: Optional[str] = None
_storage_lock: Optional[asyncio.Lock] = None


def _get_storage_lock() -> asyncio.Lock:
    global _storage_lock
    if _storage_lock is None:
        _storage_lock = asyncio.Lock()
    return _storage_lock


async def _init_storage(force: bool = False) -> str:
    if _STORAGE_MODE == "local":
        return "local"
    global _storage_key
    if _storage_key and not force:
        return _storage_key
    async with _get_storage_lock():
        if _storage_key and not force:
            return _storage_key
        if not _EMERGENT_KEY:
            raise HTTPException(status_code=500,
                                detail="EMERGENT_LLM_KEY tanımlı değil")
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(f"{_STORAGE_URL}/init",
                                  json={"emergent_key": _EMERGENT_KEY})
            r.raise_for_status()
            _storage_key = r.json()["storage_key"]
        log.info("Emergent object storage initialized")
        return _storage_key


async def _put_object(path: str, data: bytes, content_type: str) -> dict:
    if _STORAGE_MODE == "local":
        full = os.path.join(_UPLOAD_PATH, path)
        os.makedirs(os.path.dirname(full), exist_ok=True)
        with open(full, "wb") as f:
            f.write(data)
        import hashlib as _hl
        return {"path": path, "size": len(data), "etag": _hl.md5(data).hexdigest()}
    key = await _init_storage()
    async with httpx.AsyncClient(timeout=120) as client:
        r = await client.put(
            f"{_STORAGE_URL}/objects/{path}",
            headers={"X-Storage-Key": key, "Content-Type": content_type},
            content=data,
        )
        if r.status_code == 404:
            key = await _init_storage(force=True)
            r = await client.put(
                f"{_STORAGE_URL}/objects/{path}",
                headers={"X-Storage-Key": key, "Content-Type": content_type},
                content=data,
            )
        r.raise_for_status()
        return r.json()


async def _get_object(path: str) -> tuple:
    if _STORAGE_MODE == "local":
        full = os.path.join(_UPLOAD_PATH, path)
        if not os.path.exists(full):
            raise HTTPException(status_code=404, detail="Dosya bulunamadı")
        with open(full, "rb") as f:
            data = f.read()
        import mimetypes as _mt
        ct, _ = _mt.guess_type(path)
        return data, ct or "application/octet-stream"
    key = await _init_storage()
    async with httpx.AsyncClient(timeout=60) as client:
        r = await client.get(
            f"{_STORAGE_URL}/objects/{path}",
            headers={"X-Storage-Key": key},
        )
        if r.status_code == 404:
            key = await _init_storage(force=True)
            r = await client.get(
                f"{_STORAGE_URL}/objects/{path}",
                headers={"X-Storage-Key": key},
            )
        r.raise_for_status()
        return r.content, r.headers.get("Content-Type", "application/octet-stream")

# -----------------------------------------------------------------------------
# Utility: Turkish-aware text search
# -----------------------------------------------------------------------------
def _tr_lower(value: str) -> str:
    """Türkçe I/İ dönüşümünü doğru yaparak küçük harfe çevirir."""
    return (value or "").translate(str.maketrans({"I": "ı", "İ": "i"})).lower()


def _tr_search_regex(value: str) -> str:
    """MongoDB regex aramasında Türkçe büyük/küçük harf eşleşmesini güvenilir yapar."""
    equivalents = {
        "i": "[iİ]",
        "İ": "[iİ]",
        "ı": "[ıI]",
        "I": "[ıI]",
        "ş": "[şŞ]",
        "Ş": "[şŞ]",
        "ğ": "[ğĞ]",
        "Ğ": "[ğĞ]",
        "ü": "[üÜ]",
        "Ü": "[üÜ]",
        "ö": "[öÖ]",
        "Ö": "[öÖ]",
        "ç": "[çÇ]",
        "Ç": "[çÇ]",
    }
    return "".join(equivalents.get(ch, re.escape(ch)) for ch in (value or "").strip())


# -----------------------------------------------------------------------------
# Utility: password + jwt
# -----------------------------------------------------------------------------
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=JWT_EXP_MIN),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

async def get_current_user(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Oturum bulunamadı")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Oturum süresi doldu")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Geçersiz oturum")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
    return user

def require_roles(*roles: str):
    async def dep(user: dict = Depends(get_current_user)) -> dict:
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        return user
    return dep

# -----------------------------------------------------------------------------
# Turkish holidays (static 2024-2027 including religious days) — Arife = half day
# -----------------------------------------------------------------------------
STATIC_HOLIDAYS = [
    # Fixed
    ("2024-01-01", "Yılbaşı", "full"),
    ("2024-04-09", "Ramazan Bayramı Arifesi", "half"),
    ("2024-04-10", "Ramazan Bayramı 1. Gün", "full"),
    ("2024-04-11", "Ramazan Bayramı 2. Gün", "full"),
    ("2024-04-12", "Ramazan Bayramı 3. Gün", "full"),
    ("2024-04-23", "Ulusal Egemenlik ve Çocuk Bayramı", "full"),
    ("2024-05-01", "Emek ve Dayanışma Günü", "full"),
    ("2024-05-19", "Atatürk'ü Anma Gençlik ve Spor Bayramı", "full"),
    ("2024-06-15", "Kurban Bayramı Arifesi", "half"),
    ("2024-06-16", "Kurban Bayramı 1. Gün", "full"),
    ("2024-06-17", "Kurban Bayramı 2. Gün", "full"),
    ("2024-06-18", "Kurban Bayramı 3. Gün", "full"),
    ("2024-06-19", "Kurban Bayramı 4. Gün", "full"),
    ("2024-07-15", "Demokrasi ve Milli Birlik Günü", "full"),
    ("2024-08-30", "Zafer Bayramı", "full"),
    ("2024-10-28", "Cumhuriyet Bayramı Arifesi", "half"),
    ("2024-10-29", "Cumhuriyet Bayramı", "full"),
    # 2025
    ("2025-01-01", "Yılbaşı", "full"),
    ("2025-03-29", "Ramazan Bayramı Arifesi", "half"),
    ("2025-03-30", "Ramazan Bayramı 1. Gün", "full"),
    ("2025-03-31", "Ramazan Bayramı 2. Gün", "full"),
    ("2025-04-01", "Ramazan Bayramı 3. Gün", "full"),
    ("2025-04-23", "Ulusal Egemenlik ve Çocuk Bayramı", "full"),
    ("2025-05-01", "Emek ve Dayanışma Günü", "full"),
    ("2025-05-19", "Atatürk'ü Anma Gençlik ve Spor Bayramı", "full"),
    ("2025-06-05", "Kurban Bayramı Arifesi", "half"),
    ("2025-06-06", "Kurban Bayramı 1. Gün", "full"),
    ("2025-06-07", "Kurban Bayramı 2. Gün", "full"),
    ("2025-06-08", "Kurban Bayramı 3. Gün", "full"),
    ("2025-06-09", "Kurban Bayramı 4. Gün", "full"),
    ("2025-07-15", "Demokrasi ve Milli Birlik Günü", "full"),
    ("2025-08-30", "Zafer Bayramı", "full"),
    ("2025-10-28", "Cumhuriyet Bayramı Arifesi", "half"),
    ("2025-10-29", "Cumhuriyet Bayramı", "full"),
    # 2026
    ("2026-01-01", "Yılbaşı", "full"),
    ("2026-03-19", "Ramazan Bayramı Arifesi", "half"),
    ("2026-03-20", "Ramazan Bayramı 1. Gün", "full"),
    ("2026-03-21", "Ramazan Bayramı 2. Gün", "full"),
    ("2026-03-22", "Ramazan Bayramı 3. Gün", "full"),
    ("2026-04-23", "Ulusal Egemenlik ve Çocuk Bayramı", "full"),
    ("2026-05-01", "Emek ve Dayanışma Günü", "full"),
    ("2026-05-19", "Atatürk'ü Anma Gençlik ve Spor Bayramı", "full"),
    ("2026-05-26", "Kurban Bayramı Arifesi", "half"),
    ("2026-05-27", "Kurban Bayramı 1. Gün", "full"),
    ("2026-05-28", "Kurban Bayramı 2. Gün", "full"),
    ("2026-05-29", "Kurban Bayramı 3. Gün", "full"),
    ("2026-05-30", "Kurban Bayramı 4. Gün", "full"),
    ("2026-07-15", "Demokrasi ve Milli Birlik Günü", "full"),
    ("2026-08-30", "Zafer Bayramı", "full"),
    ("2026-10-28", "Cumhuriyet Bayramı Arifesi", "half"),
    ("2026-10-29", "Cumhuriyet Bayramı", "full"),
    # 2027
    ("2027-01-01", "Yılbaşı", "full"),
    ("2027-03-09", "Ramazan Bayramı Arifesi", "half"),
    ("2027-03-10", "Ramazan Bayramı 1. Gün", "full"),
    ("2027-03-11", "Ramazan Bayramı 2. Gün", "full"),
    ("2027-03-12", "Ramazan Bayramı 3. Gün", "full"),
    ("2027-04-23", "Ulusal Egemenlik ve Çocuk Bayramı", "full"),
    ("2027-05-01", "Emek ve Dayanışma Günü", "full"),
    ("2027-05-15", "Kurban Bayramı Arifesi", "half"),
    ("2027-05-16", "Kurban Bayramı 1. Gün", "full"),
    ("2027-05-17", "Kurban Bayramı 2. Gün", "full"),
    ("2027-05-18", "Kurban Bayramı 3. Gün", "full"),
    ("2027-05-19", "Kurban Bayramı 4. Gün", "full"),
    ("2027-07-15", "Demokrasi ve Milli Birlik Günü", "full"),
    ("2027-08-30", "Zafer Bayramı", "full"),
    ("2027-10-28", "Cumhuriyet Bayramı Arifesi", "half"),
    ("2027-10-29", "Cumhuriyet Bayramı", "full"),
]

async def get_all_holidays() -> dict:
    """Return {date_iso: {'name', 'type'}} merging static + custom + imported.
    'full' type wins over 'half' if multiple entries on same date."""
    m = {d: {"name": n, "type": t, "source": "system"} for d, n, t in STATIC_HOLIDAYS}
    async for h in db.holidays.find({}, {"_id": 0}):
        m[h["date"]] = {"name": h["name"], "type": h["type"], "source": "custom"}
    # Imported records (holidays_import) — aynı tarih için 'full' 'half'ı ezer
    async for h in db.holidays_import.find({"active": True}, {"_id": 0}):
        d = h["date"]
        cur = m.get(d)
        if not cur or (cur.get("type") == "half" and h.get("type") == "full"):
            m[d] = {"name": h["name"], "type": h["type"], "source": "imported"}
    return m

# -----------------------------------------------------------------------------
# Leave calc helpers
# -----------------------------------------------------------------------------
def entitlement_for(hire_date: date, prev_seniority_years: float, birth_date: Optional[date],
                    as_of: date) -> int:
    total_years = ((as_of - hire_date).days / 365.25) + (prev_seniority_years or 0)
    if total_years < 1:
        base = 0
    elif total_years < 5:
        base = 14
    elif total_years < 15:
        base = 20
    else:
        base = 26
    # 18 yaş altı ve 50+ minimum 20
    if birth_date:
        age = (as_of - birth_date).days / 365.25
        if (age < 18 or age >= 50) and base > 0:
            base = max(base, 20)
    return int(base)

def _fmt_tr_num(x) -> str:
    """20.5 → '20,5' ; 20.0 → '20' — Turkish decimal display."""
    try:
        n = float(x)
    except Exception:
        return str(x)
    if abs(n - round(n)) < 1e-9:
        return str(int(round(n)))
    return f"{n:.2f}".rstrip("0").rstrip(".").replace(".", ",")


def _next_working_day(after_end: date, holidays: dict) -> date:
    """Return first working day after `after_end` (skipping Sat/Sun and full-day holidays;
    half-day arife counts as a working day / return day)."""
    d = after_end + timedelta(days=1)
    while True:
        if d.weekday() >= 5:
            d += timedelta(days=1); continue
        h = holidays.get(d.isoformat())
        if h and h.get("type") != "half":
            d += timedelta(days=1); continue
        return d


TR_WEEKDAYS = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]


async def calc_leave_days(start: date, end: date) -> dict:
    """Return {'days', 'breakdown', 'return_date', 'return_weekday'}.
    Deducts weekends & Turkish holidays; arife = 0.5. return_date = ilk gerçek çalışma günü."""
    holidays = await get_all_holidays()
    days = 0.0
    breakdown = []
    d = start
    while d <= end:
        iso = d.isoformat()
        weekday = d.weekday()  # 0=Mon .. 5=Sat, 6=Sun
        h = holidays.get(iso)
        if weekday >= 5:
            breakdown.append({"date": iso, "value": 0, "reason": "Hafta sonu"})
        elif h:
            if h["type"] == "half":
                days += 0.5
                breakdown.append({"date": iso, "value": 0.5, "reason": f"Arife ({h['name']})"})
            else:
                breakdown.append({"date": iso, "value": 0, "reason": h["name"]})
        else:
            days += 1.0
            breakdown.append({"date": iso, "value": 1, "reason": "Çalışma günü"})
        d += timedelta(days=1)
    ret = _next_working_day(end, holidays)
    return {
        "days": days,
        "breakdown": breakdown,
        "return_date": ret.isoformat(),
        "return_weekday": TR_WEEKDAYS[ret.weekday()],
    }

async def allocate_leaves_fifo(entitlements: list, leaves: list,
                                 personnel: Optional[dict] = None) -> list:
    """FIFO: en eski hak edişten başlayarak her izin gününü tüketir. Aynı izin
    birden fazla hak edişe bölünebilir. Doğru gün hesabı ile birleşince,
    yalnızca gerçekten hak ediş bakiyesi tükenen izinler bölünür.

    Iter 29: avans izin (henüz hak ediş oluşmamışsa) hipotetik hak ediş kaydı
    üretilir — cetvelde YILLAR + İşyerindeki Kıdem hak ETSEYDİ olacak değerle
    doldurulsun diye. Hipotetik entitlement: `hire + (n years)` (leave.start_date
    <= o tarih) + `n + prev_years` toplam kıdem.
    """
    holidays = await get_all_holidays()
    ents = sorted(entitlements, key=lambda x: x["date"])
    state = [{"date": e["date"], "days": float(e["days"]),
              "seniority_at": e.get("total_seniority", 0), "used": 0.0} for e in ents]
    # Hipotetik hak ediş üretici — sadece avans dilimlerinde kullanılır.
    hire_d = _parse_date((personnel or {}).get("ise_giris")) if personnel else None
    prev_years = int((personnel or {}).get("onceki_kidem_yil") or 0)

    def _hypo_ent(leave_start: date) -> tuple:
        """(entitlement_date_iso, seniority_years). hire yoksa (None, prev+1)."""
        if not hire_d:
            return (None, prev_years + 1)
        n = 1
        while True:
            anniv = _safe_anniv(hire_d, n)
            if anniv >= leave_start:
                return (anniv.isoformat(), n + prev_years)
            n += 1
            if n > 100:
                return (None, prev_years + 1)
    allocations: list = []
    for L in sorted(leaves, key=lambda x: x["start_date"]):
        s = _parse_date(L["start_date"]); e_d = _parse_date(L["end_date"])
        if not s or not e_d:
            continue
        day_list = []
        cur = s
        while cur <= e_d:
            iso = cur.isoformat()
            h = holidays.get(iso)
            wd = cur.weekday()
            if wd >= 5:
                val = 0.0
            elif h:
                val = 0.5 if h.get("type") == "half" else 0.0
            else:
                val = 1.0
            day_list.append({"date": cur, "value": val})
            cur += timedelta(days=1)

        # Iter 26: İzin kaydında saklanan `days` (Excel/manuel giriş) AUTORİTEDİR.
        # Takvim hesabı farklı bir toplam veriyorsa, FIFO dağıtımı için day_list'i
        # kayıtta yazılı gün sayısına eşit olacak şekilde düzelt. Böylece hak ediş
        # bölünmesi (17+3 gibi) izin kaydındaki gerçek gün sayısına göre yapılır.
        target_days = L.get("days")
        try:
            target_days = float(target_days) if target_days is not None else None
        except Exception:
            target_days = None
        if target_days is not None:
            cal_total = sum(d["value"] for d in day_list)
            diff = round(target_days - cal_total, 4)
            if abs(diff) > 1e-6:
                if diff > 0:
                    # Eksik günü sıfır-değerli son güne ekle (çalışma yapılmış say)
                    added = False
                    for d in reversed(day_list):
                        if d["value"] == 0:
                            d["value"] = diff; added = True; break
                    if not added and day_list:
                        day_list[-1]["value"] += diff
                elif diff < 0:
                    # Fazla günü sondan indir (pozitif kaldığı sürece)
                    remaining = -diff
                    for d in reversed(day_list):
                        if remaining <= 1e-9: break
                        if d["value"] > 0:
                            take = min(d["value"], remaining)
                            d["value"] -= take; remaining -= take

        def first_available_idx():
            for i, r in enumerate(state):
                if (r["days"] - r["used"]) > 1e-9:
                    return i
            return None

        cur_idx = first_available_idx()
        slices: list = []
        cur_slice = None
        for day in day_list:
            v = day["value"]
            if v == 0:
                if cur_slice is not None:
                    cur_slice[3] = day["date"]
                continue
            remaining_v = v
            while remaining_v > 1e-9:
                if cur_idx is None:
                    if not cur_slice or cur_slice[0] != -1:
                        if cur_slice: slices.append(list(cur_slice))
                        cur_slice = [-1, 0.0, day["date"], day["date"]]
                    cur_slice[1] += remaining_v
                    cur_slice[3] = day["date"]
                    remaining_v = 0
                    break
                r = state[cur_idx]
                avail = r["days"] - r["used"]
                if avail <= 1e-9:
                    if cur_slice is not None:
                        slices.append(list(cur_slice)); cur_slice = None
                    cur_idx = first_available_idx()
                    continue
                take = min(remaining_v, avail)
                r["used"] += take
                remaining_v -= take
                if cur_slice is None or cur_slice[0] != cur_idx:
                    if cur_slice is not None:
                        slices.append(list(cur_slice))
                    cur_slice = [cur_idx, take, day["date"], day["date"]]
                else:
                    cur_slice[1] += take
                    cur_slice[3] = day["date"]
                if (r["days"] - r["used"]) <= 1e-9 and remaining_v > 1e-9:
                    slices.append(list(cur_slice)); cur_slice = None
                    cur_idx = first_available_idx()
        if cur_slice is not None:
            slices.append(cur_slice)

        if slices:
            slices[0][2] = s

        n = len(slices)
        for k, sl in enumerate(slices):
            ei, alloc_days, first_d, last_d = sl
            if ei == -1:
                # Iter 29: avans için hipotetik hak ediş üret
                hypo_date, hypo_sen = _hypo_ent(first_d)
                ent_date = hypo_date or ""
                try:
                    ent_year = int(ent_date[:4]) if ent_date else None
                except Exception:
                    ent_year = None
                sen_at = hypo_sen
            else:
                r = state[ei]
                ent_date = r["date"]
                try:
                    ent_year = int(ent_date[:4])
                except Exception:
                    ent_year = None
                sen_at = r["seniority_at"]
            if k < n - 1:
                ret_d = slices[k + 1][2].isoformat()
            else:
                ret_d = _next_working_day(e_d, holidays).isoformat()
            allocations.append({
                "leave_id": L["id"],
                "entitlement_date": ent_date,
                "entitlement_year": ent_year,
                "seniority_at": sen_at,
                "days": round(alloc_days, 2),
                "slice_start": first_d.isoformat(),
                "slice_end": last_d.isoformat(),
                "return_date": ret_d,
                "izin_turu": L.get("izin_turu", ""),
                "aciklama": L.get("aciklama", ""),
                "start_date": L["start_date"],
                "end_date": L["end_date"],
                "is_advance": ei == -1,
            })
    return allocations



# -----------------------------------------------------------------------------
# Models
# -----------------------------------------------------------------------------
Role = Literal["admin", "hr", "viewer"]

class LoginIn(BaseModel):
    email: EmailStr
    password: str

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Role

class UserOut(BaseModel):
    id: str
    email: str
    name: str
    role: Role

class Personnel(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sicil_no: str
    ad_soyad: str
    tc_no: Optional[str] = ""
    ise_giris: str  # ISO date
    isten_cikis: Optional[str] = None
    dogum_tarihi: Optional[str] = None
    departman: str = ""
    gorev: str = ""
    sirket: str = ""
    aktif: bool = True
    onceki_kidem_yil: int = 0  # sadece tam yıl
    telefon: str = ""
    email: Optional[str] = ""
    aciklama: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class PersonnelIn(BaseModel):
    sicil_no: str
    ad_soyad: str
    tc_no: Optional[str] = ""
    ise_giris: str
    isten_cikis: Optional[str] = None
    dogum_tarihi: Optional[str] = None
    departman: str = ""
    gorev: str = ""
    sirket: str = ""
    aktif: bool = True
    onceki_kidem_yil: int = 0
    telefon: str = ""
    email: Optional[str] = ""
    aciklama: str = ""

class LeaveRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    personnel_id: str
    start_date: str
    end_date: str
    days: float
    izin_turu: str = "Yıllık İzin"
    aciklama: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    created_by: Optional[str] = None

class LeaveIn(BaseModel):
    personnel_id: str
    start_date: str
    end_date: str
    izin_turu: str = "Yıllık İzin"
    aciklama: str = ""

# === Iter 54: ÖZEL İZİNLER (yıllık izin sisteminden TAM AYRI koleksiyon) ===
SPECIAL_LEAVE_TYPES = ["gebelik", "dogum", "sut_izni", "evlilik", "cenaze", "diger"]


class SpecialLeaveIn(BaseModel):
    personnel_id: str
    tur: str
    process_id: str = ""
    start_date: str = ""
    end_date: str = ""
    gun_sayisi: Optional[float] = None
    aciklama: str = ""
    durum: str = ""
    yakinlik: str = ""
    gebelik_teblig_tarihi: str = ""
    tahmini_dogum_tarihi: str = ""
    calisamaz_rapor_tarihi: str = ""
    calisamaz_rapor_bitis: str = ""
    dogum_tarihi_kayit: str = ""
    dogum_sonrasi_isbasi: str = ""
    sut_izni_bitis: str = ""
    cocuk_dogum_tarihi: str = ""
    ucretsiz_izin_baslangic: str = ""
    ucretsiz_izin_bitis: str = ""


@api.get("/special-leaves")
async def list_special_leaves(personnel_id: Optional[str] = None, tur: Optional[str] = None,
                                durum: Optional[str] = None, limit: int = 200, skip: int = 0,
                                _: dict = Depends(get_current_user)):
    q: dict = {"deleted": {"$ne": True}}
    if personnel_id: q["personnel_id"] = personnel_id
    if tur: q["tur"] = tur
    if durum: q["durum"] = durum
    total = await db.special_leaves.count_documents(q)
    items = await db.special_leaves.find(q, {"_id": 0}).sort([("start_date", -1), ("created_at", -1)]).skip(skip).limit(limit).to_list(None)
    pids = list({x["personnel_id"] for x in items if x.get("personnel_id")})
    pmap = {}
    if pids:
        async for p in db.personnel.find({"id": {"$in": pids}}, {"_id": 0, "id": 1, "sicil_no": 1, "ad_soyad": 1, "departman": 1, "aktif": 1}):
            pmap[p["id"]] = p
    # Ek dosya sayıları (batch)
    sids = [x["id"] for x in items if x.get("id")]
    att_counts: dict = {}
    if sids:
        pipeline = [
            {"$match": {"special_leave_id": {"$in": sids}, "is_deleted": False}},
            {"$group": {"_id": "$special_leave_id", "count": {"$sum": 1}}},
        ]
        async for row in db.special_leave_attachments.aggregate(pipeline):
            att_counts[row["_id"]] = row["count"]
    for it in items:
        p = pmap.get(it["personnel_id"], {})
        it["sicil_no"] = p.get("sicil_no")
        it["ad_soyad"] = p.get("ad_soyad")
        it["departman"] = p.get("departman")
        it["_personnel_active"] = p.get("aktif", True)
        it["attachment_count"] = att_counts.get(it.get("id"), 0)
    return {"total": total, "items": items}


@api.get("/special-leaves/personnel/{pid}/process")
async def personnel_process(pid: str, _: dict = Depends(get_current_user)):
    """Iter 68: Bir personelin tüm gebelik/doğum/süt izni kayıtlarını + durum + belge
    sayılarını tek response'ta döner. Süreç Detay ekranı bu endpoint'i kullanır."""
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    rows = await db.special_leaves.find(
        {"personnel_id": pid, "deleted": {"$ne": True}}, {"_id": 0}
    ).sort("created_at", -1).to_list(None)
    ids = [r["id"] for r in rows]
    att_counts: dict = {}
    if ids:
        async for a in db.special_leave_attachments.aggregate([
            {"$match": {"special_leave_id": {"$in": ids}, "is_deleted": False}},
            {"$group": {"_id": "$special_leave_id", "count": {"$sum": 1}}},
        ]):
            att_counts[a["_id"]] = a["count"]
    for r in rows:
        r["attachment_count"] = att_counts.get(r.get("id"), 0)
    return {"personnel": p, "records": rows}


@api.get("/special-leaves/status-panel")
async def status_panel(_: dict = Depends(get_current_user)):
    """Iter 67: Gebelik/doğum/süt izni yaşam döngüsü — tek deterministik durum motoru.
    Statü öncelik sırası (aynı personel yalnız 1 kartta):
      MILK_LEAVE — child_dob + isbasi ikisi de var AND today < child_dob+1yr
      MATERNITY  — (crt_start VAR ya da child_dob VAR) AND isbasi YOK
      PREGNANT   — teblig VAR AND crt_start YOK AND child_dob YOK
      COMPLETED  — yukarıdakilerin hiçbiri
    Sadece aktif personel dahil. Yıllık izin sistemi etkilenmez.
    """
    today = date.today()
    today_iso = today.isoformat()
    future_10 = (today + timedelta(days=10)).isoformat()

    def _days_between(a: str, b: str) -> Optional[int]:
        try:
            if not a or not b:
                return None
            da = datetime.strptime(a[:10], "%Y-%m-%d").date()
            db_ = datetime.strptime(b[:10], "%Y-%m-%d").date()
            return (db_ - da).days + 1  # dahil-dahil takvim günü
        except Exception:
            return None

    def _add_years(iso: str, yrs: int) -> str:
        try:
            d = datetime.strptime(iso[:10], "%Y-%m-%d").date()
            return date(d.year + yrs, d.month, d.day).isoformat()
        except Exception:
            return ""

    def _days_until(iso: str) -> Optional[int]:
        try:
            d = datetime.strptime(iso[:10], "%Y-%m-%d").date()
            return (d - today).days
        except Exception:
            return None

    active_pids: set = set()
    async for p in db.personnel.find({"aktif": True}, {"_id": 0, "id": 1}):
        active_pids.add(p["id"])

    by_pid: dict = {}
    async for it in db.special_leaves.find(
        {"deleted": {"$ne": True}, "tur": {"$in": ["gebelik", "dogum", "sut_izni"]}},
        {"_id": 0},
    ):
        pid = it.get("personnel_id")
        if not pid or pid not in active_pids:
            continue
        by_pid.setdefault(pid, []).append(it)

    def _merge(rows: list) -> dict:
        """Personelin gebelik+doğum+süt izni kayıtlarını tek pencerede birleştirir.
        Iter 73: process_id varsa yalnız EN SON process'in kayıtları merge edilir
        (aynı personelin eski/tamamlanmış süreçleri karışmaz).
        Iter 68: Eski kayıtlarda `gebelik_teblig_tarihi` boş olup yalnız `start_date`
        girilmiş olabilir — bu durumda `start_date` teblig olarak alınır."""
        # process_id'ler ve created_at'lerine bak; en son (max created_at) process'i seç
        rows_sorted = sorted(rows, key=lambda r: r.get("created_at", ""), reverse=True)
        pids_seen = [r.get("process_id") for r in rows_sorted if r.get("process_id")]
        active_pid = pids_seen[0] if pids_seen else None
        if active_pid:
            filtered = [r for r in rows_sorted if r.get("process_id") == active_pid]
        else:
            filtered = rows_sorted  # eski kayıtlar (process_id yok) — hepsini merge
        m = {"id": None}
        for r in filtered:
            for k in ("gebelik_teblig_tarihi", "tahmini_dogum_tarihi",
                      "calisamaz_rapor_tarihi", "calisamaz_rapor_bitis",
                      "cocuk_dogum_tarihi", "dogum_tarihi_kayit",
                      "dogum_sonrasi_isbasi", "sut_izni_bitis",
                      "ucretsiz_izin_baslangic", "ucretsiz_izin_bitis",
                      "start_date", "end_date", "durum", "aciklama",
                      "process_id"):
                if not m.get(k) and r.get(k):
                    m[k] = r[k]
            if not m["id"]:
                m["id"] = r.get("id")
            if r.get("tur") == "gebelik" and not m.get("gebelik_teblig_tarihi") and r.get("start_date"):
                m["gebelik_teblig_tarihi"] = r["start_date"]
        return m

    gebe: list = []
    dogum: list = []
    sut: list = []
    gebe_upcoming_10 = 0
    sut_ending_10 = 0

    for pid, rows in by_pid.items():
        m = _merge(rows)
        teblig = m.get("gebelik_teblig_tarihi") or ""
        crt = m.get("calisamaz_rapor_tarihi") or ""
        crt_end = m.get("calisamaz_rapor_bitis") or ""
        child_dob = m.get("cocuk_dogum_tarihi") or m.get("dogum_tarihi_kayit") or ""
        isbasi = m.get("dogum_sonrasi_isbasi") or ""
        ui_bas = m.get("ucretsiz_izin_baslangic") or ""
        ui_bit = m.get("ucretsiz_izin_bitis") or ""

        # Süt izni bitişi: child_dob varsa DAİMA otomatik (child_dob + 1 yıl).
        # Eski/hatalı persist edilmiş değerlere güvenilmez.
        sut_bitis = _add_years(child_dob, 1) if child_dob else (m.get("sut_izni_bitis") or "")

        # Süt izni bitişi bugünden önce mi?
        milk_expired = bool(sut_bitis and sut_bitis < today_iso)

        # DURUM MOTORU (Iter 70)
        # MILK_LEAVE koşulu: işbaşı geçmiş olmak ZORUNDA değil — yeter ki
        # işbaşı ≤ çocuk_dob+1yıl ve bugün henüz 1 yaşı geçmemiş olsun.
        status = None
        if child_dob and isbasi and not milk_expired:
            one_year = _add_years(child_dob, 1) if child_dob else ""
            if one_year and isbasi <= one_year and today_iso < one_year:
                status = "MILK_LEAVE"

        if not status:
            # MATERNITY: crt VEYA child_dob var, işbaşı henüz yapılmadı (veya işbaşı 1 yaş sonrası).
            # Iter 72 fix: child_dob + 1 yıl geçmişse süreç tamamlanmış demektir, MATERNITY'de kalma.
            one_year_for_maternity = _add_years(child_dob, 1) if child_dob else ""
            still_within_year = (not child_dob) or (one_year_for_maternity and today_iso < one_year_for_maternity)
            if (crt or child_dob) and (not isbasi or (
                one_year_for_maternity and isbasi > one_year_for_maternity
            )) and still_within_year:
                status = "MATERNITY"

        if not status:
            if teblig and not crt and not child_dob:
                status = "PREGNANT"

        if not status:
            continue  # gösterilmeye uygun aktif hiçbir kart yok

        # ORTAK ZENGİNLEŞTİRME
        item = {
            "id": m.get("id"),
            "personnel_id": pid,
            "gebelik_teblig_tarihi": teblig,
            "calisamaz_rapor_tarihi": crt,
            "calisamaz_rapor_bitis": crt_end,
            "cocuk_dogum_tarihi": child_dob,
            "ucretsiz_izin_baslangic": ui_bas,
            "ucretsiz_izin_bitis": ui_bit,
            "dogum_sonrasi_isbasi": isbasi,
            "sut_izni_bitis": sut_bitis,
            "durum": m.get("durum") or "",
        }
        if status == "PREGNANT":
            item["_date_hint"] = teblig
            item["_next_critical"] = {
                "label": "Çalışamaz Raporu (planlı)" if crt else "Tebliğ",
                "date": crt or teblig,
                "days_left": _days_until(crt) if crt else None,
            }
            if crt and today_iso < crt <= future_10:
                gebe_upcoming_10 += 1
            gebe.append(item)
        elif status == "MATERNITY":
            rapor_gun = _days_between(crt, crt_end) if (crt and crt_end) else None
            ui_gun = _days_between(ui_bas, ui_bit) if (ui_bas and ui_bit) else None
            item["rapor_gun_sayisi"] = rapor_gun
            item["ucretsiz_izin_gun_sayisi"] = ui_gun
            item["toplam_uzak_gun"] = (rapor_gun or 0) + (ui_gun or 0)
            item["_date_hint"] = crt or child_dob
            # Yaklaşan kritik tarih: rapor bitiş > ücretsiz izin bitiş > işbaşı
            using_ui = ui_bas and ui_bit and ui_bas <= today_iso <= ui_bit
            if using_ui:
                item["durum"] = item["durum"] or "Ücretsiz İzin Kullanıyor"
                item["_next_critical"] = {
                    "label": "Ücretsiz İzin Bitişi",
                    "date": ui_bit,
                    "days_left": _days_until(ui_bit),
                }
            elif crt_end and today_iso <= crt_end:
                item["_next_critical"] = {
                    "label": "Rapor Bitişi",
                    "date": crt_end,
                    "days_left": _days_until(crt_end),
                }
            elif isbasi:
                item["_next_critical"] = {
                    "label": "İşbaşı",
                    "date": isbasi,
                    "days_left": _days_until(isbasi),
                }
            else:
                item["_next_critical"] = {"label": "İşbaşı", "date": "", "days_left": None}
            dogum.append(item)
        elif status == "MILK_LEAVE":
            item["_date_hint"] = sut_bitis
            item["kalan_gun"] = _days_until(sut_bitis)
            # Iter 70: Toplam takip süresi = işbaşı → çocuk 1 yaş
            item["toplam_takip_gun"] = _days_between(isbasi, sut_bitis) if (isbasi and sut_bitis) else None
            item["_next_critical"] = {
                "label": "Süt İzni Bitişi",
                "date": sut_bitis,
                "days_left": item["kalan_gun"],
            }
            item["takip_baslangic"] = isbasi
            if sut_bitis and today_iso <= sut_bitis <= future_10:
                sut_ending_10 += 1
            sut.append(item)

    # Personel bilgisi ekle
    all_pids = list({x["personnel_id"] for x in (gebe + dogum + sut)})
    pmap: dict = {}
    if all_pids:
        async for p in db.personnel.find(
            {"id": {"$in": all_pids}},
            {"_id": 0, "id": 1, "sicil_no": 1, "ad_soyad": 1, "departman": 1},
        ):
            pmap[p["id"]] = p
    for lst in (gebe, dogum, sut):
        for x in lst:
            p = pmap.get(x["personnel_id"], {})
            x["sicil_no"] = p.get("sicil_no")
            x["ad_soyad"] = p.get("ad_soyad")
            x["departman"] = p.get("departman")

    # Sırala (kritik tarihe göre)
    def _sort_key(x): return (x.get("_next_critical") or {}).get("date") or x.get("_date_hint") or "9999-12-31"
    gebe.sort(key=_sort_key)
    dogum.sort(key=_sort_key)
    sut.sort(key=_sort_key)

    # Kart altı bilgi: en yakın kritik tarih
    def _card_summary(lst):
        if not lst:
            return None
        head = lst[0]
        nc = head.get("_next_critical") or {}
        return {"date": nc.get("date"), "label": nc.get("label"),
                "days_left": nc.get("days_left")}

    return {
        "gebe_calisan": {
            "count": len(gebe), "items": gebe,
            "upcoming_report_10d": gebe_upcoming_10,
            "next_critical": _card_summary(gebe),
        },
        "dogum_izninde": {
            "count": len(dogum), "items": dogum,
            "next_critical": _card_summary(dogum),
        },
        "sut_izni_kullanan": {
            "count": len(sut), "items": sut,
            "ending_soon_10d": sut_ending_10,
            "next_critical": _card_summary(sut),
        },
    }


@api.get("/special-leaves/gebelik-alerts")
async def gebelik_alerts(threshold_days: int = 10,
                          _: dict = Depends(get_current_user)):
    """Iter 62: Çalışamaz Raporu (D-8h) tarihi yaklaşan gebelik/doğum kayıtları.
    - upcoming: calisamaz_rapor_tarihi bugün ile bugün+threshold_days arasında
    """
    today = date.today()
    today_iso = today.isoformat()
    future_cutoff = (today + timedelta(days=threshold_days)).isoformat()
    upcoming: list = []
    async for it in db.special_leaves.find(
        {"deleted": {"$ne": True}, "tur": {"$in": ["gebelik", "dogum"]}},
        {"_id": 0},
    ):
        d = it.get("calisamaz_rapor_tarihi") or ""
        if not d:
            continue
        if today_iso <= d <= future_cutoff:
            upcoming.append({**it, "_alert_date": d})
    pids = list({x["personnel_id"] for x in upcoming if x.get("personnel_id")})
    pmap: dict = {}
    if pids:
        async for p in db.personnel.find(
            {"id": {"$in": pids}},
            {"_id": 0, "id": 1, "sicil_no": 1, "ad_soyad": 1, "departman": 1, "aktif": 1},
        ):
            pmap[p["id"]] = p
    for x in upcoming:
        p = pmap.get(x["personnel_id"], {})
        x["sicil_no"] = p.get("sicil_no")
        x["ad_soyad"] = p.get("ad_soyad")
        x["departman"] = p.get("departman")
        x["_personnel_active"] = p.get("aktif", True)
    upcoming.sort(key=lambda x: x["_alert_date"])
    return {"threshold_days": threshold_days, "upcoming": upcoming}


@api.get("/special-leaves/milk-alerts")
async def milk_alerts(threshold_days: int = 10,
                        _: dict = Depends(get_current_user)):
    """Iter 59: Süt İzni yaklaşan / biten personeller.
    - yaklaşan: end_date bugünden itibaren threshold_days içinde (dahil)
    - biten: end_date son threshold_days içinde geçmiş
    """
    today = date.today()
    today_iso = today.isoformat()
    future_cutoff = (today + timedelta(days=threshold_days)).isoformat()
    past_cutoff = (today - timedelta(days=threshold_days)).isoformat()
    upcoming: list = []
    ended: list = []
    async for it in db.special_leaves.find(
        {"deleted": {"$ne": True}, "tur": "sut_izni"}, {"_id": 0}
    ):
        end_val = it.get("end_date") or it.get("sut_izni_bitis") or ""
        if not end_val:
            continue
        entry = {**it, "_end": end_val}
        if today_iso <= end_val <= future_cutoff:
            upcoming.append(entry)
        elif past_cutoff <= end_val < today_iso:
            ended.append(entry)
    pids = list({x["personnel_id"] for x in upcoming + ended if x.get("personnel_id")})
    pmap: dict = {}
    if pids:
        async for p in db.personnel.find(
            {"id": {"$in": pids}},
            {"_id": 0, "id": 1, "sicil_no": 1, "ad_soyad": 1, "departman": 1, "aktif": 1},
        ):
            pmap[p["id"]] = p
    for lst in (upcoming, ended):
        for x in lst:
            p = pmap.get(x["personnel_id"], {})
            x["sicil_no"] = p.get("sicil_no")
            x["ad_soyad"] = p.get("ad_soyad")
            x["departman"] = p.get("departman")
            x["_personnel_active"] = p.get("aktif", True)
    upcoming.sort(key=lambda x: x["_end"])
    ended.sort(key=lambda x: x["_end"], reverse=True)
    return {
        "threshold_days": threshold_days,
        "upcoming": upcoming,
        "ended": ended,
    }


@api.post("/special-leaves")
async def create_special_leave(body: SpecialLeaveIn, request: Request,
                                 user: dict = Depends(require_roles("admin", "hr"))):
    # Iter 69: Yeni kayıt yalnız birleşik süreç türü + evlilik/cenaze/diger kabul eder.
    ALLOWED_NEW_TYPES = {"gebelik", "evlilik", "cenaze", "diger"}
    if body.tur not in ALLOWED_NEW_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Yeni kayıtta izin türü '{body.tur}' artık desteklenmiyor. "
                   f"Gebelik/Doğum/Süt süreci için 'Gebelik, Doğum ve Süt İzni Takibi' seçin.",
        )
    if body.tur not in SPECIAL_LEAVE_TYPES:
        raise HTTPException(status_code=400, detail=f"Geçersiz tür")
    p = await db.personnel.find_one({"id": body.personnel_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    rec = body.model_dump()
    rec["id"] = str(uuid.uuid4())
    rec["created_at"] = datetime.now(timezone.utc).isoformat()
    rec["created_by"] = user.get("id")
    rec["deleted"] = False
    await db.special_leaves.insert_one(rec)
    await _audit(action="create", module="special_leave", entity_type="special_leave",
                 entity_id=rec["id"], entity_name=f"{p.get('ad_soyad')} — {body.tur}",
                 new_values=rec,
                 description=f"Özel izin eklendi: {p.get('ad_soyad')} — {body.tur} ({body.start_date} → {body.end_date})",
                 request=request, user=user)
    return {"ok": True, "id": rec["id"]}


@api.put("/special-leaves/{sid}")
async def update_special_leave(sid: str, body: SpecialLeaveIn, request: Request,
                                 user: dict = Depends(require_roles("admin", "hr"))):
    existing = await db.special_leaves.find_one({"id": sid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    upd = body.model_dump()
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.special_leaves.update_one({"id": sid}, {"$set": upd})
    await _audit(action="update", module="special_leave", entity_type="special_leave",
                 entity_id=sid, entity_name=body.tur,
                 old_values=existing, new_values=upd,
                 description="Özel izin güncellendi", request=request, user=user)
    return {"ok": True}


@api.delete("/special-leaves/{sid}")
async def delete_special_leave(sid: str, reason: str = "", request: Request = None,
                                 user: dict = Depends(require_roles("admin", "hr"))):
    existing = await db.special_leaves.find_one({"id": sid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    await db.special_leaves.update_one({"id": sid},
        {"$set": {"deleted": True, "deleted_at": datetime.now(timezone.utc).isoformat(),
                  "deleted_by": user.get("id"), "delete_reason": reason or ""}})
    await _audit(action="delete", module="special_leave", entity_type="special_leave",
                 entity_id=sid, entity_name=existing.get("tur"),
                 old_values=existing, new_values={"deleted": True, "reason": reason},
                 description=f"Özel izin (soft) silindi — gerekçe: {reason[:200] if reason else '—'}",
                 request=request, user=user)
    return {"ok": True}
# === Iter 54 sonu ===


# === Iter 59: Özel İzin Belge Yükleme (Emergent Object Storage) ===
ALLOWED_ATTACHMENT_MIME = {"application/pdf", "image/jpeg", "image/png"}
MAX_ATTACHMENT_SIZE = 10 * 1024 * 1024  # 10 MB


@api.post("/special-leaves/{sid}/attachments")
async def upload_special_leave_attachments(
    sid: str, request: Request,
    files: List[UploadFile] = File(...),
    user: dict = Depends(require_roles("admin", "hr")),
):
    parent = await db.special_leaves.find_one(
        {"id": sid, "deleted": {"$ne": True}}, {"_id": 0}
    )
    if not parent:
        raise HTTPException(status_code=404, detail="Özel izin kaydı bulunamadı")
    results = []
    for f in files:
        ct = (f.content_type or "").lower()
        if ct == "image/jpg":
            ct = "image/jpeg"
        if ct not in ALLOWED_ATTACHMENT_MIME:
            raise HTTPException(
                status_code=400,
                detail=f"Kabul edilmeyen dosya türü: {f.filename} ({ct}) — PDF/JPG/PNG yükleyin",
            )
        data = await f.read()
        if len(data) > MAX_ATTACHMENT_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"Dosya çok büyük: {f.filename} ({len(data)/(1024*1024):.1f} MB) — Maks 10 MB",
            )
        ext = (f.filename.rsplit(".", 1)[-1] if "." in (f.filename or "") else "bin").lower()
        aid = str(uuid.uuid4())
        path = f"{_APP_NAME}/special_leaves/{sid}/{aid}.{ext}"
        stored = await _put_object(path, data, ct)
        rec = {
            "id": aid,
            "special_leave_id": sid,
            "storage_path": stored.get("path", path),
            "original_filename": f.filename or f"belge.{ext}",
            "content_type": ct,
            "size": stored.get("size", len(data)),
            "is_deleted": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user.get("id"),
        }
        await db.special_leave_attachments.insert_one(rec)
        results.append({k: v for k, v in rec.items() if k != "_id"})
    await _audit(
        action="upload", module="special_leave_attachment",
        entity_type="special_leave_attachment", entity_id=sid,
        entity_name=f"{parent.get('tur')} — {len(results)} dosya",
        description=f"Özel izin belgesi yüklendi: {len(results)} dosya",
        request=request, user=user,
    )
    return {"ok": True, "attachments": results}


@api.get("/special-leaves/{sid}/attachments")
async def list_special_leave_attachments(
    sid: str, _: dict = Depends(get_current_user)
):
    items = await db.special_leave_attachments.find(
        {"special_leave_id": sid, "is_deleted": False}, {"_id": 0}
    ).sort("created_at", -1).to_list(None)
    return {"items": items}


@api.get("/special-leaves/{sid}/attachments/{aid}/download")
async def download_special_leave_attachment(
    sid: str, aid: str, _: dict = Depends(get_current_user)
):
    rec = await db.special_leave_attachments.find_one(
        {"id": aid, "special_leave_id": sid, "is_deleted": False}, {"_id": 0}
    )
    if not rec:
        raise HTTPException(status_code=404, detail="Belge bulunamadı")
    data, ct = await _get_object(rec["storage_path"])
    fname = rec.get("original_filename") or "belge"
    return Response(
        content=data,
        media_type=rec.get("content_type", ct),
        headers={"Content-Disposition": f'inline; filename="{fname}"'},
    )


@api.delete("/special-leaves/{sid}/attachments/{aid}")
async def delete_special_leave_attachment(
    sid: str, aid: str, request: Request,
    user: dict = Depends(require_roles("admin", "hr")),
):
    rec = await db.special_leave_attachments.find_one(
        {"id": aid, "special_leave_id": sid, "is_deleted": False}, {"_id": 0}
    )
    if not rec:
        raise HTTPException(status_code=404, detail="Belge bulunamadı")
    await db.special_leave_attachments.update_one(
        {"id": aid},
        {"$set": {"is_deleted": True,
                  "deleted_at": datetime.now(timezone.utc).isoformat(),
                  "deleted_by": user.get("id")}},
    )
    await _audit(
        action="delete", module="special_leave_attachment",
        entity_type="special_leave_attachment", entity_id=aid,
        entity_name=rec.get("original_filename"),
        description=f"Özel izin belgesi silindi: {rec.get('original_filename')}",
        request=request, user=user,
    )
    return {"ok": True}
# === Iter 59 sonu ===


# === Iter 73: Muvafakatname Takip (Pilot) ===
def _is_annual_type(t: Optional[str]) -> bool:
    tt = (t or "").strip().lower()
    return (not tt) or ("yıl" in tt) or ("yil" in tt) or (tt == "annual")


async def _consent_items(pilot: bool = False) -> list:
    """Muvafakatname gereken (avans/hak edişsiz kullanılmış) yıllık izinleri döner."""
    p_map: dict = {}
    async for p in db.personnel.find({}, {"_id": 0}):
        p_map[p["id"]] = p
    ent_by_pid: dict = {}
    async for e in db.entitlements.find({}, {"_id": 0}):
        ent_by_pid.setdefault(e["personnel_id"], []).append(e)
    for arr in ent_by_pid.values():
        arr.sort(key=lambda x: x["date"])
    by_pid: dict = {}
    async for L in db.leaves.find({}, {"_id": 0}):
        by_pid.setdefault(L["personnel_id"], []).append(L)
    items = []
    for pid, leaves in by_pid.items():
        p = p_map.get(pid)
        if not p:
            continue
        ents = ent_by_pid.get(pid, [])
        annual = [L for L in leaves if _is_annual_type(L.get("izin_turu"))]
        annual.sort(key=lambda x: x.get("start_date") or "")
        cum_used = 0.0
        current_year = date.today().year
        for L in annual:
            sd = L.get("start_date") or ""
            days = float(L.get("days") or 0)
            entitled_so_far = sum(float(e.get("days") or 0) for e in ents if e.get("date", "") <= sd)
            over = (cum_used + days) - entitled_so_far
            zero_day_no_balance = (days == 0) and ((entitled_so_far - cum_used) <= 0)

            # Muvafakatname sadece mevcut takvim yilindaki izinler icin istenir.
            # Eski izinler kullanilan izin / bakiye hesabinda kalmaya devam eder.
            try:
                leave_year = date.fromisoformat(sd[:10]).year
            except Exception:
                leave_year = None
            is_current_year_leave = leave_year == current_year

            advance_days = (
                round(min(days, max(0.0, over)), 2)
                if is_current_year_leave
                else 0
            )

            if is_current_year_leave and ((over > 0) or zero_day_no_balance):
                # Bir sonraki hak ediş tarihi (izinden sonra)
                next_ent = next((e.get("date") for e in ents if e.get("date", "") > sd), None)
                items.append({
                    "id": L["id"],
                    "personnel_id": pid,
                    "sicil_no": p.get("sicil_no"),
                    "ad_soyad": p.get("ad_soyad"),
                    "departman": p.get("departman"),
                    "tc_no": p.get("tc_no"),
                    "ise_giris": p.get("ise_giris"),
                    "start_date": sd,
                    "end_date": L.get("end_date"),
                    "days": days,
                    "entitled_so_far": round(entitled_so_far, 2),
                    "used_before": round(cum_used, 2),
                    "advance_days": advance_days,
                    "next_entitlement_date": next_ent,
                    "consent_status": L.get("consent_status") or "pending",
                    "consent_printed_at": L.get("consent_printed_at"),
                    "consent_signed_at": L.get("consent_signed_at"),
                    "attachment_count": 0,
                    "created_at": L.get("created_at"),
                    "izin_turu": L.get("izin_turu"),
                    "personnel_active": p.get("aktif", True),
                })
            cum_used += days
    lids = [it["id"] for it in items]
    if lids:
        counts: dict = {}
        async for r in db.leave_consent_attachments.aggregate([
            {"$match": {"leave_id": {"$in": lids}, "is_deleted": False}},
            {"$group": {"_id": "$leave_id", "count": {"$sum": 1}}},
        ]):
            counts[r["_id"]] = r["count"]
        for it in items:
            it["attachment_count"] = counts.get(it["id"], 0)
    items.sort(key=lambda x: x.get("start_date") or "", reverse=True)
    if pilot:
        items = items[:1]
    return items


@api.get("/personnel/consent-tracking")
async def personnel_consent_tracking(
    q: Optional[str] = None,
    departman: Optional[str] = None,
    sirket: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    view: str = "latest",  # latest | all
    ids_only: bool = False,
    limit: int = 500,
    skip: int = 0,
    _: dict = Depends(get_current_user),
):
    """Muvafakatname listesi.
    - Aktif personel içinden `remaining < 0` olanların avans yıllık izin kayıtlarını döner.
    - view="latest" (varsayılan): her personel için yalnız EN SON avans izin kaydı.
    - view="all": tüm avans izin kayıtları (kişi başına birden fazla satır olabilir).
    - Belge şablonu / hesaplama mantığı DEĞİŞMEZ — bu endpoint yalnızca hangi izin kaydının
      mevcut muvafakatname şablonuna gideceğini belirler.
    """
    # 1) Aktif personel + remaining < 0
    neg_map: dict = {}
    async for p in db.personnel.find({"aktif": {"$ne": False}}, {"_id": 0}):
        try:
            bal = await _compute_entitlements(p)
            rem = float(bal.get("remaining", 0) or 0)
        except Exception:
            continue
        if rem < 0:
            neg_map[p["id"]] = round(rem, 2)

    # 2) Mevcut avans izin satırları (izin-bazlı) — sadece eksi bakiyeli personeller
    all_rows = await _consent_items(pilot=False)
    rows = [r for r in all_rows if r.get("personnel_id") in neg_map]
    for r in rows:
        r["remaining"] = neg_map.get(r["personnel_id"], 0)

    # 3) Filtreler
    if q:
        qq = _tr_lower(q.strip())
        rows = [r for r in rows if qq in _tr_lower(r.get("ad_soyad") or "")
                or qq in _tr_lower(r.get("sicil_no") or "")]
    if departman:
        rows = [r for r in rows if (r.get("departman") or "") == departman]
    if sirket:
        # sirket bilgisi _consent_items içinde yok — personnel'den doldur
        p_sirket: dict = {}
        async for p in db.personnel.find({"id": {"$in": list({r["personnel_id"] for r in rows})}},
                                          {"_id": 0, "id": 1, "sirket": 1}):
            p_sirket[p["id"]] = p.get("sirket")
        for r in rows:
            r["sirket"] = p_sirket.get(r["personnel_id"])
        rows = [r for r in rows if (r.get("sirket") or "") == sirket]
    else:
        # sirket alanını yine de ekle (frontend için)
        p_sirket: dict = {}
        async for p in db.personnel.find({"id": {"$in": list({r["personnel_id"] for r in rows})}},
                                          {"_id": 0, "id": 1, "sirket": 1}):
            p_sirket[p["id"]] = p.get("sirket")
        for r in rows:
            r["sirket"] = p_sirket.get(r["personnel_id"])
    if start:
        rows = [r for r in rows if (r.get("start_date") or "") >= start]
    if end:
        rows = [r for r in rows if (r.get("start_date") or "") <= end]

    # 4) Sıralama: remaining ASC (en negatif en üstte) — start_date DESC ikincil sıralama
    rows.sort(key=lambda x: (float(x.get("remaining") or 0), -(int((x.get("start_date") or "0000-00-00").replace("-", "")))))

    # 5) view = latest: her personel için sadece en yeni satır
    if view == "latest":
        seen = set()
        latest = []
        for r in rows:
            pid = r.get("personnel_id")
            if pid in seen:
                continue
            seen.add(pid)
            latest.append(r)
        rows = latest

    total = len(rows)
    persons = len({r.get("personnel_id") for r in rows})
    if ids_only:
        return {"ids": [r["id"] for r in rows], "total": total, "persons": persons}
    limit = max(1, min(int(limit), 5000))
    skip = max(0, int(skip))
    return {"items": rows[skip: skip + limit], "total": total, "persons": persons}


@api.get("/personnel/consent-batch")
async def personnel_consent_batch(pids: str, _: dict = Depends(get_current_user)):
    """Toplu muvafakatname ön izleme için birden fazla personelin verisini döner."""
    ids = [x.strip() for x in (pids or "").split(",") if x.strip()][:500]
    if not ids:
        return {"items": []}
    p_map: dict = {}
    async for p in db.personnel.find({"id": {"$in": ids}}, {"_id": 0}):
        p_map[p["id"]] = p
    last_map: dict = {}
    async for L in db.leaves.find(
        {"personnel_id": {"$in": ids},
         "izin_turu": {"$regex": "yıl|yil|annual", "$options": "i"}},
        {"_id": 0},
    ).sort("start_date", -1):
        pid = L.get("personnel_id")
        if pid and pid not in last_map:
            last_map[pid] = L
    out = []
    for pid in ids:
        p = p_map.get(pid)
        if not p:
            continue
        try:
            bal = await _compute_entitlements(p)
            rem = float(bal.get("remaining", 0) or 0)
        except Exception:
            continue
        out.append({
            "personnel": p,
            "remaining": round(rem, 2),
            "muvafakat_days": round(abs(rem), 2),
            "last_leave": last_map.get(pid),
            "next_entitlement_date": (bal.get("next_entitlement") or {}).get("date"),
        })
    return {"items": out}


@api.get("/leaves/consent-tracking")
async def consent_tracking(
    q: Optional[str] = None,
    departman: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    year: Optional[int] = None,
    status: Optional[str] = None,
    ids_only: bool = False,
    pilot: bool = False,
    limit: int = 100,
    skip: int = 0,
    _: dict = Depends(get_current_user),
):
    """Muvafakatname takip listesi.
    - Filtreler: q (ad soyad / sicil), departman, start–end (izin start_date bazlı),
      year (opsiyonel, tarih aralığı yoksa), status (pending/printed/signed_uploaded/all).
    - pilot=True → tek örnek kayıt (eski davranış).
    - ids_only=True → yalnızca id listesi (toplu seç için, pagination'dan bağımsız).
    """
    items = await _consent_items(pilot=False)
    if q:
        qq = _tr_lower(q.strip())
        items = [it for it in items if
                 qq in _tr_lower(it.get("ad_soyad") or "") or
                 qq in _tr_lower(it.get("sicil_no") or "")]
    if departman:
        items = [it for it in items if (it.get("departman") or "") == departman]
    if start:
        items = [it for it in items if (it.get("start_date") or "") >= start]
    if end:
        items = [it for it in items if (it.get("start_date") or "") <= end]
    if year and not (start or end):
        y_s = f"{int(year):04d}-01-01"
        y_e = f"{int(year):04d}-12-31"
        items = [it for it in items if y_s <= (it.get("start_date") or "") <= y_e]
    if status and status not in ("all", ""):
        items = [it for it in items if (it.get("consent_status") or "pending") == status]
    total = len(items)
    if pilot:
        return {"items": items[:1], "total": total}
    if ids_only:
        return {"ids": [it["id"] for it in items], "total": total}
    limit = max(1, min(int(limit), 5000))
    skip = max(0, int(skip))
    return {"items": items[skip: skip + limit], "total": total}


@api.get("/leaves/consent-batch")
async def consent_batch(ids: str, _: dict = Depends(get_current_user)):
    """Toplu muvafakatname ön izleme için birden fazla izin verisini tek sorguda döner."""
    lids = [x.strip() for x in (ids or "").split(",") if x.strip()]
    if not lids:
        return {"items": []}
    lids = lids[:500]  # güvenlik sınırı
    out = []
    async for L in db.leaves.find({"id": {"$in": lids}}, {"_id": 0}):
        p = await db.personnel.find_one({"id": L["personnel_id"]}, {"_id": 0})
        bal = await _compute_entitlements(p) if p else {"remaining": 0, "next_entitlement": None}
        out.append({"leave": L, "personnel": p, "balance": bal})
    # istenen sırayla döndür
    ord_map = {lid: i for i, lid in enumerate(lids)}
    out.sort(key=lambda x: ord_map.get(x["leave"]["id"], 999999))
    return {"items": out}


@api.post("/leaves/{lid}/consent/mark-printed")
async def mark_consent_printed(lid: str, request: Request,
                                 user: dict = Depends(require_roles("admin", "hr"))):
    L = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not L:
        raise HTTPException(status_code=404, detail="İzin kaydı bulunamadı")
    now_iso = datetime.now(timezone.utc).isoformat()
    # İmzalı yüklendiyse status'u geriye almayız; sadece printed_at güncellenir
    cur = L.get("consent_status") or "pending"
    new_status = cur if cur == "signed_uploaded" else "printed"
    await db.leaves.update_one(
        {"id": lid},
        {"$set": {"consent_status": new_status,
                  "consent_printed_at": now_iso,
                  "consent_printed_by": user.get("id")}},
    )
    await _audit(
        action="print", module="leave_consent",
        entity_type="leave_consent", entity_id=lid,
        entity_name=str(L.get("personnel_id")),
        description="Muvafakatname yazdırıldı olarak işaretlendi",
        request=request, user=user,
    )
    return {"ok": True, "consent_status": new_status, "consent_printed_at": now_iso}


@api.post("/leaves/{lid}/consent/attachments")
async def upload_consent_attachments(
    lid: str, request: Request,
    files: List[UploadFile] = File(...),
    user: dict = Depends(require_roles("admin", "hr")),
):
    L = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not L:
        raise HTTPException(status_code=404, detail="İzin kaydı bulunamadı")
    results = []
    for f in files:
        ct = (f.content_type or "").lower()
        if ct == "image/jpg":
            ct = "image/jpeg"
        if ct not in ALLOWED_ATTACHMENT_MIME:
            raise HTTPException(status_code=400,
                detail=f"Kabul edilmeyen dosya türü: {f.filename} ({ct}) — PDF/JPG/PNG yükleyin")
        data = await f.read()
        if len(data) > MAX_ATTACHMENT_SIZE:
            raise HTTPException(status_code=400,
                detail=f"Dosya çok büyük: {f.filename} ({len(data)/(1024*1024):.1f} MB) — Maks 10 MB")
        ext = (f.filename.rsplit(".", 1)[-1] if "." in (f.filename or "") else "bin").lower()
        aid = str(uuid.uuid4())
        path = f"{_APP_NAME}/leave_consents/{lid}/{aid}.{ext}"
        stored = await _put_object(path, data, ct)
        rec = {
            "id": aid,
            "leave_id": lid,
            "personnel_id": L["personnel_id"],
            "storage_path": stored.get("path", path),
            "original_filename": f.filename or f"muvafakatname.{ext}",
            "content_type": ct,
            "size": stored.get("size", len(data)),
            "is_deleted": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user.get("id"),
        }
        await db.leave_consent_attachments.insert_one(rec)
        results.append({k: v for k, v in rec.items() if k != "_id"})
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.leaves.update_one(
        {"id": lid},
        {"$set": {"consent_status": "signed_uploaded",
                  "consent_signed_at": now_iso,
                  "consent_signed_by": user.get("id")}},
    )
    await _audit(
        action="upload", module="leave_consent_attachment",
        entity_type="leave_consent_attachment", entity_id=lid,
        entity_name=f"{len(results)} dosya",
        description=f"İmzalı muvafakatname yüklendi: {len(results)} dosya",
        request=request, user=user,
    )
    return {"ok": True, "attachments": results, "consent_status": "signed_uploaded"}


@api.get("/leaves/{lid}/consent/attachments")
async def list_consent_attachments(lid: str, _: dict = Depends(get_current_user)):
    items = await db.leave_consent_attachments.find(
        {"leave_id": lid, "is_deleted": False}, {"_id": 0}
    ).sort("created_at", -1).to_list(None)
    return {"items": items}


@api.get("/leaves/{lid}/consent/attachments/{aid}/download")
async def download_consent_attachment(lid: str, aid: str, _: dict = Depends(get_current_user)):
    rec = await db.leave_consent_attachments.find_one(
        {"id": aid, "leave_id": lid, "is_deleted": False}, {"_id": 0}
    )
    if not rec:
        raise HTTPException(status_code=404, detail="Belge bulunamadı")
    data, ct = await _get_object(rec["storage_path"])
    fname = rec.get("original_filename") or "muvafakatname"
    return Response(
        content=data,
        media_type=rec.get("content_type", ct),
        headers={"Content-Disposition": f'inline; filename="{fname}"'},
    )
# === Iter 73 sonu ===


class HolidayIn(BaseModel):
    date: str
    name: str
    type: Literal["full", "half"] = "full"

# -----------------------------------------------------------------------------
# Audit Log — models + helpers
# -----------------------------------------------------------------------------
class AuditLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: Optional[str] = None
    user_name: Optional[str] = None
    user_role: Optional[str] = None
    action: str
    module: str
    entity_type: Optional[str] = None
    entity_id: Optional[str] = None
    entity_name: Optional[str] = None
    old_values: Optional[dict] = None
    new_values: Optional[dict] = None
    description: str = ""
    ip_address: Optional[str] = None
    device_name: Optional[str] = None
    client_type: Optional[str] = None
    success: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UserExtendedIn(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Role
    username: Optional[str] = ""
    departman: Optional[str] = ""
    aktif: bool = True
    aciklama: Optional[str] = ""

class UserUpdateIn(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    username: Optional[str] = None
    role: Optional[Role] = None
    departman: Optional[str] = None
    aktif: Optional[bool] = None
    aciklama: Optional[str] = None

class ResetPasswordIn(BaseModel):
    new_password: str

SENSITIVE_KEYS = ("password", "password_hash", "token", "access_token", "new_password", "secret", "authorization", "jwt")

def _scrub_audit(node):
    if isinstance(node, list):
        return [_scrub_audit(x) for x in node]
    if not isinstance(node, dict):
        return node
    out = {}
    for k, v in node.items():
        if any(s in str(k).lower() for s in SENSITIVE_KEYS):
            out[k] = "***"
        else:
            out[k] = _scrub_audit(v)
    return out

def _client_info(request: Optional[Request]) -> dict:
    if not request:
        return {"ip_address": None, "device_name": None, "client_type": None}
    xff = request.headers.get("x-forwarded-for")
    ip = (xff.split(",")[0].strip() if xff else None) or (request.client.host if request.client else None)
    ua = request.headers.get("user-agent", "") or ""
    ual = ua.lower()
    device = "Diğer"
    if "windows" in ual: device = "Windows"
    elif "mac" in ual: device = "macOS"
    elif "android" in ual: device = "Android"
    elif "iphone" in ual or "ipad" in ual: device = "iOS"
    elif "linux" in ual: device = "Linux"
    return {"ip_address": ip, "device_name": device, "client_type": ua[:220]}

async def _audit(*, action: str, module: str,
                 entity_type: Optional[str] = None, entity_id: Optional[str] = None, entity_name: Optional[str] = None,
                 old_values: Optional[dict] = None, new_values: Optional[dict] = None,
                 description: str = "", request: Optional[Request] = None,
                 user: Optional[dict] = None, user_id: Optional[str] = None, user_name: Optional[str] = None,
                 user_role: Optional[str] = None, success: bool = True) -> dict:
    info = _client_info(request)
    doc = AuditLog(
        user_id=(user["id"] if user else user_id),
        user_name=(user["name"] if user else user_name),
        user_role=(user["role"] if user else user_role),
        action=action, module=module,
        entity_type=entity_type, entity_id=entity_id, entity_name=entity_name,
        old_values=_scrub_audit(old_values) if old_values else None,
        new_values=_scrub_audit(new_values) if new_values else None,
        description=description,
        ip_address=info["ip_address"], device_name=info["device_name"], client_type=info["client_type"],
        success=success,
    ).model_dump()
    try:
        await db.audit_log.insert_one(doc)
        if doc["user_id"] and success:
            await db.users.update_one({"id": doc["user_id"]}, {"$set": {"last_action": doc["created_at"]}})
    except Exception as ex:
        log.warning("audit insert failed: %s", ex)
    return doc

def _dict_diff(old: dict, new: dict) -> tuple:
    """Returns (changed_old, changed_new) with only fields whose value differs."""
    changed_old, changed_new = {}, {}
    old = old or {}; new = new or {}
    for k, v in new.items():
        if old.get(k) != v:
            changed_old[k] = old.get(k)
            changed_new[k] = v
    return changed_old, changed_new

# -----------------------------------------------------------------------------
# Startup
# -----------------------------------------------------------------------------
@app.on_event("startup")
async def startup():
    # Kritik indexler — büyük veri kümesinde sorgu hızı için
    try:
        await db.leaves.create_index("start_date")
        await db.leaves.create_index("personnel_id")
        await db.leaves.create_index([("personnel_id", 1), ("start_date", 1)])
        await db.leaves.create_index("end_date")
        # NOT: personnel.sicil_no ve users.email UNIQUE index'leri aşağıda
        # canonical bloğunda tanımlı — burada tekrar tanımlanmıyor.
        await db.personnel.create_index("aktif")
        await db.audit_log.create_index([("timestamp", -1)])
        await db.audit_log.create_index("module")
        await db.audit_log.create_index("action")
        await db.holidays_import.create_index("year")
        await db.holidays_import.create_index("date")
        await db.bulk_upload_history.create_index([("uploaded_at", -1)])
    except Exception:
        pass

    # ------------------------------------------------------------------------
    # CANONICAL UNIQUE INDEXES — idempotent (aynı key farklı unique spec varsa
    # güvenli şekilde yeniden oluşturulur). Iter 60: sicil_no non-unique →
    # unique migration için legacy state guard eklendi.
    # ------------------------------------------------------------------------
    async def _ensure_unique_index(coll, key_spec, name, **opts):
        """Idempotent unique index oluşturucu. Aynı ad'la mismatch spec varsa drop et."""
        try:
            existing = await coll.index_information()
        except Exception:
            existing = {}
        if name in existing:
            info = existing[name]
            # Beklenen key ve unique=True eşleşiyor mu?
            expected_key = [(key_spec, 1)] if isinstance(key_spec, str) else list(key_spec)
            same_key = info.get("key") == expected_key
            same_unique = bool(info.get("unique")) == bool(opts.get("unique", False))
            if same_key and same_unique:
                return  # zaten doğru; no-op
            # Spec farklı (ör. legacy non-unique) → drop ve yeniden oluştur
            try:
                await coll.drop_index(name)
                log.info("Migrated legacy index %s.%s → %s", coll.name, name, opts)
            except Exception as e:
                log.warning("drop_index(%s.%s) failed: %s", coll.name, name, e)
        try:
            await coll.create_index(key_spec, name=name, **opts)
        except Exception as e:
            log.warning("create_index(%s.%s) failed: %s", coll.name, name, e)

    await _ensure_unique_index(db.users, "email", "email_1", unique=True)
    await _ensure_unique_index(db.personnel, "sicil_no", "sicil_no_1", unique=True)
    await _ensure_unique_index(db.holidays, "date", "date_1", unique=True)
    await _ensure_unique_index(db.entitlements, [("personnel_id", 1), ("date", 1)], "personnel_id_1_date_1", unique=True)
    await _ensure_unique_index(db.audit_log, "id", "id_1", unique=True)
    await db.audit_log.create_index([("created_at", -1)])
    await db.audit_log.create_index([("user_id", 1), ("created_at", -1)])
    await db.audit_log.create_index([("entity_id", 1)])
    # Migrate legacy user documents to new fields
    await db.users.update_many({"aktif": {"$exists": False}}, {"$set": {"aktif": True}})
    await db.users.update_many({"username": {"$exists": False}}, [{"$set": {"username": "$email"}}])
    await db.users.update_many({"departman": {"$exists": False}}, {"$set": {"departman": ""}})
    await db.users.update_many({"aciklama": {"$exists": False}}, {"$set": {"aciklama": ""}})

    # seed admin (idempotent + duplicate-key guard)
    # Iter 60: Windows local Docker'da restart sonrası E11000 duplicate key hatası
    # oluşuyordu. Case-insensitive lookup + try/except ile tam idempotent yapıldı.
    email = os.environ["ADMIN_EMAIL"].lower()
    pw = os.environ["ADMIN_PASSWORD"]
    name = os.environ.get("ADMIN_NAME", "Yönetici")
    # Case-insensitive lookup (production'dan farklı case ile geçmiş olabilir)
    existing = await db.users.find_one({
        "email": {"$regex": f"^{re.escape(email)}$", "$options": "i"}
    })
    if existing is None:
        try:
            await db.users.insert_one({
                "id": str(uuid.uuid4()),
                "email": email,
                "password_hash": hash_password(pw),
                "name": name,
                "role": "admin",
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            log.info("Seeded admin %s", email)
        except Exception as e:
            # DuplicateKeyError vs. race — startup'ı düşürme, atla ve log'la
            log.warning("Admin seed skipped (already exists or race): %s", e)

    else:
        log.info(
            "Admin already exists; password left unchanged: %s",
            existing.get("email")
        )

# -----------------------------------------------------------------------------
# Auth routes
# -----------------------------------------------------------------------------
@api.post("/auth/login")
async def login(body: LoginIn, request: Request, response: Response):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user.get("password_hash", "")):
        await _audit(action="login_failed", module="auth", entity_type="user", entity_name=email,
                     description=f"Başarısız giriş: {email}", request=request, success=False)
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    if user.get("aktif") is False:
        await _audit(action="login_failed", module="auth", entity_type="user",
                     entity_id=user["id"], entity_name=user.get("name"),
                     description="Pasif kullanıcı giriş denedi",
                     request=request, user=user, success=False)
        raise HTTPException(status_code=403, detail="Kullanıcı pasif durumda")
    token = create_token(user["id"], user["email"], user["role"])
    response.set_cookie(
        key="access_token", value=token, httponly=True, secure=True,
        samesite="none", max_age=JWT_EXP_MIN * 60, path="/",
    )
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"id": user["id"]}, {"$set": {"last_login": now_iso}})
    await _audit(action="login_success", module="auth", entity_type="user",
                 entity_id=user["id"], entity_name=user.get("name"),
                 description="Başarılı giriş", request=request, user=user, success=True)
    return {
        "token": token,
        "user": {"id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"]},
    }

@api.post("/auth/logout")
async def logout(response: Response, request: Request):
    try:
        auth = request.headers.get("Authorization", "")
        token = auth[7:] if auth.startswith("Bearer ") else request.cookies.get("access_token")
        if token:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
            u = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
            if u:
                await _audit(action="logout", module="auth", entity_type="user",
                             entity_id=u["id"], entity_name=u.get("name"),
                             description="Çıkış yapıldı", request=request, user=u)
    except Exception:
        pass
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


# -----------------------------------------------------------------------------
# Health check — dış izleme + frontend teşhis için
# -----------------------------------------------------------------------------
@api.get("/health")
async def health_check():
    """Backend ve MongoDB canlılık kontrolü. Frontend Network Error tanısında da kullanılır.
    - 200 + {"status":"ok"} → backend + db erişilebilir
    - 503 → mongo bağlantı hatası
    """
    try:
        await db.command("ping")
        return {"status": "ok",
                "time": datetime.now(timezone.utc).isoformat(),
                "service": "merkoteks-hr",
                "database": "up"}
    except Exception as e:
        raise HTTPException(status_code=503,
                            detail=f"Veritabanı bağlantı hatası: {e}")

@api.get("/users")
async def list_users(_: dict = Depends(require_roles("admin"))):
    out = []
    async for u in db.users.find({}, {"_id": 0, "password_hash": 0}):
        out.append(u)
    out.sort(key=lambda x: (x.get("aktif", True) is False, (x.get("name") or "").lower()))
    return out

@api.post("/users")
async def create_user(body: UserExtendedIn, request: Request, current: dict = Depends(require_roles("admin"))):
    email = body.email.lower()
    username = (body.username or email).strip().lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Bu e-posta zaten kayıtlı")
    if username and await db.users.find_one({"username": username}):
        raise HTTPException(status_code=400, detail="Bu kullanıcı adı zaten kullanılıyor")
    if len(body.password) < 4:
        raise HTTPException(status_code=400, detail="Şifre en az 4 karakter olmalı")
    doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        "username": username or email,
        "password_hash": hash_password(body.password),
        "name": body.name,
        "role": body.role,
        "departman": body.departman or "",
        "aktif": bool(body.aktif),
        "aciklama": body.aciklama or "",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(doc)
    await _audit(action="create", module="users", entity_type="user", entity_id=doc["id"], entity_name=doc["name"],
                 new_values={k: doc[k] for k in ("email","username","name","role","departman","aktif","aciklama")},
                 description=f"Yeni kullanıcı oluşturuldu: {doc['name']} ({doc['email']})",
                 request=request, user=current)
    return {k: v for k, v in doc.items() if k not in ("password_hash", "_id")}

@api.put("/users/{user_id}")
async def update_user(user_id: str, body: UserUpdateIn, request: Request, current: dict = Depends(require_roles("admin"))):
    existing = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    updates = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    if "email" in updates:
        updates["email"] = updates["email"].lower()
        if updates["email"] != existing.get("email") and await db.users.find_one({"email": updates["email"], "id": {"$ne": user_id}}):
            raise HTTPException(status_code=400, detail="Bu e-posta başka bir kullanıcıda kayıtlı")
    if "username" in updates:
        updates["username"] = (updates["username"] or "").strip().lower()
        if updates["username"] and updates["username"] != (existing.get("username") or ""):
            if await db.users.find_one({"username": updates["username"], "id": {"$ne": user_id}}):
                raise HTTPException(status_code=400, detail="Bu kullanıcı adı başka bir kullanıcıda kayıtlı")
    if user_id == current["id"] and updates.get("aktif") is False:
        raise HTTPException(status_code=400, detail="Kendi hesabınızı pasif yapamazsınız")
    if user_id == current["id"] and "role" in updates and updates["role"] != current["role"]:
        raise HTTPException(status_code=400, detail="Kendi rolünüzü değiştiremezsiniz")
    if not updates:
        return existing
    await db.users.update_one({"id": user_id}, {"$set": updates})
    new_doc = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    changed_old, changed_new = _dict_diff(existing, updates)
    await _audit(action="update", module="users", entity_type="user", entity_id=user_id,
                 entity_name=new_doc.get("name"),
                 old_values=changed_old, new_values=changed_new,
                 description=f"Kullanıcı güncellendi: {new_doc.get('name')}",
                 request=request, user=current)
    return new_doc

@api.post("/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, body: ResetPasswordIn, request: Request, current: dict = Depends(require_roles("admin"))):
    existing = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    if len(body.new_password) < 4:
        raise HTTPException(status_code=400, detail="Şifre en az 4 karakter olmalı")
    await db.users.update_one({"id": user_id}, {"$set": {"password_hash": hash_password(body.new_password)}})
    await _audit(action="reset_password", module="users", entity_type="user", entity_id=user_id,
                 entity_name=existing.get("name"),
                 description=f"{existing.get('name')} kullanıcısının şifresi sıfırlandı",
                 request=request, user=current)
    return {"ok": True}

@api.post("/users/{user_id}/toggle-active")
async def toggle_user_active(user_id: str, request: Request, current: dict = Depends(require_roles("admin"))):
    existing = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    if user_id == current["id"]:
        raise HTTPException(status_code=400, detail="Kendi hesabınızı pasif yapamazsınız")
    new_state = not bool(existing.get("aktif", True))
    await db.users.update_one({"id": user_id}, {"$set": {"aktif": new_state}})
    await _audit(action=("activate" if new_state else "deactivate"), module="users",
                 entity_type="user", entity_id=user_id, entity_name=existing.get("name"),
                 old_values={"aktif": existing.get("aktif", True)}, new_values={"aktif": new_state},
                 description=f"{existing.get('name')} → {'aktif' if new_state else 'pasif'}",
                 request=request, user=current)
    return {"aktif": new_state}


def _clean_bson(v):
    """Nested dict/list içindeki BSON ObjectId'leri stringe çevirir.
    Bazı eski audit_log kayıtlarında old_values/new_values içine gömülü ObjectId'ler
    olabilir — FastAPI jsonable_encoder bunları serialize edemez ve 500 döner."""
    try:
        from bson import ObjectId as _OID
    except Exception:
        _OID = None
    if _OID is not None and isinstance(v, _OID):
        return str(v)
    if isinstance(v, dict):
        return {k: _clean_bson(x) for k, x in v.items()}
    if isinstance(v, list):
        return [_clean_bson(x) for x in v]
    return v


@api.get("/users/{user_id}/audit-log")
async def user_audit_log(user_id: str, limit: int = 200, _: dict = Depends(require_roles("admin"))):
    items = []
    async for a in db.audit_log.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).limit(min(1000, max(1, limit))):
        items.append(_clean_bson(a))
    return {"total": len(items), "items": items}

@api.get("/leaves/export.xlsx")
async def leaves_export_xlsx(personnel_id: Optional[str] = None,
                              start: Optional[str] = None, end: Optional[str] = None,
                              recent_days: Optional[int] = None,
                              izin_turu: Optional[str] = None,
                              q: Optional[str] = None,
                              departman: Optional[str] = None,
                              sirket: Optional[str] = None,
                              _: dict = Depends(get_current_user)):
    """Mevcut filtreye uyan izinleri Excel olarak indir."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    filt = await _build_leaves_filter(personnel_id, start, end, recent_days,
                                        izin_turu, q, departman, sirket)
    if filt is None:
        filt = {"_id": "___never_matches___"}
    p_cache: dict = {}
    async for p in db.personnel.find({}, {"_id": 0, "id": 1, "sicil_no": 1, "ad_soyad": 1, "departman": 1}):
        p_cache[p["id"]] = p
    holidays = await get_all_holidays()
    wb = Workbook(); ws = wb.active; ws.title = "İzinler"
    headers = ["Sicil No", "Ad Soyad", "Departman", "İzin Türü",
               "Başlangıç", "Bitiş", "Gün", "İşbaşı", "Açıklama", "Kayıt Tarihi"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1D4ED8")
        c.alignment = Alignment(horizontal="center")
    r = 2
    async for L in db.leaves.find(filt, {"_id": 0}).sort("start_date", -1):
        p = p_cache.get(L["personnel_id"], {})
        try:
            isbasi = _next_working_day(date.fromisoformat(L["end_date"]), holidays).isoformat()
        except Exception:
            isbasi = ""
        ws.cell(row=r, column=1, value=p.get("sicil_no", ""))
        ws.cell(row=r, column=2, value=p.get("ad_soyad", ""))
        ws.cell(row=r, column=3, value=p.get("departman", ""))
        ws.cell(row=r, column=4, value=L.get("izin_turu", ""))
        ws.cell(row=r, column=5, value=L.get("start_date", ""))
        ws.cell(row=r, column=6, value=L.get("end_date", ""))
        ws.cell(row=r, column=7, value=float(L.get("days", 0)))
        ws.cell(row=r, column=8, value=isbasi)
        ws.cell(row=r, column=9, value=L.get("aciklama", ""))
        ws.cell(row=r, column=10, value=(L.get("created_at") or "")[:10])
        r += 1
    for col, w in enumerate([14, 32, 22, 18, 14, 14, 8, 14, 30, 14], 1):
        ws.column_dimensions[chr(64 + col)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"izinler_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@api.get("/reports/excel-overrides")
async def reports_excel_overrides(limit: int = 100, skip: int = 0,
                                    _: dict = Depends(require_roles("admin", "hr"))):
    """Toplu Excel yüklemesinde kullanıcının sistem hesabını ezerek manuel gün belirttiği
    kayıtları listeler (audit_log.action = 'excel_days_override').
    """
    filt = {"action": "excel_days_override", "module": "leaves"}
    total = await db.audit_log.count_documents(filt)
    limit = max(1, min(int(limit), 500))
    skip = max(0, int(skip))
    items = []
    async for a in (db.audit_log.find(filt, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit)):
        nv = a.get("new_values") or {}
        items.append({
            "id": a.get("id"),
            "timestamp": a.get("timestamp"),
            "user_email": a.get("user_email"),
            "user_name": a.get("user_name"),
            "entity_name": a.get("entity_name"),
            "sicil_no": nv.get("sicil_no"),
            "ad_soyad": nv.get("ad_soyad") or a.get("entity_name"),
            "start_date": nv.get("start_date"),
            "end_date": nv.get("end_date"),
            "system_days": nv.get("system_days"),
            "manual_days": nv.get("excel_days") if nv.get("excel_days") is not None else nv.get("manual_days"),
            "difference": (
                None if nv.get("system_days") is None or (nv.get("excel_days") is None and nv.get("manual_days") is None)
                else round(float(nv.get("excel_days") if nv.get("excel_days") is not None else nv.get("manual_days")) - float(nv["system_days"]), 2)
            ),
            "reason": nv.get("reason") or nv.get("override_reason"),
            "filename": nv.get("filename") or a.get("description"),
        })
    return {"total": total, "items": items}


@api.post("/holidays/bulk-import-xlsx")
async def holidays_bulk_import_xlsx(file: UploadFile = File(...),
                                     request: Request = None,
                                     user: dict = Depends(require_roles("admin", "hr"))):
    """Excel dosyasından tatil listesi içe aktarır. İlk 3 kolon: Tarih | Tatil Adı | Gün Değeri."""
    from openpyxl import load_workbook
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Excel okunamadı")
    lines = []
    for r in range(1, (ws.max_row or 1) + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 4)]
        if not any(row): continue
        parts = []
        for v in row:
            if v is None: parts.append("")
            elif isinstance(v, (datetime, date)):
                parts.append(v.strftime("%d.%m.%Y") if isinstance(v, (datetime, date)) else str(v))
            else: parts.append(str(v).strip())
        if not parts[0]: continue
        lines.append("\t".join(parts))
    text = "\n".join(lines)
    return await holidays_bulk_import_text(
        {"text": text, "filename": file.filename or "tatil_listesi.xlsx"},
        request, user)


@api.delete("/holidays/records/{rid}")
async def delete_holiday_record(rid: str, request: Request,
                                 user: dict = Depends(require_roles("admin"))):
    existing = await db.holidays_import.find_one({"id": rid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    await db.holidays_import.delete_one({"id": rid})
    await _audit(action="delete", module="holidays", entity_type="holiday",
                 entity_id=rid, entity_name=existing.get("name"),
                 old_values=existing, description=f"Tatil kaydı silindi: {existing.get('date')} — {existing.get('name')}",
                 request=request, user=user)
    return {"ok": True}


@api.put("/holidays/records/{rid}")
async def update_holiday_record(rid: str, payload: dict, request: Request,
                                 user: dict = Depends(require_roles("admin", "hr"))):
    existing = await db.holidays_import.find_one({"id": rid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    updates = {}
    if "name" in payload and payload["name"] is not None:
        nm = str(payload["name"]).strip()
        if nm:
            updates["name"] = nm
            updates["needs_review"] = False
            updates["category"] = _classify_holiday(nm)
    if "type" in payload and payload["type"] in ("full", "half"):
        updates["type"] = payload["type"]
        updates["day_value"] = 0.5 if payload["type"] == "half" else 1.0
    if "day_value" in payload:
        try:
            dv = float(str(payload["day_value"]).replace(",", "."))
            updates["day_value"] = 0.5 if abs(dv - 0.5) < 1e-9 else 1.0
            updates["type"] = "half" if updates["day_value"] == 0.5 else "full"
        except Exception:
            pass
    if "date" in payload and payload["date"]:
        d_obj = _parse_import_date(str(payload["date"])) if not str(payload["date"]).startswith(str(payload["date"])[:4] + "-") else None
        if d_obj is None:
            try:
                d_obj = date.fromisoformat(str(payload["date"])[:10])
            except Exception:
                d_obj = None
        if d_obj:
            updates["date"] = d_obj.isoformat()
            updates["year"] = d_obj.year
    if "active" in payload:
        updates["active"] = bool(payload["active"])
    if not updates:
        return existing
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.holidays_import.update_one({"id": rid}, {"$set": updates})
    await _audit(action="update", module="holidays", entity_type="holiday",
                 entity_id=rid, entity_name=updates.get("name") or existing.get("name"),
                 old_values=existing, new_values=updates,
                 description=f"Tatil kaydı güncellendi: {existing.get('date')} — {updates.get('name') or existing.get('name')}",
                 request=request, user=user)
    return {**existing, **updates}


@api.post("/holidays/import/excel/preview")
async def holidays_import_excel_preview(file: UploadFile = File(...),
                                          user: dict = Depends(require_roles("admin", "hr"))):
    """Excel dosyasını önizler. DB'ye YAZMAZ. Beklenen kolonlar: Tarih | Tatil Tanımı | Gün Değeri.
    Satır statüleri: valid | duplicate | invalid | review.
    """
    from openpyxl import load_workbook
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Excel okunamadı")

    rows_out: list = []
    seen_pairs: set = set()
    stats = {"total": 0, "valid": 0, "duplicate": 0, "invalid": 0, "review": 0}

    for r in range(1, (ws.max_row or 1) + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, 4)]
        if not any(v not in (None, "") for v in cells):
            continue
        raw0 = "" if cells[0] is None else str(cells[0]).strip()
        raw1 = "" if cells[1] is None else str(cells[1]).strip()
        low0 = raw0.lower()
        low1 = raw1.lower()
        if ("tarih" in low0 and ("tatil" in low1 or "tanım" in low1)) or low0 == "tarih":
            continue

        stats["total"] += 1
        row = {"row": r, "raw_date": raw0, "raw_name": raw1,
               "raw_day": ("" if cells[2] is None else str(cells[2]))}

        # tarih
        d_obj = None
        if isinstance(cells[0], datetime):
            d_obj = cells[0].date()
        elif isinstance(cells[0], date):
            d_obj = cells[0]
        else:
            d_obj = _parse_import_date(raw0)
        if not d_obj:
            row["status"] = "invalid"
            row["reason"] = "Geçersiz tarih"
            stats["invalid"] += 1
            rows_out.append(row)
            continue
        row["date"] = d_obj.isoformat()
        row["date_tr"] = d_obj.strftime("%d.%m.%Y")
        row["year"] = d_obj.year

        # gün değeri
        try:
            if cells[2] is None or str(cells[2]).strip() == "":
                dv = 1.0
            else:
                dv = float(str(cells[2]).replace(",", "."))
        except Exception:
            row["status"] = "invalid"
            row["reason"] = "Geçersiz gün değeri"
            stats["invalid"] += 1
            rows_out.append(row)
            continue
        row["day_value"] = 0.5 if abs(dv - 0.5) < 1e-9 else 1.0
        row["type"] = "half" if row["day_value"] == 0.5 else "full"
        row["type_label"] = "Yarım Gün" if row["type"] == "half" else "Tam Gün"

        # ad
        name_clean = raw1.strip() if raw1 else ""
        display_name = name_clean if name_clean else "Tatil Tanımı Belirtilmemiş"
        row["name"] = display_name
        row["needs_review"] = not name_clean

        # dosya içi mükerrer
        pair = (row["date"], display_name.lower())
        if pair in seen_pairs:
            row["status"] = "duplicate"
            row["reason"] = "Dosya içinde mükerrer"
            stats["duplicate"] += 1
            rows_out.append(row)
            continue
        seen_pairs.add(pair)

        # DB'de mükerrer kontrolü
        existing = await db.holidays_import.find_one(
            {"date": row["date"], "name": display_name}, {"_id": 0, "id": 1})
        if existing:
            row["status"] = "duplicate"
            row["reason"] = "Kayıt zaten mevcut (aynı tarih + tanım)"
            stats["duplicate"] += 1
        elif not name_clean:
            row["status"] = "review"
            stats["review"] += 1
            stats["valid"] += 1  # yine de yazılabilir
        else:
            row["status"] = "valid"
            stats["valid"] += 1
        rows_out.append(row)

    return {"filename": file.filename or "tatil_listesi.xlsx",
            "stats": stats, "rows": rows_out}


@api.post("/holidays/import/excel/confirm")
async def holidays_import_excel_confirm(payload: dict, request: Request,
                                          user: dict = Depends(require_roles("admin", "hr"))):
    """Önizlemede onaylanan satırları DB'ye yazar."""
    rows_in = (payload or {}).get("rows", [])
    filename = (payload or {}).get("filename", "tatil_listesi.xlsx")
    if not isinstance(rows_in, list) or not rows_in:
        raise HTTPException(status_code=400, detail="Yazılacak kayıt bulunamadı")

    added = 0
    skipped = 0
    affected_years: set = set()
    for r in rows_in:
        if r.get("status") not in ("valid", "review"):
            skipped += 1
            continue
        iso = r.get("date")
        name = str(r.get("name") or "Tatil Tanımı Belirtilmemiş").strip()
        if not iso:
            skipped += 1
            continue
        try:
            dv = float(str(r.get("day_value", 1.0)).replace(",", "."))
        except Exception:
            dv = 1.0
        dv = 0.5 if abs(dv - 0.5) < 1e-9 else 1.0
        h_type = "half" if dv == 0.5 else "full"

        existing = await db.holidays_import.find_one(
            {"date": iso, "name": name}, {"_id": 0, "id": 1})
        if existing:
            skipped += 1
            continue

        needs_review = bool(r.get("needs_review")) or (name == "Tatil Tanımı Belirtilmemiş")
        doc = {
            "id": str(uuid.uuid4()),
            "date": iso,
            "year": int(r.get("year") or iso[:4]),
            "name": name,
            "day_value": dv,
            "type": h_type,
            "category": _classify_holiday(name) if not needs_review else "Diğer",
            "source": f"Excel: {filename}",
            "active": True,
            "needs_review": needs_review,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.holidays_import.insert_one(doc)
        added += 1
        affected_years.add(doc["year"])

    result = {"added": added, "skipped": skipped,
              "affected_years": sorted(list(affected_years))}
    await _audit(
        action="bulk_import_excel", module="holidays", entity_type="holiday",
        entity_id=None, entity_name=f"Excel aktarım: {added} kayıt",
        new_values={"filename": filename, **result},
        description=f"Excel tatil aktarımı: +{added} yeni, {skipped} atlandı, yıllar: {result['affected_years']}",
        request=request, user=user,
    )
    return result


@api.post("/holidays/records/bulk-update")
async def holidays_records_bulk_update(payload: dict, request: Request,
                                         user: dict = Depends(require_roles("admin", "hr"))):
    """Seçili tatil kayıtlarını toplu günceller (ad ve/veya aktif alanı)."""
    ids = payload.get("ids") or []
    new_name = payload.get("name")
    active = payload.get("active")
    if not ids:
        raise HTTPException(status_code=400, detail="Kayıt seçilmedi")
    updates: dict = {}
    if new_name is not None and str(new_name).strip():
        nm = str(new_name).strip()
        updates["name"] = nm
        updates["needs_review"] = False
        updates["category"] = _classify_holiday(nm)
    if active is not None:
        updates["active"] = bool(active)
    if not updates:
        raise HTTPException(status_code=400, detail="Güncellenecek alan yok")
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.holidays_import.update_many({"id": {"$in": ids}}, {"$set": updates})
    await _audit(action="bulk_update", module="holidays", entity_type="holiday",
                 entity_id=None, entity_name=f"Toplu güncelleme ({len(ids)})",
                 new_values={"ids": ids, **updates},
                 description=f"Toplu tatil güncellemesi: {result.modified_count} kayıt",
                 request=request, user=user)
    return {"updated": result.modified_count}


@api.post("/holidays/records/bulk-delete")
async def holidays_records_bulk_delete(payload: dict, request: Request,
                                         user: dict = Depends(require_roles("admin"))):
    """Seçili tatil kayıtlarını toplu siler. Yalnızca admin — şifre + gerekçe zorunlu."""
    ids = payload.get("ids") or []
    password = payload.get("password") or ""
    reason = (payload.get("reason") or "").strip()
    if not ids:
        raise HTTPException(status_code=400, detail="Kayıt seçilmedi")
    if not reason:
        raise HTTPException(status_code=400, detail="Gerekçe zorunlu")
    admin = await db.users.find_one({"id": user["id"]})
    if not admin or not verify_password(password, admin.get("password_hash", "")):
        await _audit(action="delete_failed", module="holidays", entity_type="holiday",
                     description=f"Toplu tatil silme başarısız (yanlış şifre): {len(ids)} kayıt",
                     request=request, user=user)
        raise HTTPException(status_code=403, detail="Şifre hatalı")
    docs = []
    async for h in db.holidays_import.find({"id": {"$in": ids}}, {"_id": 0}):
        docs.append(h)
    result = await db.holidays_import.delete_many({"id": {"$in": ids}})
    await _audit(action="bulk_delete", module="holidays", entity_type="holiday",
                 entity_id=None, entity_name=f"Toplu silme ({len(ids)})",
                 old_values={"records": docs, "reason": reason},
                 description=f"Toplu tatil silme: {result.deleted_count} kayıt — Gerekçe: {reason}",
                 request=request, user=user)
    return {"deleted": result.deleted_count}


@api.get("/audit-log")
async def list_audit_log(
    user_id: Optional[str] = None,
    user_role: Optional[str] = None,
    module: Optional[str] = None,
    action: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    success: Optional[bool] = None,
    entity_name: Optional[str] = None,
    entity_id: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    _: dict = Depends(require_roles("admin")),
):
    filt: dict = {}
    if user_id: filt["user_id"] = user_id
    if user_role: filt["user_role"] = user_role
    if module: filt["module"] = module
    if action: filt["action"] = action
    if success is not None: filt["success"] = success
    if entity_id: filt["entity_id"] = entity_id
    if entity_name:
        filt["entity_name"] = {"$regex": entity_name, "$options": "i"}
    if start or end:
        rng: dict = {}
        if start: rng["$gte"] = start
        if end: rng["$lte"] = end + "T23:59:59"
        filt["created_at"] = rng
    if q:
        filt["$or"] = [
            {"description": {"$regex": q, "$options": "i"}},
            {"entity_name": {"$regex": q, "$options": "i"}},
            {"user_name": {"$regex": q, "$options": "i"}},
        ]
    total = await db.audit_log.count_documents(filt)
    items = []
    async for a in db.audit_log.find(filt, {"_id": 0}).sort("created_at", -1).skip(max(0, skip)).limit(min(500, max(1, limit))):
        items.append(_clean_bson(a))
    return {"total": total, "items": items}


async def _audit_query(user_id, user_role, module, action, start, end,
                        success, entity_name, entity_id, q, limit, skip):
    """Ortak audit sorgu — export endpointleri için."""
    filt: dict = {}
    if user_id: filt["user_id"] = user_id
    if user_role: filt["user_role"] = user_role
    if module: filt["module"] = module
    if action: filt["action"] = action
    if success is not None: filt["success"] = success
    if entity_id: filt["entity_id"] = entity_id
    if entity_name:
        filt["entity_name"] = {"$regex": entity_name, "$options": "i"}
    if start or end:
        rng: dict = {}
        if start: rng["$gte"] = start
        if end: rng["$lte"] = end + "T23:59:59"
        filt["created_at"] = rng
    if q:
        filt["$or"] = [
            {"description": {"$regex": q, "$options": "i"}},
            {"entity_name": {"$regex": q, "$options": "i"}},
            {"user_name": {"$regex": q, "$options": "i"}},
        ]
    items = []
    async for a in db.audit_log.find(filt, {"_id": 0}).sort("created_at", -1).skip(max(0, skip)).limit(min(10000, max(1, limit))):
        items.append(a)
    return items


def _fmt_datetime_tr(iso: Optional[str]) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return iso[:16].replace("T", " ")

@api.get("/audit-log/export.xlsx")
async def export_audit_log_xlsx(
    user_id: Optional[str] = None, user_role: Optional[str] = None,
    module: Optional[str] = None, action: Optional[str] = None,
    start: Optional[str] = None, end: Optional[str] = None,
    success: Optional[bool] = None, entity_name: Optional[str] = None,
    entity_id: Optional[str] = None, q: Optional[str] = None,
    limit: int = 5000, skip: int = 0,
    request: Request = None, current: dict = Depends(require_roles("admin")),
):
    items = await _audit_query(user_id, user_role, module, action, start, end,
                                success, entity_name, entity_id, q, limit, skip)
    _module_label = {"auth": "Kimlik Doğrulama", "users": "Kullanıcılar",
                      "personnel": "Personel", "leaves": "İzinler",
                      "holidays": "Tatiller", "reports": "Raporlar"}
    _action_label = {"create": "Oluştur", "update": "Güncelle", "delete": "Sil",
                      "terminate": "İşten Ayrılış", "reactivate": "Aktife Al",
                      "bulk_create": "Toplu Oluşturma", "reset_password": "Şifre Sıfırla",
                      "activate": "Aktifleştir", "deactivate": "Pasifleştir",
                      "login_success": "Başarılı Giriş", "login_failed": "Başarısız Giriş",
                      "logout": "Çıkış"}
    _role_label = {"admin": "Yönetici", "hr": "İnsan Kaynakları", "viewer": "Sadece Rapor"}
    headers = ["Tarih & Saat", "Kullanıcı", "Rol", "Modül", "İşlem",
               "Etkilenen Kayıt", "Açıklama", "IP Adresi", "Cihaz", "Durum"]
    rows = []
    for a in items:
        rows.append([
            _fmt_datetime_tr(a.get("created_at")),
            a.get("user_name") or "",
            _role_label.get(a.get("user_role"), a.get("user_role") or ""),
            _module_label.get(a.get("module"), a.get("module") or ""),
            _action_label.get(a.get("action"), a.get("action") or ""),
            a.get("entity_name") or "",
            a.get("description") or "",
            a.get("ip_address") or "",
            a.get("device_name") or "",
            "Başarılı" if a.get("success") else "Başarısız",
        ])
    xlsx = _xlsx(headers, rows, "Denetim Kayıtları")
    await _audit(action="export", module="reports", entity_type="audit_log_export",
                 entity_name=f"Denetim Kayıtları (xlsx, {len(rows)} satır)",
                 description=f"Denetim kayıtları Excel olarak indirildi: {len(rows)} kayıt",
                 request=request, user=current)
    return Response(content=xlsx,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="denetim-kayitlari.xlsx"'})


@api.get("/audit-log/export.pdf")
async def export_audit_log_pdf(
    user_id: Optional[str] = None, user_role: Optional[str] = None,
    module: Optional[str] = None, action: Optional[str] = None,
    start: Optional[str] = None, end: Optional[str] = None,
    success: Optional[bool] = None, entity_name: Optional[str] = None,
    entity_id: Optional[str] = None, q: Optional[str] = None,
    limit: int = 5000, skip: int = 0,
    request: Request = None, current: dict = Depends(require_roles("admin")),
):
    items = await _audit_query(user_id, user_role, module, action, start, end,
                                success, entity_name, entity_id, q, limit, skip)
    _module_label = {"auth": "Kimlik Doğrulama", "users": "Kullanıcılar",
                      "personnel": "Personel", "leaves": "İzinler",
                      "holidays": "Tatiller", "reports": "Raporlar"}
    _action_label = {"create": "Oluştur", "update": "Güncelle", "delete": "Sil",
                      "terminate": "İşten Ayrılış", "reactivate": "Aktife Al",
                      "bulk_create": "Toplu Oluşturma", "reset_password": "Şifre Sıfırla",
                      "activate": "Aktifleştir", "deactivate": "Pasifleştir",
                      "login_success": "Başarılı Giriş", "login_failed": "Başarısız Giriş",
                      "logout": "Çıkış"}
    _role_label = {"admin": "Yönetici", "hr": "İnsan Kaynakları", "viewer": "Sadece Rapor"}
    headers = ["Tarih & Saat", "Kullanıcı", "Rol", "Modül", "İşlem",
               "Etkilenen Kayıt", "IP", "Durum"]
    rows = []
    for a in items:
        rows.append([
            _fmt_datetime_tr(a.get("created_at")),
            a.get("user_name") or "",
            _role_label.get(a.get("user_role"), a.get("user_role") or ""),
            _module_label.get(a.get("module"), a.get("module") or ""),
            _action_label.get(a.get("action"), a.get("action") or ""),
            a.get("entity_name") or "",
            a.get("ip_address") or "",
            "✔" if a.get("success") else "✖",
        ])
    pdf = _pdf(f"Denetim Kayıtları Raporu ({len(rows)} kayıt)", headers, rows)
    await _audit(action="export", module="reports", entity_type="audit_log_export",
                 entity_name=f"Denetim Kayıtları (pdf, {len(rows)} satır)",
                 description=f"Denetim kayıtları PDF olarak indirildi: {len(rows)} kayıt",
                 request=request, user=current)
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": 'attachment; filename="denetim-kayitlari.pdf"'})


@api.get("/audit-log/{aid}")
async def get_audit(aid: str, _: dict = Depends(require_roles("admin"))):
    a = await db.audit_log.find_one({"id": aid}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Audit kaydı bulunamadı")
    return a

class UserDeleteIn(BaseModel):
    password: str
    reason: str

@api.post("/users/{user_id}/delete")
async def user_hard_delete(user_id: str, body: UserDeleteIn, request: Request,
                            current: dict = Depends(require_roles("admin"))):
    if not body.password or not body.reason or not body.reason.strip():
        raise HTTPException(status_code=400, detail="Yönetici şifresi ve silme gerekçesi zorunlu")
    if user_id == current["id"]:
        raise HTTPException(status_code=400, detail="Kendi hesabınızı silemezsiniz")
    admin = await db.users.find_one({"id": current["id"]})
    if not admin or not verify_password(body.password, admin.get("password_hash", "")):
        await _audit(action="delete_failed", module="users", entity_type="user",
                     entity_id=user_id,
                     description=f"Kullanıcı silme reddedildi (şifre) — gerekçe: {body.reason[:200]}",
                     request=request, user=current, success=False)
        raise HTTPException(status_code=403, detail="Yönetici şifresi doğrulanamadı. Kullanıcı silinmedi.")
    existing = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    # Son aktif admin koruması
    if existing.get("role") == "admin" and existing.get("aktif") is not False:
        active_admins = await db.users.count_documents({"role": "admin", "aktif": {"$ne": False}})
        if active_admins <= 1:
            raise HTTPException(status_code=400, detail="Sistemde tek aktif yönetici kalmış — silinemez")
    await db.users.delete_one({"id": user_id})
    await _audit(action="hard_delete", module="users", entity_type="user",
                 entity_id=user_id, entity_name=existing.get("name"),
                 old_values={"email": existing.get("email"), "username": existing.get("username"),
                              "role": existing.get("role"), "aktif": existing.get("aktif", True)},
                 new_values={"reason": body.reason.strip()},
                 description=f"Kullanıcı KALICI silindi: {existing.get('name')} ({existing.get('email')}) — gerekçe: {body.reason.strip()[:200]}",
                 request=request, user=current)
    return {"ok": True}

class WipeIn(BaseModel):
    password: str
    reason: str
    confirm: str  # "SİL" yazılmalı


@api.post("/personnel/wipe-all")
async def personnel_wipe_all(body: WipeIn, request: Request,
                              current: dict = Depends(require_roles("admin"))):
    """Iter 50: TÜM personel ve ilgili tüm izin/hak ediş kayıtları kalıcı silinir.
    Admin şifresi + gerekçe + 'SİL' onayı zorunlu. Audit_log korunur.
    """
    if not body.password or not body.reason or not body.reason.strip():
        raise HTTPException(status_code=400, detail="Yönetici şifresi ve silme gerekçesi zorunlu")
    if body.confirm != "SİL":
        raise HTTPException(status_code=400, detail="Onay için 'SİL' yazmalısınız")
    admin = await db.users.find_one({"id": current["id"]})
    if not admin or not verify_password(body.password, admin.get("password_hash", "")):
        raise HTTPException(status_code=403, detail="Yönetici şifresi doğrulanamadı")
    p_count = await db.personnel.count_documents({})
    l_count = await db.leaves.count_documents({})
    e_count = await db.entitlements.count_documents({})
    await db.leaves.delete_many({})
    await db.entitlements.delete_many({})
    await db.personnel.delete_many({})
    await _audit(action="wipe_all", module="personnel", entity_type="personnel",
                 entity_id="*", entity_name=f"{p_count} personel",
                 old_values={"personnel": p_count, "leaves": l_count, "entitlements": e_count},
                 new_values={"reason": body.reason.strip()},
                 description=f"TÜM PERSONEL SİLİNDİ: {p_count} personel, {l_count} izin, {e_count} hak ediş — gerekçe: {body.reason.strip()[:200]}",
                 request=request, user=current)
    return {"ok": True, "personnel": p_count, "leaves": l_count, "entitlements": e_count}


@api.post("/leaves/wipe-all")
async def leaves_wipe_all(body: WipeIn, request: Request,
                           current: dict = Depends(require_roles("admin"))):
    """Iter 50: TÜM izin ve hak ediş kayıtları kalıcı silinir. Personel korunur.
    Admin şifresi + gerekçe + 'SİL' onayı zorunlu.
    """
    if not body.password or not body.reason or not body.reason.strip():
        raise HTTPException(status_code=400, detail="Yönetici şifresi ve silme gerekçesi zorunlu")
    if body.confirm != "SİL":
        raise HTTPException(status_code=400, detail="Onay için 'SİL' yazmalısınız")
    admin = await db.users.find_one({"id": current["id"]})
    if not admin or not verify_password(body.password, admin.get("password_hash", "")):
        raise HTTPException(status_code=403, detail="Yönetici şifresi doğrulanamadı")
    l_count = await db.leaves.count_documents({})
    e_count = await db.entitlements.count_documents({})
    await db.leaves.delete_many({})
    await db.entitlements.delete_many({})
    await _audit(action="wipe_all", module="leaves", entity_type="leave",
                 entity_id="*", entity_name=f"{l_count} izin kaydı",
                 old_values={"leaves": l_count, "entitlements": e_count},
                 new_values={"reason": body.reason.strip()},
                 description=f"TÜM İZİN KAYITLARI SİLİNDİ: {l_count} izin, {e_count} hak ediş — gerekçe: {body.reason.strip()[:200]}",
                 request=request, user=current)
    return {"ok": True, "leaves": l_count, "entitlements": e_count}


class AuditWipeIn(BaseModel):
    password: str
    reason: str
    confirm: str  # "SİL"
    include_admin: bool = False   # Yönetici işlem geçmişi
    include_non_admin: bool = False  # Kullanıcı (admin dışı) işlem geçmişi


@api.post("/audit-log/wipe")
async def audit_wipe(body: AuditWipeIn, request: Request,
                      current: dict = Depends(require_roles("admin"))):
    """Iter 52: İşlem geçmişini seçime göre kalıcı siler.
    include_admin=True → user_role='admin' kayıtları
    include_non_admin=True → user_role in ('hr','viewer',None,'') kayıtları
    En az biri True olmalı. Silme işleminin kendi kaydı (bu wipe) her zaman KORUNUR.
    """
    if not body.password or not body.reason or not body.reason.strip():
        raise HTTPException(status_code=400, detail="Yönetici şifresi ve silme gerekçesi zorunlu")
    if body.confirm != "SİL":
        raise HTTPException(status_code=400, detail="Onay için 'SİL' yazmalısınız")
    if not (body.include_admin or body.include_non_admin):
        raise HTTPException(status_code=400, detail="En az bir kategori seçilmeli")
    admin = await db.users.find_one({"id": current["id"]})
    if not admin or not verify_password(body.password, admin.get("password_hash", "")):
        raise HTTPException(status_code=403, detail="Yönetici şifresi doğrulanamadı")
    q: dict = {}
    if body.include_admin and body.include_non_admin:
        q = {}  # hepsi
    elif body.include_admin:
        q = {"user_role": "admin"}
    else:
        q = {"user_role": {"$ne": "admin"}}
    count = await db.audit_log.count_documents(q)
    await db.audit_log.delete_many(q)
    await _audit(action="wipe_all", module="audit", entity_type="audit_log",
                 entity_id="*", entity_name=f"{count} kayıt",
                 old_values={"count": count, "filter": str(q)},
                 new_values={"reason": body.reason.strip(),
                              "include_admin": body.include_admin,
                              "include_non_admin": body.include_non_admin},
                 description=f"İşlem geçmişi silindi: {count} kayıt — gerekçe: {body.reason.strip()[:200]}",
                 request=request, user=current)
    return {"ok": True, "deleted": count}


# NOTE: DELETE /users/{user_id} kaldırıldı — fiziksel silme yerine toggle-active kullanılır.

# -----------------------------------------------------------------------------
# Personnel routes
# -----------------------------------------------------------------------------
async def _invalidate_cetvel_for(personnel_ids) -> None:
    """Iter 37: Yeni izin girişi yapıldığında personelin cetvel_generated_at bayrağı
    silinir — böylece Personel listesindeki 'İzin Cetveli' sütunu tekrar
    'Doldurulmalı' yazar. Kullanıcı yeni cetveli üretip onayladığında bayrak yeniden
    set edilir (POST /api/personnel/{pid}/cetvel-mark).

    Iter 57: Daha önce cetveli 'oluşturuldu' olarak işaretlenmiş personel için
    'cetvel_needs_refill=true' bayrağı set edilir; böylece Personel Detay ekranında
    'İzin Cetveli Yeniden Doldurulmalıdır' uyarısı gösterilebilir. Cetveli hiç
    oluşturulmamış personelde bu bayrak set edilmez.
    """
    if not personnel_ids:
        return
    if isinstance(personnel_ids, str):
        personnel_ids = [personnel_ids]
    ids = [pid for pid in {p for p in personnel_ids if p}]
    if not ids:
        return
    try:
        # Sadece daha önce cetveli oluşturulmuş personel için needs_refill bayrağını set et.
        await db.personnel.update_many(
            {"id": {"$in": ids}, "cetvel_generated_at": {"$exists": True}},
            {"$set": {"cetvel_needs_refill": True},
             "$unset": {"cetvel_generated_at": ""}},
        )
    except Exception:
        pass


@api.post("/personnel/{pid}/cetvel-mark")
async def personnel_cetvel_mark(pid: str, request: Request,
                                  user: dict = Depends(require_roles("admin", "hr"))):
    """Iter 37: Personelin izin cetveli oluşturuldu olarak işaretlenir.
    Sonraki yeni izin girişine kadar 'Uygun/Oluşturuldu' olarak kalır."""
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    ts = datetime.now(timezone.utc).isoformat()
    await db.personnel.update_one({"id": pid},
                                    {"$set": {"cetvel_generated_at": ts,
                                              "cetvel_generated_by": user.get("id")},
                                     "$unset": {"cetvel_needs_refill": ""}})
    await _audit(action="cetvel_mark", module="personnel", entity_type="personnel",
                 entity_id=pid, entity_name=p.get("ad_soyad"),
                 new_values={"cetvel_generated_at": ts},
                 description=f"İzin Cetveli oluşturuldu olarak işaretlendi: {p.get('ad_soyad')}",
                 request=request, user=user)
    return {"ok": True, "cetvel_generated_at": ts}


@api.post("/personnel/{pid}/cetvel-unmark")
async def personnel_cetvel_unmark(pid: str, request: Request,
                                    user: dict = Depends(require_roles("admin", "hr"))):
    """Iter 41: İzin Cetveli 'oluşturuldu' bayrağını iptal eder — Personel listesinde
    tekrar 'Doldurulmalı' olarak gösterilir."""
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    prev = p.get("cetvel_generated_at")
    await db.personnel.update_one({"id": pid},
                                    {"$unset": {"cetvel_generated_at": "",
                                                "cetvel_generated_by": "",
                                                "cetvel_needs_refill": ""}})
    await _audit(action="cetvel_unmark", module="personnel", entity_type="personnel",
                 entity_id=pid, entity_name=p.get("ad_soyad"),
                 old_values={"cetvel_generated_at": prev},
                 new_values={"cetvel_generated_at": None},
                 description=f"İzin Cetveli iptal edildi: {p.get('ad_soyad')}",
                 request=request, user=user)
    return {"ok": True, "cetvel_generated_at": None}


def _ten_day_check_from_leaves(entitlements: list, next_ent_iso: Optional[str],
                                 leaves: list, today: date) -> dict:
    """4857 s.K. m.56 — Yıllık iznin bir bölümü 10 günden aşağı olamaz.

    Kontrol İÇİNDE BULUNULAN TAKVİM YILI (01.01 – 31.12) için sıfırdan yapılır.
    Yıl değiştiğinde durum otomatik olarak sıfırlanır — geçmiş yıllardaki 10+
    günlük izinler cari yıl kontrolünü ETKİLEMEZ.

    Sadece Yıllık İzin türü dikkate alınır. Özel izinler (evlilik, cenaze, süt,
    gebelik/doğum, ücretsiz, rapor vb.) bu kontrolden tamamen bağımsızdır.

    Kural: TEK BİR izin kaydının süresi (Cmt/Paz/tam tatil hariç motor günü) >= 10
    olmalıdır. Parça toplamı (5+5, 4+3+3) sayılmaz.

    Durumlar:
      - earned_ok  → Bu yıl içinde 10+ günlük tek parça yıllık izin var VE iznin
                      başlangıcından önce (veya aynı gün) bir hak ediş tarihi
                      mevcut. (Hak edilmiş bakiyeden kullanım.)
      - advance_ok → Bu yıl içinde 10+ günlük tek parça yıllık izin var, ancak
                      hiçbir 10+ parça için hak ediş oluşmamış. (Avans / erken.)
      - missing    → Bu yıl içinde 10+ günlük tek parça yıllık izin yok.

    Öncelik earned_ok > advance_ok > missing. Aktif her personel bu 3 durumdan
    birini alır (— yerine mutlaka bir ikon gösterilir).
    """
    year = today.year
    year_start_iso = f"{year:04d}-01-01"
    year_end_iso_excl = f"{year + 1:04d}-01-01"

    def _is_annual(L: dict) -> bool:
        t = (L.get("izin_turu") or "").strip().lower()
        # Boş tür varsayılan olarak yıllık kabul edilir
        return (not t) or ("yıllık" in t) or ("yillik" in t) or (t == "annual")

    # Sadece bu takvim yılında BAŞLAYAN yıllık izinler
    annual_leaves = []
    for L in (leaves or []):
        sd = L.get("start_date") or ""
        if not sd:
            continue
        if not (year_start_iso <= sd < year_end_iso_excl):
            continue
        if _is_annual(L):
            annual_leaves.append(L)

    # Tek parça >= 10 gün olan (bu yılki) yıllık izinler
    big_slices = [L for L in annual_leaves if float(L.get("days") or 0) >= 10.0 - 1e-9]

    if not big_slices:
        max_slice = max((float(L.get("days") or 0) for L in annual_leaves), default=0.0)
        return {
            "status": "missing",
            "year": year,
            "max_slice_days": round(max_slice, 2),
            "big_slice_count": 0,
            "earned_big_count": 0,
            "advance_big_count": 0,
        }

    # Hak ediş tarihleri (ISO string) — tüm entitlementler (yıl bağımsız),
    # çünkü kullanılan iznin niteliğini o iznin başlangıç tarihine göre belirliyoruz.
    ent_dates_iso = sorted([e.get("date") for e in (entitlements or []) if e.get("date")])
    earned, advance = [], []
    for L in big_slices:
        sd = L.get("start_date") or ""
        # İznin başlangıcında hak edilmiş bir dönem başlamışsa earned
        if ent_dates_iso and any(ed <= sd for ed in ent_dates_iso):
            earned.append(L)
        else:
            advance.append(L)

    max_earned = max((float(L.get("days") or 0) for L in earned), default=0.0)
    max_advance = max((float(L.get("days") or 0) for L in advance), default=0.0)
    max_slice = max(max_earned, max_advance)

    # Öncelik: earned_ok > advance_ok
    status = "earned_ok" if earned else "advance_ok"
    return {
        "status": status,
        "year": year,
        "max_slice_days": round(max_slice, 2),
        "big_slice_count": len(big_slices),
        "earned_big_count": len(earned),
        "advance_big_count": len(advance),
    }


@api.get("/personnel/balance-summary")
async def personnel_balance_summary(aktif: Optional[bool] = None, _: dict = Depends(get_current_user)):
    """Tüm personelin (veya aktif filtresine göre) kalan izin özetini tek sorguda döner (N+1 önler).
    Iter 36: Her personel için `ten_day_check` alanı eklendi — mevcut hak ediş döneminde
    en az bir 10-gün kesintisiz yıllık izin kullanımı var mı kontrolü.
    """
    q: dict = {}
    if aktif is not None:
        q["aktif"] = aktif
    # Tüm izinleri tek seferde çek — personel bazlı grupla (N+1 önle)
    leaves_by_pid: dict = {}
    async for L in db.leaves.find({}, {"_id": 0, "personnel_id": 1, "start_date": 1,
                                         "end_date": 1, "days": 1, "izin_turu": 1}):
        leaves_by_pid.setdefault(L.get("personnel_id"), []).append(L)
    today = date.today()
    out = []
    async for p in db.personnel.find(q, {"_id": 0}):
        try:
            bal = await _compute_entitlements(p)
            ten = _ten_day_check_from_leaves(
                bal.get("entitlements", []),
                (bal.get("next_entitlement") or {}).get("date"),
                leaves_by_pid.get(p["id"], []),
                today,
            )
            out.append({
                "id": p["id"], "sicil_no": p.get("sicil_no"),
                "ad_soyad": p.get("ad_soyad"), "departman": p.get("departman"),
                "entitled_total": bal["entitled_total"],
                "used_total": bal["used_total"],
                "remaining": bal["remaining"],
                "ten_day_check": ten,
                "cetvel_generated_at": p.get("cetvel_generated_at"),
            })
        except Exception:
            pass
    return out


def _build_personnel_filter(q: Optional[str], departman: Optional[str],
                             sirket: Optional[str], aktif: Optional[bool]) -> dict:
    filt: dict = {}
    if aktif is not None:
        filt["aktif"] = aktif
    if departman:
        filt["departman"] = {"$regex": f"^{re.escape(departman)}$", "$options": "i"}
    if sirket:
        filt["sirket"] = {"$regex": f"^{re.escape(sirket)}$", "$options": "i"}
    if q:
        # Türkçe karakter duyarlı büyük/küçük harf araması.
        esc_name = _tr_search_regex(q)
        esc_other = re.escape(q.strip())
        filt["$or"] = [
            {"ad_soyad": {"$regex": esc_name, "$options": "i"}},
            {"sicil_no": {"$regex": esc_other, "$options": "i"}},
            {"tc_no": {"$regex": esc_other, "$options": "i"}},
        ]
    return filt


async def _advance_ok_pids() -> set:
    """ten_day_check.status == 'advance_ok' olan aktif personel ID'leri.
    Sarı Rozet Filtresi için: bu yıl 10+ günlük tek parça yıllık izin kullanan ama
    o iznin başlangıcında hak edişi olmayan (avans) personeller."""
    today = date.today()
    # Tek sorguda tüm aktif personel + tüm leaves
    p_list = []
    async for p in db.personnel.find({"aktif": {"$ne": False}}, {"_id": 0}):
        p_list.append(p)
    leaves_by_pid: dict = {}
    async for L in db.leaves.find({"personnel_id": {"$in": [p["id"] for p in p_list]}}, {"_id": 0}):
        leaves_by_pid.setdefault(L["personnel_id"], []).append(L)
    pids = set()
    for p in p_list:
        try:
            bal = await _compute_entitlements(p)
        except Exception:
            continue
        tdc = _ten_day_check_from_leaves(
            bal.get("entitlements", []),
            (bal.get("next_entitlement") or {}).get("date"),
            leaves_by_pid.get(p["id"], []),
            today,
        )
        if tdc.get("status") == "advance_ok":
            pids.add(p["id"])
    return pids


@api.get("/personnel")
async def list_personnel(
    q: Optional[str] = None, departman: Optional[str] = None,
    sirket: Optional[str] = None, aktif: Optional[bool] = None,
    sort_by: str = "ad_soyad", sort_dir: str = "asc",
    limit: Optional[int] = None, skip: int = 0,
    consent_advance: bool = False,
    _: dict = Depends(get_current_user),
):
    """Personel listesi. Yeni: sort_by/sort_dir/limit/skip server-side. Tüm listeye ihtiyaç varsa limit=None.
    consent_advance=True → sadece ten_day_check.status='advance_ok' olan personeller (Sarı Rozet).
    """
    filt = _build_personnel_filter(q, departman, sirket, aktif)
    if consent_advance:
        allow_ids = list(await _advance_ok_pids())
        filt = {**filt, "id": {"$in": allow_ids}} if allow_ids else {"id": "__none__"}
    allowed_sort = {"ad_soyad", "sicil_no", "departman", "sirket", "ise_giris", "aktif"}
    field = sort_by if sort_by in allowed_sort else "ad_soyad"
    direction = -1 if str(sort_dir).lower() == "desc" else 1
    cursor = db.personnel.find(filt, {"_id": 0}).sort(field, direction)
    if skip:
        cursor = cursor.skip(int(skip))
    if limit is not None:
        cursor = cursor.limit(max(1, min(int(limit), 5000)))
    items = []
    async for p in cursor:
        items.append(p)
    return items


@api.get("/personnel/count")
async def personnel_count(q: Optional[str] = None, departman: Optional[str] = None,
                           sirket: Optional[str] = None, aktif: Optional[bool] = None,
                           consent_advance: bool = False,
                           _: dict = Depends(get_current_user)):
    filt = _build_personnel_filter(q, departman, sirket, aktif)
    if consent_advance:
        allow_ids = list(await _advance_ok_pids())
        filt = {**filt, "id": {"$in": allow_ids}} if allow_ids else {"id": "__none__"}
    return {"total": await db.personnel.count_documents(filt)}


@api.get("/personnel/facets")
async def personnel_facets(include_inactive: bool = False,
                            _: dict = Depends(get_current_user)):
    """Departman + şirket + görev listeleri.
    include_inactive=True ise pasif personelin değerleri de dahil edilir
    (otomatik tamamlama için tavsiye edilir — mevcut tüm departmanlar önerilsin)."""
    filt = {} if include_inactive else {"aktif": True}
    deps = await db.personnel.distinct("departman", filt)
    sirks = await db.personnel.distinct("sirket", filt)
    gorevs = await db.personnel.distinct("gorev", filt)
    deps = sorted([d for d in deps if d])
    sirks = sorted([s for s in sirks if s])
    gorevs = sorted([g for g in gorevs if g])
    return {"departments": deps, "companies": sirks, "roles": gorevs}

@api.post("/personnel")
async def create_personnel(body: PersonnelIn, request: Request, current: dict = Depends(require_roles("admin", "hr"))):
    normalized = _normalize_sicil(body.sicil_no)
    if not normalized:
        raise HTTPException(status_code=400, detail="Sicil numarası zorunlu")
    body.sicil_no = normalized
    if await db.personnel.find_one({"sicil_no": normalized}):
        raise HTTPException(status_code=400, detail=f"Sicil numarası '{normalized}' zaten kayıtlı")
    p = Personnel(**body.model_dump())
    d = p.model_dump()
    d["updated_at"] = d.get("created_at")
    await db.personnel.insert_one(d)
    await _audit(action="create", module="personnel", entity_type="personnel",
                 entity_id=p.id, entity_name=p.ad_soyad,
                 new_values=body.model_dump(),
                 description=f"Personel eklendi: {p.ad_soyad} ({p.sicil_no})",
                 request=request, user=current)
    return {k: v for k, v in d.items() if k != "_id"}

@api.get("/personnel/{pid}")
async def get_personnel(pid: str, _: dict = Depends(get_current_user)):
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    return p

@api.put("/personnel/{pid}")
async def update_personnel(pid: str, body: PersonnelIn, request: Request, current: dict = Depends(require_roles("admin", "hr"))):
    existing = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    if_match = request.headers.get("If-Match")
    current_updated_at = existing.get("updated_at") or existing.get("created_at")
    if if_match and current_updated_at and if_match != current_updated_at:
        raise HTTPException(status_code=412, detail={
            "message": "Bu personel kaydı başka bir kullanıcı tarafından değiştirildi. Lütfen sayfayı yenileyip tekrar deneyin.",
            "current_updated_at": current_updated_at,
        })
    new_data = body.model_dump()
    new_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.personnel.update_one({"id": pid}, {"$set": new_data})
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    changed_old, changed_new = _dict_diff(existing, body.model_dump())
    action = "update"
    desc = f"Personel güncellendi: {p.get('ad_soyad')}"
    if existing.get("aktif") and not new_data.get("aktif"):
        action = "terminate"
        desc = f"İşten ayrılış: {p.get('ad_soyad')} — çıkış: {new_data.get('isten_cikis') or '—'}"
    elif not existing.get("aktif") and new_data.get("aktif"):
        action = "reactivate"
        desc = f"Aktife alındı: {p.get('ad_soyad')}"
    await _audit(action=action, module="personnel", entity_type="personnel",
                 entity_id=pid, entity_name=p.get("ad_soyad"),
                 old_values=changed_old, new_values=changed_new,
                 description=desc, request=request, user=current)
    return p

@api.delete("/personnel/{pid}")
async def delete_personnel(pid: str, request: Request, current: dict = Depends(require_roles("admin"))):
    existing = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    await db.leaves.delete_many({"personnel_id": pid})
    await db.personnel.delete_one({"id": pid})
    await _audit(action="delete", module="personnel", entity_type="personnel",
                 entity_id=pid, entity_name=existing.get("ad_soyad"),
                 old_values={k: existing.get(k) for k in ("sicil_no", "ad_soyad", "departman", "gorev")},
                 description=f"Personel silindi: {existing.get('ad_soyad')} ({existing.get('sicil_no')})",
                 request=request, user=current)
    return {"ok": True}


class PersonnelDeleteIn(BaseModel):
    password: str
    reason: str

@api.get("/personnel/{pid}/delete-preview")
async def personnel_delete_preview(pid: str, _: dict = Depends(require_roles("admin"))):
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    leaves_count = await db.leaves.count_documents({"personnel_id": pid})
    ent_count = await db.entitlements.count_documents({"personnel_id": pid})
    return {
        "personnel": {"id": p["id"], "sicil_no": p.get("sicil_no"),
                       "ad_soyad": p.get("ad_soyad"), "departman": p.get("departman"),
                       "ise_giris": p.get("ise_giris"), "aktif": p.get("aktif", True)},
        "leaves_count": leaves_count,
        "entitlements_count": ent_count,
    }

@api.post("/personnel/{pid}/delete")
async def personnel_hard_delete(pid: str, body: PersonnelDeleteIn, request: Request,
                                 current: dict = Depends(require_roles("admin"))):
    """Yönetici şifresi + gerekçe ile 2 aşamalı kalıcı silme. Audit korunur."""
    if not body.password or not body.reason or not body.reason.strip():
        raise HTTPException(status_code=400, detail="Yönetici şifresi ve silme gerekçesi zorunlu")
    admin = await db.users.find_one({"id": current["id"]})
    if not admin or not verify_password(body.password, admin.get("password_hash", "")):
        await _audit(action="delete_failed", module="personnel", entity_type="personnel",
                     entity_id=pid, entity_name=None,
                     description=f"Personel silme reddedildi (şifre doğrulanamadı) — gerekçe: {body.reason[:200]}",
                     request=request, user=current, success=False)
        raise HTTPException(status_code=403, detail="Yönetici şifresi doğrulanamadı. Personel silinmedi.")
    existing = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    try:
        leaves_removed = (await db.leaves.delete_many({"personnel_id": pid})).deleted_count
        ent_removed = (await db.entitlements.delete_many({"personnel_id": pid})).deleted_count
        await db.personnel.delete_one({"id": pid})
    except Exception as ex:
        await _audit(action="delete_failed", module="personnel", entity_type="personnel",
                     entity_id=pid, entity_name=existing.get("ad_soyad"),
                     description=f"Silme sırasında hata: {ex}", request=request, user=current, success=False)
        raise HTTPException(status_code=500, detail="Silme sırasında hata oluştu. Kayıtlar korundu.")
    # T.C. kimlik gibi hassas alanları audit'te maskele (son 4 hane hariç)
    tc = existing.get("tc_no") or ""
    tc_masked = ("*" * max(0, len(tc) - 4) + tc[-4:]) if tc else ""
    await _audit(action="hard_delete", module="personnel", entity_type="personnel",
                 entity_id=pid, entity_name=existing.get("ad_soyad"),
                 old_values={
                     "sicil_no": existing.get("sicil_no"),
                     "ad_soyad": existing.get("ad_soyad"),
                     "departman": existing.get("departman"),
                     "gorev": existing.get("gorev"),
                     "ise_giris": existing.get("ise_giris"),
                     "tc_no": tc_masked,
                     "leaves_removed": leaves_removed,
                     "entitlements_removed": ent_removed,
                 },
                 new_values={"reason": body.reason.strip()},
                 description=f"Personel KALICI silindi: {existing.get('ad_soyad')} ({existing.get('sicil_no')}) — gerekçe: {body.reason.strip()[:200]}",
                 request=request, user=current)
    return {"ok": True, "leaves_removed": leaves_removed, "entitlements_removed": ent_removed}

# -----------------------------------------------------------------------------
# Leave balance
# -----------------------------------------------------------------------------
def _parse_date(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except Exception:
        return None

def _days_for_seniority(total_seniority: int, age_at: Optional[float]) -> tuple:
    """4857 sayılı İş Kanunu m.53 — Yıllık ücretli izin süresi:
       - Toplam izin kıdemi 1..5 yıl (5 dahil)  → 14 gün
       - 5 < toplam < 15                       → 20 gün
       - toplam ≥ 15                            → 26 gün
       Yaş istisnası: hak ediş tarihinde yaşı 18 ve daha küçük VEYA 50 ve daha büyük
       olan işçilere yıllık izin süresi 20 günden az olamaz.
       Personel lehine yüksek olan uygulanır.

       Tuple döner: (final_days, base_days_by_seniority, age_min_days).
    """
    if total_seniority <= 5:
        base = 14
    elif total_seniority < 15:
        base = 20
    else:
        base = 26
    # Yaş kuralı: "18 veya daha küçük" → age < 19 (yaşı 18'i tamamlamamış yani 18 dahil)
    # "50 veya daha büyük" → age >= 50.
    age_days = 20 if (age_at is not None and (age_at < 19 or age_at >= 50)) else 0
    return max(base, age_days), base, age_days

async def _compute_entitlements(personnel: dict, as_of: Optional[date] = None) -> dict:
    """Yeni hak ediş sistemi:
       - İlk hak ediş = son işe giriş + 1 yıl
       - Önceki kıdem geçmiş tarihe yansıtılmaz, sadece hak ediş gününde toplam kıdeme eklenir
       - Her hak ediş kaydı immutable — DB'ye yazıldıktan sonra parametreler değişse de bozulmaz
    """
    hire = _parse_date(personnel["ise_giris"])
    empty = {"entitled_total": 0, "used_total": 0, "remaining": 0,
             "entitlements": [], "next_entitlement": None,
             "new_period_years": 0, "total_seniority": 0}
    if not hire:
        return empty
    today = as_of or date.today()
    birth = _parse_date(personnel.get("dogum_tarihi"))
    prev_years = int(personnel.get("onceki_kidem_yil") or 0)

    existing = {}
    async for e in db.entitlements.find({"personnel_id": personnel["id"]}, {"_id": 0}):
        existing[e["date"]] = e

    entitlements: list = []
    y = 1
    while y <= 60:
        anniv = _safe_anniv(hire, y)
        if anniv > today:
            break
        iso = anniv.isoformat()
        if iso in existing:
            entitlements.append(existing[iso])
        else:
            new_period = y
            total = prev_years + new_period
            age_at = ((anniv - birth).days / 365.25) if birth else None
            days, base, age_days = _days_for_seniority(total, age_at)
            expl = f"Önceki kıdem {prev_years} yıl + Yeni dönem {new_period} yıl = {total} yıl → {base} gün"
            if age_days > base:
                expl += f". Yaş kuralı ({age_days} gün) uygulandı."
            rec = {
                "id": str(uuid.uuid4()),
                "personnel_id": personnel["id"],
                "date": iso,
                "prev_years": prev_years,
                "new_period_years": new_period,
                "total_seniority": total,
                "age_at": int(age_at) if age_at is not None else None,
                "days": days,
                "explanation": expl,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            try:
                await db.entitlements.insert_one(rec.copy())
            except Exception:
                pass
            entitlements.append(rec)
        y += 1

    # Bir sonraki hak ediş
    next_anniv = _safe_anniv(hire, y)
    new_period_next = y
    total_next = prev_years + new_period_next
    age_at_next = ((next_anniv - birth).days / 365.25) if birth else None
    next_days, base_n, age_n = _days_for_seniority(total_next, age_at_next)
    next_info = {
        "date": next_anniv.isoformat(),
        "days": next_days,
        "total_seniority": total_next,
        "new_period_years": new_period_next,
    }

    entitled_total = sum(x["days"] for x in entitlements)
    used = 0.0
    async for L in db.leaves.find({"personnel_id": personnel["id"]}, {"_id": 0, "days": 1}):
        used += float(L.get("days", 0))
    remaining = entitled_total - used

    # Yeni çalışma döneminde tamamlanan hizmet yılı
    new_period_completed = len(entitlements)
    total_seniority_now = prev_years + new_period_completed

    return {
        "entitled_total": entitled_total,
        "used_total": used,
        "remaining": remaining,
        "entitlements": sorted(entitlements, key=lambda x: x["date"]),
        "next_entitlement": next_info,
        "new_period_years": new_period_completed,
        "total_seniority": total_seniority_now,
        "prev_years": prev_years,
        "hire_date": hire.isoformat(),
        "last_entitlement_date": entitlements[-1]["date"] if entitlements else None,
    }

async def _balance_for(personnel: dict, as_of: Optional[date] = None) -> dict:
    """Legacy shape for older callers — wraps new engine."""
    r = await _compute_entitlements(personnel, as_of)
    return {
        "entitled_total": r["entitled_total"],
        "used_total": r["used_total"],
        "remaining": r["remaining"],
        "per_year": [{"year_index": i + 1, "granted_on": e["date"], "days": e["days"]}
                     for i, e in enumerate(r["entitlements"])],
    }


async def _recompute_entitlements_for(personnel: dict) -> dict:
    """Admin bakım işlemi (UI'da düğme YOK — sadece backend/dahili çağrı):
       Kişinin tüm entitlement kayıtlarını yeniden üretir. Eski kayıtlar tarih başına
       aynı ID ile üzerine yazılır; ise_giris/doğum/önceki_kidem güncellenmişse
       satırlar yeni değerlere göre yeniden yazılır. Kullanılmış izinler (leaves)
       KESİNLİKLE silinmez, dokunulmaz.
    """
    pid = personnel["id"]
    # Öncesini raporla
    old_docs = []
    async for e in db.entitlements.find({"personnel_id": pid}, {"_id": 0}):
        old_docs.append(e)
    old_total = round(sum(float(x.get("days", 0) or 0) for x in old_docs), 2)

    # Sıfırla ve yeniden hesapla (mevcut _compute_entitlements FIFO satırları üretip DB'ye yazar)
    await db.entitlements.delete_many({"personnel_id": pid})
    fresh = await _compute_entitlements(personnel)

    new_ents = fresh.get("entitlements", [])
    return {
        "personnel_id": pid,
        "sicil_no": personnel.get("sicil_no"),
        "ad_soyad": personnel.get("ad_soyad"),
        "ise_giris": personnel.get("ise_giris"),
        "dogum_tarihi": personnel.get("dogum_tarihi"),
        "onceki_kidem_yil": int(personnel.get("onceki_kidem_yil") or 0),
        "old_count": len(old_docs),
        "old_total": old_total,
        "new_count": len(new_ents),
        "new_total": round(sum(float(x.get("days", 0) or 0) for x in new_ents), 2),
        "used_total": fresh.get("used_total", 0),
        "remaining": fresh.get("remaining", 0),
        "entitlements": [
            {
                "entitlement_date": e["date"],
                "previous_seniority": e.get("prev_years"),
                "new_period_years": e.get("new_period_years"),
                "total_seniority": e.get("total_seniority"),
                "age_at_entitlement": e.get("age_at"),
                "entitlement_days": e.get("days"),
                "explanation": e.get("explanation"),
            } for e in new_ents
        ],
    }

def _valid_ymd(y, m, d):
    try:
        date(y, m, d); return True
    except Exception:
        return False

def _safe_anniv(hire: date, years: int) -> date:
    # handle Feb 29
    try:
        return date(hire.year + years, hire.month, hire.day)
    except ValueError:
        return date(hire.year + years, hire.month, 28)

@api.get("/personnel/{pid}/balance")
async def personnel_balance(pid: str, _: dict = Depends(get_current_user)):
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    detail = await _compute_entitlements(p)
    return {"personnel": p, "balance": detail}

@api.post("/personnel/{pid}/recalculate")
async def recalculate_entitlements(pid: str, _: dict = Depends(require_roles("admin", "hr"))):
    """Kişinin bekleyen kayıtlarını temizler; yeni parametrelerle yeniden üretir.
    Kullanım: kıdem veya işe giriş sonradan düzeltildiğinde admin manuel çalıştırır.
    """
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    await db.entitlements.delete_many({"personnel_id": pid})
    detail = await _compute_entitlements(p)
    return {"ok": True, "balance": detail}

@api.get("/personnel/{pid}/izin-cetveli")
async def izin_cetveli(pid: str, _: dict = Depends(get_current_user)):
    """Yıllık İzin Cetveli: yıllara göre hak ediş, kullanım ve FIFO dağıtım satırları."""
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    bal = await _compute_entitlements(p)
    ents = sorted(bal["entitlements"], key=lambda x: x["date"])
    all_leaves = []
    async for L in db.leaves.find({"personnel_id": pid}, {"_id": 0}).sort("start_date", 1):
        all_leaves.append(L)
    allocations = await allocate_leaves_fifo(ents, all_leaves, personnel=p)
    rows = []
    cum_entitled = 0.0
    cum_used = 0.0
    for i, e in enumerate(ents):
        start = e["date"]
        end = ents[i + 1]["date"] if i + 1 < len(ents) else "9999-12-31"
        year_leaves = [L for L in all_leaves if start <= L["start_date"] < end]
        used = sum(L["days"] for L in year_leaves)
        cum_entitled += e["days"]
        cum_used += used
        rows.append({
            "yil_no": e["new_period_years"],
            "hak_edis_tarihi": e["date"],
            "gecerlilik_bitis": (ents[i + 1]["date"] if i + 1 < len(ents) else None),
            "onceki_kidem": e["prev_years"],
            "yeni_donem": e["new_period_years"],
            "toplam_kidem": e["total_seniority"],
            "yas": e["age_at"],
            "hak_edilen": e["days"],
            "aciklama": e["explanation"],
            "kullanilan": used,
            "kalan_yil_ici": e["days"] - used,
            "kumule_hak": cum_entitled,
            "kumule_kullanim": cum_used,
            "kumule_kalan": cum_entitled - cum_used,
            "leaves": [{
                "start_date": L["start_date"], "end_date": L["end_date"],
                "days": L["days"], "izin_turu": L.get("izin_turu", ""),
                "aciklama": L.get("aciklama", ""),
            } for L in year_leaves],
        })
    return {
        "personnel": p,
        "balance": {
            "entitled_total": bal["entitled_total"],
            "used_total": bal["used_total"],
            "remaining": bal["remaining"],
            "prev_years": bal["prev_years"],
            "new_period_years": bal["new_period_years"],
            "total_seniority": bal["total_seniority"],
            "next_entitlement": bal["next_entitlement"],
            "hire_date": bal["hire_date"],
        },
        "rows": rows,
        "allocations": allocations,
        "form_meta": {"dokuman_no": "İK.FR.09", "duzenleme_tarihi": "02.01.2023", "revizyon_no": "01"},
    }

# -----------------------------------------------------------------------------
# Leaves
# -----------------------------------------------------------------------------
@api.post("/leaves/preview")
async def leaves_preview(body: LeaveIn, _: dict = Depends(get_current_user)):
    s = _parse_date(body.start_date); e = _parse_date(body.end_date)
    if not s or not e or e < s:
        raise HTTPException(status_code=400, detail="Geçersiz tarih aralığı")
    return await calc_leave_days(s, e)


# -----------------------------------------------------------------------------
# Admin bakım — entitlement yeniden hesaplama (UI'da düğme YOK).
# 4857 sayılı İş Kanunu m.53 kurallarına göre entitlements koleksiyonundaki
# eski/hatalı satırları temizler ve yeniden yazar. İzin (leaves) kayıtlarına
# dokunulmaz. Sadece admin.
# -----------------------------------------------------------------------------
@api.post("/personnel/{pid}/entitlements/recompute")
async def recompute_person_entitlements(pid: str, _: dict = Depends(require_roles("admin"))):
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    return await _recompute_entitlements_for(p)

@api.post("/personnel/entitlements/recompute-all")
async def recompute_all_entitlements(_: dict = Depends(require_roles("admin"))):
    results = []
    async for p in db.personnel.find({}, {"_id": 0}):
        r = await _recompute_entitlements_for(p)
        results.append({
            "id": r["personnel_id"], "sicil_no": r["sicil_no"], "ad_soyad": r["ad_soyad"],
            "old_total": r["old_total"], "new_total": r["new_total"],
            "delta": round(r["new_total"] - r["old_total"], 2),
            "used_total": r["used_total"], "remaining": r["remaining"],
        })
    return {"count": len(results), "results": results,
             "total_delta": round(sum(x["delta"] for x in results), 2)}


async def _resolve_person_ids_for_search(q: Optional[str] = None,
                                          departman: Optional[str] = None,
                                          sirket: Optional[str] = None) -> Optional[List[str]]:
    """q/departman/sirket ile personel eşleşen ID listesi. Filtre yoksa None döner (unbounded)."""
    if not (q or departman or sirket):
        return None
    filt: dict = {}
    if departman:
        filt["departman"] = {"$regex": f"^{re.escape(departman)}$", "$options": "i"}
    if sirket:
        filt["sirket"] = {"$regex": f"^{re.escape(sirket)}$", "$options": "i"}
    if q:
        esc_tr = _tr_search_regex(q)
        esc_other = re.escape(q.strip())
        filt["$or"] = [
            {"ad_soyad": {"$regex": esc_tr, "$options": "i"}},
            {"sicil_no": {"$regex": esc_other, "$options": "i"}},
            {"departman": {"$regex": esc_tr, "$options": "i"}},
        ]
    ids: List[str] = []
    async for p in db.personnel.find(filt, {"_id": 0, "id": 1}):
        ids.append(p["id"])
    return ids


async def _build_leaves_filter(personnel_id, start, end, recent_days, izin_turu,
                                 q, departman, sirket,
                                 personnel_active: Optional[bool] = None) -> Optional[dict]:
    """None dönerse filtre boş sonuç anlamına gelir (personel eşleşmedi).
    Iter 47: personnel_active=True → yalnız aktif personellerin izinleri,
    False → yalnız işten ayrılanların izinleri, None → hepsi.
    """
    filt: dict = {}
    if personnel_id:
        filt["personnel_id"] = personnel_id
    else:
        matched = await _resolve_person_ids_for_search(q, departman, sirket)
        active_matched = None
        if personnel_active is not None:
            ids = []
            async for p in db.personnel.find({"aktif": bool(personnel_active)}, {"_id": 0, "id": 1}):
                ids.append(p["id"])
            active_matched = ids
        # Kombine et: matched ∩ active_matched
        combined = None
        if matched is not None and active_matched is not None:
            combined = list(set(matched) & set(active_matched))
        elif matched is not None:
            combined = matched
        elif active_matched is not None:
            combined = active_matched
        if combined is not None:
            if not combined:
                return None  # boş sonuç
            filt["personnel_id"] = {"$in": combined}
    if recent_days and not start and not end:
        cutoff = (date.today() - timedelta(days=int(recent_days))).isoformat()
        filt["start_date"] = {"$gte": cutoff}
    elif start or end:
        rng: dict = {}
        if start: rng["$gte"] = start
        if end: rng["$lte"] = end
        filt["start_date"] = rng
    if izin_turu:
        filt["izin_turu"] = izin_turu
    return filt


@api.get("/leaves")
async def list_leaves(personnel_id: Optional[str] = None,
                       start: Optional[str] = None, end: Optional[str] = None,
                       recent_days: Optional[int] = None,
                       izin_turu: Optional[str] = None,
                       q: Optional[str] = None,
                       departman: Optional[str] = None,
                       sirket: Optional[str] = None,
                       personnel_active: Optional[bool] = None,
                       sort_by: str = "start_date", sort_dir: str = "desc",
                       limit: int = 100, skip: int = 0,
                       include_isbasi: bool = True,
                       include_consent: bool = False,
                       _: dict = Depends(get_current_user)):
    """İzin listesi. Server-side filtre + sort + pagination.
    Iter 47: personnel_active True/False → aktif/işten ayrılan personel izinleri.
    """
    filt = await _build_leaves_filter(personnel_id, start, end, recent_days,
                                       izin_turu, q, departman, sirket,
                                       personnel_active=personnel_active)
    if filt is None:
        return []
    allowed_sort = {"start_date", "end_date", "days", "created_at", "izin_turu"}
    field = sort_by if sort_by in allowed_sort else "start_date"
    direction = -1 if str(sort_dir).lower() == "desc" else 1
    limit = max(1, min(int(limit), 5000))
    skip = max(0, int(skip))
    items = []
    async for L in db.leaves.find(filt, {"_id": 0}).sort(field, direction).skip(skip).limit(limit):
        items.append(L)

    # personel bilgisi tek sorguda (batch)
    pids = list({L["personnel_id"] for L in items})
    pmap: dict = {}
    if pids:
        async for p in db.personnel.find({"id": {"$in": pids}},
                                          {"_id": 0, "id": 1, "sicil_no": 1, "ad_soyad": 1, "departman": 1, "sirket": 1}):
            pmap[p["id"]] = p
    for L in items:
        p = pmap.get(L["personnel_id"], {})
        L["sicil_no"] = p.get("sicil_no")
        L["ad_soyad"] = p.get("ad_soyad")
        L["_departman"] = p.get("departman")
        L["_sirket"] = p.get("sirket")
        # Iter 47: personelin aktif durumu + işten çıkış tarihi (İşten Ayrılanların İzinleri sekmesi için)
        L["_personnel_active"] = p.get("aktif", True)
        L["_personnel_isten_cikis"] = p.get("isten_cikis")

    if include_isbasi and items:
        holidays = await get_all_holidays()
        for L in items:
            try:
                end_d = date.fromisoformat(L["end_date"])
                L["isbasi_tarihi"] = _next_working_day(end_d, holidays).isoformat()
            except Exception:
                L["isbasi_tarihi"] = None

    if include_consent:
        per_person_ents: dict = {}
        for L in items:
            pid = L["personnel_id"]
            if pid not in per_person_ents:
                ents = []
                async for e in db.entitlements.find({"personnel_id": pid}, {"_id": 0, "date": 1, "days": 1}):
                    ents.append(e)
                ents.sort(key=lambda x: x["date"])
                per_person_ents[pid] = ents
        cum_used: dict = {}
        for L in sorted(items, key=lambda x: x["start_date"]):
            pid = L["personnel_id"]
            prior_used = cum_used.get(pid, 0.0)
            entitled_so_far = sum(e["days"] for e in per_person_ents.get(pid, []) if e["date"] <= L["start_date"])
            over = (prior_used + L["days"]) - entitled_so_far
            # Iter 60: 0 günlük izin girildiğinde de bakiye 0/negatif ise muvafakatname zorunlu
            # olur (yer-tutucu kayıt senaryosu). Bakiye pozitif ise 0 günlük izin muvafakatname
            # gerektirmez.
            zero_day_no_balance = (L["days"] == 0) and ((entitled_so_far - prior_used) <= 0)

            # Muvafakatname uyarisi sadece mevcut takvim yilindaki izinlerde gosterilir.
            # Eski izinler bakiye ve kullanilan izin hesabinda kalmaya devam eder.
            try:
                leave_year = date.fromisoformat((L.get("start_date") or "")[:10]).year
            except Exception:
                leave_year = None
            is_current_year_leave = leave_year == date.today().year

            L["consent_required"] = (
                is_current_year_leave
                and ((over > 0) or zero_day_no_balance)
            )

            # Eski yillardaki izinler icin muvafakatname/avans gunu gosterme.
            L["consent_advance_days"] = (
                round(min(float(L["days"]), max(0.0, over)), 2)
                if is_current_year_leave
                else 0
            )

            cum_used[pid] = prior_used + L["days"]
    return items


@api.get("/leaves/count")
async def leaves_count(personnel_id: Optional[str] = None,
                        start: Optional[str] = None, end: Optional[str] = None,
                        recent_days: Optional[int] = None,
                        izin_turu: Optional[str] = None,
                        q: Optional[str] = None,
                        departman: Optional[str] = None,
                        sirket: Optional[str] = None,
                        personnel_active: Optional[bool] = None,
                        _: dict = Depends(get_current_user)):
    filt = await _build_leaves_filter(personnel_id, start, end, recent_days,
                                       izin_turu, q, departman, sirket,
                                       personnel_active=personnel_active)
    if filt is None:
        return {"total": 0}
    total = await db.leaves.count_documents(filt)
    return {"total": total}


@api.get("/leaves/facets")
async def leaves_facets(_: dict = Depends(get_current_user)):
    """Filtre dropdown'ları için izin türü listesi."""
    turler = await db.leaves.distinct("izin_turu")
    turler = sorted([t for t in turler if t])
    return {"izin_turleri": turler}


class BulkTarget(BaseModel):
    type: Literal["all", "department", "selected"]
    department: Optional[str] = None
    personnel_ids: Optional[List[str]] = None

class BulkLeaveIn(BaseModel):
    target: BulkTarget
    start_date: str
    end_date: str
    izin_turu: str = "Yıllık İzin"
    aciklama: str = ""

async def _resolve_bulk_targets(target: BulkTarget) -> List[dict]:
    """Aktif personel filtresi hedef türüne göre uygulanır."""
    q: dict = {"aktif": True}
    if target.type == "department":
        if not target.department:
            raise HTTPException(status_code=400, detail="Departman seçilmedi")
        # tr-locale casefold uyumu için bulunan tüm eşleşen departmanları getir
        dep_upper = target.department.strip()
        q["departman"] = {"$regex": f"^{dep_upper}$", "$options": "i"}
    elif target.type == "selected":
        ids = target.personnel_ids or []
        if not ids:
            raise HTTPException(status_code=400, detail="Personel seçilmedi")
        q["id"] = {"$in": ids}
    # "all" → q sadece aktif=True
    people: List[dict] = []
    async for p in db.personnel.find(q, {"_id": 0}):
        people.append(p)
    people.sort(key=lambda x: (x.get("sicil_no") or "", x.get("ad_soyad") or ""))
    return people

async def _bulk_preview_row(p: dict, s: date, e: date, izin_turu: str) -> dict:
    warnings: List[dict] = []
    calc = await calc_leave_days(s, e)
    computed_days = calc["days"]
    # Çakışma kontrolü — aynı personel, tarih aralığı kesişen mevcut kayıt
    overlap = await db.leaves.find_one({
        "personnel_id": p["id"],
        "start_date": {"$lte": e.isoformat()},
        "end_date": {"$gte": s.isoformat()},
    }, {"_id": 0})
    conflict = None
    if overlap:
        conflict = {
            "existing_start": overlap["start_date"],
            "existing_end": overlap["end_date"],
            "existing_izin_turu": overlap.get("izin_turu", ""),
        }
        warnings.append({"level": "error", "code": "conflict",
                          "message": f"Çakışma: {overlap['start_date']} → {overlap['end_date']}"})
    # Bakiye
    bal = await _balance_for(p)
    remaining_after = float(bal["remaining"]) - float(computed_days)
    izin_norm = izin_turu.strip().casefold().replace("i̇", "i")
    if "yıllık" in izin_norm or "yillik" in izin_norm:
        if remaining_after < 0:
            warnings.append({"level": "warning", "code": "advance",
                              "message": f"Avans kullanım ({remaining_after:.1f} gün) — muvafakatname gerekli"})
        elif remaining_after < 10:
            warnings.append({"level": "warning", "code": "low_balance",
                              "message": f"Kalan bakiye 10 günün altına düşecek ({remaining_after:.1f})"})
    if computed_days <= 0:
        warnings.append({"level": "error", "code": "zero_days",
                          "message": "Hesaplanan izin 0 gün — hafta sonu/tatil"})
    can_apply = not any(w["level"] == "error" for w in warnings)
    return {
        "personnel_id": p["id"],
        "sicil_no": p.get("sicil_no", ""),
        "ad_soyad": p.get("ad_soyad", ""),
        "departman": p.get("departman", ""),
        "entitled_total": bal["entitled_total"],
        "used_total": bal["used_total"],
        "remaining": bal["remaining"],
        "computed_days": computed_days,
        "remaining_after": remaining_after,
        "conflict": conflict,
        "warnings": warnings,
        "can_apply": can_apply,
    }

@api.post("/leaves/bulk/preview")
async def leaves_bulk_preview(body: BulkLeaveIn, _: dict = Depends(require_roles("admin", "hr"))):
    s = _parse_date(body.start_date); e = _parse_date(body.end_date)
    if not s or not e or e < s:
        raise HTTPException(status_code=400, detail="Geçersiz tarih aralığı")
    people = await _resolve_bulk_targets(body.target)
    rows = []
    for p in people:
        rows.append(await _bulk_preview_row(p, s, e, body.izin_turu))
    return rows

@api.post("/leaves/bulk")
async def leaves_bulk_create(body: BulkLeaveIn, background_tasks: BackgroundTasks,
                              request: Request, user: dict = Depends(require_roles("admin", "hr"))):
    s = _parse_date(body.start_date); e = _parse_date(body.end_date)
    if not s or not e or e < s:
        raise HTTPException(status_code=400, detail="Geçersiz tarih aralığı")
    people = await _resolve_bulk_targets(body.target)
    created: List[dict] = []
    skipped: List[dict] = []
    for p in people:
        row = await _bulk_preview_row(p, s, e, body.izin_turu)
        if not row["can_apply"]:
            skipped.append({
                "personnel_id": p["id"],
                "sicil_no": p.get("sicil_no", ""),
                "ad_soyad": p.get("ad_soyad", ""),
                "reason": "; ".join(w["message"] for w in row["warnings"] if w["level"] == "error"),
            })
            continue
        calc = await calc_leave_days(s, e)
        rec = LeaveRecord(
            personnel_id=p["id"],
            start_date=body.start_date, end_date=body.end_date,
            days=calc["days"], izin_turu=body.izin_turu, aciklama=body.aciklama,
            created_by=user["id"],
        )
        await db.leaves.insert_one(rec.model_dump())
        await _invalidate_cetvel_for(p["id"])  # Iter 37
        created.append({
            "leave_id": rec.id, "personnel_id": p["id"],
            "sicil_no": p.get("sicil_no", ""), "ad_soyad": p.get("ad_soyad", ""),
            "days": calc["days"],
        })
        if p.get("email"):
            subject = f"Yıllık İzin Bildirimi — {p['ad_soyad']}"
            html = f"""<div style="font-family:Arial;padding:20px"><h3>Merkoteks — Yıllık İzin Kaydı</h3><p><b>{p['ad_soyad']}</b> için {body.start_date} → {body.end_date} arası {calc['days']} günlük {body.izin_turu} kaydı toplu izin işlemi ile oluşturuldu.</p></div>"""
            background_tasks.add_task(_send_email, p["email"], subject, html)
    await _audit(action="bulk_create", module="leaves", entity_type="leave",
                 entity_id=None, entity_name=f"Toplu izin: {len(created)} kayıt",
                 new_values={"target": body.target.model_dump(), "start_date": body.start_date,
                              "end_date": body.end_date, "izin_turu": body.izin_turu,
                              "created_count": len(created), "skipped_count": len(skipped)},
                 description=f"Toplu izin oluşturuldu: {len(created)} kayıt, {len(skipped)} atlandı",
                 request=request, user=user)
    return {"created": created, "skipped": skipped, "total_target": len(people)}

# -----------------------------------------------------------------------------
# Excel-based bulk leave upload
# -----------------------------------------------------------------------------
BULK_EXCEL_HEADERS = ["Sicil Numarası", "Adı Soyadı", "Departman", "İzin Türü",
                      "İzin Başlangıç Tarihi", "İzin Bitiş Tarihi",
                      "İzin Gün Sayısı", "Açıklama"]

def _normalize_sicil(v) -> str:
    """Sicil numarasını Excel'in float ('1001.0') / int / str / boşluk varyantlarından
    tek bir kanonik string'e çevirir. Baştaki sıfırlar (sadece string ise) korunur.
    """
    if v is None:
        return ""
    # Excel int-as-float senaryosu: 1001.0 → "1001" (gerçek ondalık ise kesme)
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v).strip()
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    # ".0" son eki (metin de olabilir): "1001.0" → "1001", ama "1001.5" gibi ondalıklar korunur
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _parse_excel_days(v):
    """Excel'den gün sayısını ayrıştır. Boş/None → None. '20,5' veya '20.5' → 20.5.
    Geçersiz metin → 'INVALID' string döner."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return "INVALID"

def _parse_excel_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None

@api.get("/leaves/bulk/excel-template")
async def leaves_bulk_excel_template(_: dict = Depends(require_roles("admin", "hr"))):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Toplu İzin"
    ws["A1"] = "Merkoteks — Toplu İzin Yükleme Şablonu"
    ws["A1"].font = Font(bold=True, size=14, color="1D4ED8")
    ws.merge_cells("A1:H1")
    ws["A2"] = ("Sicil Numarası zorunlu. Adı Soyadı ve Departman bilgi amaçlıdır (eşleşme kontrolü yapılır). "
                "İzin Türü boş bırakılırsa 'Yıllık İzin' varsayılır. "
                "Tarihler GG.AA.YYYY (örn. 05.08.2026) formatında olmalıdır. "
                "İzin Gün Sayısı isteğe bağlıdır — boş bırakılırsa sistem otomatik hesaplar.")
    ws.merge_cells("A2:H2")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"].font = Font(size=10, italic=True, color="475569")
    ws.row_dimensions[2].height = 46
    thin = Side(style="thin", color="CBD5E1")
    for i, h in enumerate(BULK_EXCEL_HEADERS, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="1D4ED8")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[4].height = 32
    # Örnek satırlar — 0,5 günlük örnek dahil
    examples = [
        ["0001", "ÖRNEK PERSONEL", "KESİMHANE", "Yıllık İzin", "05.08.2026", "15.08.2026", "", "İzin süresi otomatik hesaplanır"],
        ["0002", "İKİNCİ ÖRNEK", "MODELHANE", "Yıllık İzin", "28.10.2026", "28.10.2026", "0,5", "Sadece arife günü — 0,5 gün"],
        ["0003", "ÜÇÜNCÜ ÖRNEK", "DİKİMHANE", "Yıllık İzin", "01.10.2026", "30.10.2026", "20,5", "Excel'de girilen değer sistem hesabı ile kontrol edilir"],
    ]
    for r_i, row in enumerate(examples, start=5):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r_i, column=i, value=v)
            c.font = Font(italic=True, color="94A3B8", size=10)
            c.alignment = Alignment(horizontal="center" if i in (1, 5, 6, 7) else "left", vertical="center")
    widths = [14, 28, 22, 18, 18, 18, 14, 32]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = w
    ws.freeze_panes = "A5"

    # AÇIKLAMALAR sayfası
    ws2 = wb.create_sheet("AÇIKLAMALAR")
    ws2["A1"] = "AÇIKLAMALAR"
    ws2["A1"].font = Font(bold=True, size=16, color="1D4ED8")
    aciklamalar = [
        ("Sicil Numarası", "Zorunlu. Sistemde tanımlı personelin sicil numarası. Eşleşme bulunamazsa satır hatalı sayılır."),
        ("Adı Soyadı", "Bilgi amaçlıdır. Sistemdeki ad-soyad ile karşılaştırılır; farklıysa satırda uyarı gösterilir."),
        ("Departman", "Bilgi amaçlıdır. Personelin sistemdeki departmanı ile karşılaştırılır; farklıysa uyarı gösterilir."),
        ("İzin Türü", "Örn: 'Yıllık İzin', 'İdari İzin', 'Ücretsiz İzin', 'Mazeret İzni'. Boş bırakılırsa 'Yıllık İzin' varsayılır."),
        ("İzin Başlangıç Tarihi", "Zorunlu. GG.AA.YYYY formatında (örn. 05.08.2026)."),
        ("İzin Bitiş Tarihi", "Zorunlu. GG.AA.YYYY formatında ve başlangıçtan büyük veya eşit olmalı."),
        ("İzin Gün Sayısı", "İsteğe bağlıdır. Boş bırakılırsa sistem hafta sonu, resmî tatil, arife ve şirket tatillerini dikkate alarak otomatik hesaplar. Girilen değer sistem hesabından farklıysa yükleme öncesinde uyarı gösterilir. Ondalıklı değerler virgülle yazılabilir (0,5 / 1,5 / 20,5)."),
        ("Açıklama", "İsteğe bağlıdır. İzin kaydına iliştirilecek serbest metin."),
    ]
    for i, (h, d) in enumerate(aciklamalar, start=3):
        c1 = ws2.cell(row=i, column=1, value=h)
        c1.font = Font(bold=True, size=11, color="1D4ED8")
        c1.alignment = Alignment(vertical="top")
        c2 = ws2.cell(row=i, column=2, value=d)
        c2.font = Font(size=10, color="334155")
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[i].height = 42
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 100

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="toplu_izin_sablonu.xlsx"'},
    )


async def _parse_bulk_leave_excel(content: bytes) -> List[dict]:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı")

    # Uygun sheet'i bul: header'ında "sicil" içeren ilk sheet
    ws = None
    for sheet_name in wb.sheetnames:
        candidate = wb[sheet_name]
        found = False
        for ri in range(1, min(16, (candidate.max_row or 0) + 1)):
            row_vals = [str(candidate.cell(row=ri, column=c).value or "").strip().lower()
                        for c in range(1, min((candidate.max_column or 0) + 1, 20))]
            has_sicil = any(("sicil" in v and len(v) < 40) for v in row_vals)
            has_date = any(("başlangı" in v or "baslangi" in v or "bitiş" in v or "bitis" in v) for v in row_vals)
            if has_sicil and has_date:
                found = True; break
        if found:
            ws = candidate; break
    if ws is None:
        ws = wb.active

    header_row_idx = None
    header_map: dict = {}
    max_scan = min(15, ws.max_row or 15)
    for ri in range(1, max_scan + 1):
        vals = [(str(ws.cell(row=ri, column=c).value or "").strip())
                for c in range(1, (ws.max_column or 8) + 1)]
        # Talimat/açıklama satırlarını atla: header hücreleri kısa olmalı (< 40 karakter)
        short = [v for v in vals if v and len(v) < 40]
        low = [v.lower() for v in short]
        has_sicil = any("sicil" in v for v in low)
        has_date_hdr = any(("başlangı" in v or "baslangi" in v or "bitiş" in v or "bitis" in v) for v in low)
        if has_sicil and has_date_hdr:
            header_row_idx = ri
            for c, v in enumerate(vals, start=1):
                if not v or len(v) >= 40:
                    continue
                key = v.lower().strip()
                if not key:
                    continue
                if "sicil" in key: header_map["sicil_no"] = c
                elif "ad" in key and ("soyad" in key or "isim" in key or key == "adı"): header_map["ad_soyad"] = c
                elif "departman" in key or "bölüm" in key: header_map["departman"] = c
                elif "tür" in key or "turu" in key: header_map["izin_turu"] = c
                elif "başlangı" in key or "baslangi" in key: header_map["start_date"] = c
                elif "bitiş" in key or "bitis" in key: header_map["end_date"] = c
                elif "gün" in key or "gun sayı" in key or "sayısı" in key: header_map["days_excel"] = c
                elif "açıklama" in key or "aciklama" in key: header_map["aciklama"] = c
            break
    if header_row_idx is None or "sicil_no" not in header_map \
       or "start_date" not in header_map or "end_date" not in header_map:
        raise HTTPException(
            status_code=400,
            detail="Başlık satırı bulunamadı. Zorunlu kolonlar: Sicil Numarası, İzin Başlangıç Tarihi, İzin Bitiş Tarihi",
        )

    parsed = []
    for ri in range(header_row_idx + 1, (ws.max_row or header_row_idx) + 1):
        sicil_raw = ws.cell(row=ri, column=header_map["sicil_no"]).value
        sicil = _normalize_sicil(sicil_raw)
        ad_soyad = str(ws.cell(row=ri, column=header_map["ad_soyad"]).value or "").strip() if "ad_soyad" in header_map else ""
        # Örnek şablon satırlarını atla
        if sicil.startswith("000") and ad_soyad.startswith("ÖR"):
            continue
        if sicil in {"0001", "0002", "0003"} and "ÖRNEK" in ad_soyad.upper():
            continue
        departman = str(ws.cell(row=ri, column=header_map["departman"]).value or "").strip() if "departman" in header_map else ""
        izin_turu = str(ws.cell(row=ri, column=header_map["izin_turu"]).value or "").strip() if "izin_turu" in header_map else ""
        if not izin_turu:
            izin_turu = "Yıllık İzin"
        s_raw = ws.cell(row=ri, column=header_map["start_date"]).value
        e_raw = ws.cell(row=ri, column=header_map["end_date"]).value
        days_raw = ws.cell(row=ri, column=header_map["days_excel"]).value if "days_excel" in header_map else None
        aciklama_row = str(ws.cell(row=ri, column=header_map["aciklama"]).value or "").strip() if "aciklama" in header_map else ""
        # Boş sicil dahil TÜM satırlar dönsün — ön izlemede "Eksik Sicil" statüsü ile listelensin.
        # Ama tamamen boş bir satırı (hiçbir hücre dolu değilse) atla.
        if not any([sicil, ad_soyad, s_raw, e_raw]):
            continue
        parsed.append({
            "row": ri, "sicil_no": sicil, "ad_soyad_excel": ad_soyad,
            "departman_excel": departman, "izin_turu": izin_turu,
            "start_raw": s_raw, "end_raw": e_raw,
            "days_excel_raw": days_raw, "aciklama_row": aciklama_row,
        })
    if not parsed:
        raise HTTPException(status_code=400, detail="Dosyada geçerli satır bulunamadı")
    return parsed


async def _validate_bulk_excel_row(entry: dict) -> dict:
    warnings: List[dict] = []
    sicil = _normalize_sicil(entry.get("sicil_no"))
    izin_turu = entry["izin_turu"] or "Yıllık İzin"
    s = _parse_excel_date(entry["start_raw"])
    e = _parse_excel_date(entry["end_raw"])
    row_out = {
        "row": entry["row"],
        "sicil_no": sicil,
        "ad_soyad_excel": entry.get("ad_soyad_excel", ""),
        "departman_excel": entry.get("departman_excel", ""),
        "izin_turu": izin_turu,
        "start_date": s.isoformat() if s else (str(entry["start_raw"] or "")),
        "end_date": e.isoformat() if e else (str(entry["end_raw"] or "")),
        "personnel_id": None, "matched_ad_soyad": None, "departman": None,
        "entitled_total": 0.0, "used_total": 0.0, "remaining": 0.0,
        "computed_days": 0.0, "remaining_after": 0.0,
        "excel_days": None, "excel_days_valid": True,
        "days_diff": 0.0, "final_days": 0.0,
        "days_status": "auto",
        "aciklama_row": entry.get("aciklama_row", ""),
        "warnings": warnings, "can_apply": False,
        "status_code": "unknown",   # eksik_sicil | personel_bulunamadi | pasif | gecersiz_tarih | cakisan_izin | yetersiz_bakiye | esleti
        "status_label": "",
        "force_block": False,       # frontend "İşleme Dahil Et" seçeneğini tamamen kilitler
    }

    # 1) EKSİK SİCİL — kesin engel
    if not sicil:
        warnings.append({"level": "error", "code": "missing_sicil",
                         "message": "Sicil numarası eksik"})
        row_out["status_code"] = "eksik_sicil"
        row_out["status_label"] = "Eksik Sicil Numarası"
        row_out["force_block"] = True
        return row_out

    p = await db.personnel.find_one({"sicil_no": sicil}, {"_id": 0})

    # 2) PERSONEL BULUNAMADI — kesin engel, alternatif eşleştirme yok
    if not p:
        warnings.append({"level": "error", "code": "no_match",
                         "message": f"Sicil No {sicil} sistemde bulunamadı. Bu satır işleme alınmayacaktır."})
        row_out["status_code"] = "personel_bulunamadi"
        row_out["status_label"] = "Personel Bulunamadı"
        row_out["force_block"] = True
        return row_out

    row_out["personnel_id"] = p["id"]
    row_out["matched_ad_soyad"] = p.get("ad_soyad", "")
    row_out["departman"] = p.get("departman", "")

    # 3) PASİF PERSONEL — kesin engel
    if not p.get("aktif", True):
        warnings.append({"level": "error", "code": "inactive",
                         "message": "Personel pasif / işten ayrılmış"})
        row_out["status_code"] = "pasif"
        row_out["status_label"] = "Pasif Personel"
        row_out["force_block"] = True
        return row_out

    # 4) GEÇERSİZ TARİH
    if not s or not e:
        warnings.append({"level": "error", "code": "bad_date",
                         "message": "Tarih okunamadı (GG.AA.YYYY olmalı)"})
        row_out["status_code"] = "gecersiz_tarih"
        row_out["status_label"] = "Geçersiz Tarih"
        return row_out
    if e < s:
        warnings.append({"level": "error", "code": "date_order",
                         "message": "Bitiş başlangıçtan önce"})
        row_out["status_code"] = "gecersiz_tarih"
        row_out["status_label"] = "Geçersiz Tarih"
        return row_out

    # 5) Ad-soyad uyarısı (yalnızca uyarı — sicil doğruysa)
    if entry.get("ad_soyad_excel"):
        a = entry["ad_soyad_excel"].strip().casefold()
        b = (p.get("ad_soyad") or "").strip().casefold()
        if a and b and a != b:
            warnings.append({"level": "warning", "code": "name_mismatch",
                             "message": f"Excel'deki ad soyad ile sistem kaydı farklı (DB: {p.get('ad_soyad', '')})"})
    if entry.get("departman_excel"):
        da = entry["departman_excel"].strip().casefold()
        db_dep = (p.get("departman") or "").strip().casefold()
        if da and db_dep and da != db_dep:
            warnings.append({"level": "warning", "code": "dept_mismatch",
                             "message": f"Departman eşleşmiyor (DB: {p.get('departman', '')})"})

    calc = await calc_leave_days(s, e)
    computed_days = float(calc["days"])
    row_out["computed_days"] = computed_days

    excel_days_raw = entry.get("days_excel_raw")
    parsed_days = _parse_excel_days(excel_days_raw)
    if parsed_days == "INVALID":
        row_out["excel_days_valid"] = False
        row_out["excel_days"] = None
        row_out["days_status"] = "invalid"
        warnings.append({"level": "error", "code": "days_invalid",
                         "message": f"'İzin Gün Sayısı' geçersiz: {excel_days_raw!r}"})
    elif parsed_days is None:
        row_out["days_status"] = "auto"
        row_out["excel_days"] = None
        row_out["final_days"] = computed_days
    else:
        row_out["excel_days"] = parsed_days
        if parsed_days <= 0:
            row_out["days_status"] = "invalid"
            row_out["excel_days_valid"] = False
            warnings.append({"level": "error", "code": "days_nonpositive",
                             "message": "İzin gün sayısı 0 veya negatif olamaz"})
        else:
            diff = round(parsed_days - computed_days, 2)
            row_out["days_diff"] = diff
            if abs(diff) < 1e-9:
                row_out["days_status"] = "match"
                row_out["final_days"] = computed_days
            else:
                row_out["days_status"] = "mismatch"
                row_out["final_days"] = computed_days
                warnings.append({"level": "warning", "code": "days_mismatch",
                                 "message": f"Excel'de girilen izin gün sayısı ({_fmt_tr_num(parsed_days)}) "
                                            f"ile sistem hesabı ({_fmt_tr_num(computed_days)}) farklıdır."})

    if row_out["days_status"] == "auto":
        row_out["final_days"] = computed_days

    # 6) ÇAKIŞMA
    overlap = await db.leaves.find_one({
        "personnel_id": p["id"],
        "start_date": {"$lte": e.isoformat()},
        "end_date": {"$gte": s.isoformat()},
    }, {"_id": 0})
    if overlap:
        warnings.append({"level": "error", "code": "conflict",
                         "message": f"Çakışma: {overlap['start_date']} → {overlap['end_date']}"})
        row_out["status_code"] = "cakisan_izin"
        row_out["status_label"] = "Çakışan İzin"

    bal = await _balance_for(p)
    row_out["entitled_total"] = bal["entitled_total"]
    row_out["used_total"] = bal["used_total"]
    row_out["remaining"] = bal["remaining"]
    remaining_after = float(bal["remaining"]) - float(row_out["final_days"])
    row_out["remaining_after"] = remaining_after
    izin_norm = izin_turu.strip().casefold().replace("i̇", "i")
    if "yıllık" in izin_norm or "yillik" in izin_norm:
        if remaining_after < 0:
            warnings.append({"level": "warning", "code": "advance",
                             "message": f"Avans kullanım ({remaining_after:.1f} gün)"})
        elif remaining_after < 10:
            warnings.append({"level": "warning", "code": "low_balance",
                             "message": f"Kalan bakiye 10 gün altına ({remaining_after:.1f})"})
    if computed_days <= 0:
        warnings.append({"level": "error", "code": "zero_days",
                         "message": "0 gün — hafta sonu/tatil"})
        if row_out["status_code"] == "unknown":
            row_out["status_code"] = "gecersiz_tarih"
            row_out["status_label"] = "Geçersiz Tarih (0 gün)"

    row_out["can_apply"] = not any(w["level"] == "error" for w in warnings)
    if row_out["can_apply"] and row_out["status_code"] == "unknown":
        row_out["status_code"] = "esleti"
        row_out["status_label"] = "Eşleşti — Kayda Hazır"
    elif not row_out["can_apply"] and row_out["status_code"] == "unknown":
        row_out["status_code"] = "hatali"
        row_out["status_label"] = "Hatalı"
    return row_out


@api.post("/leaves/bulk/historical-import")
async def leaves_bulk_historical_import(file: UploadFile = File(...),
                                          request: Request = None,
                                          user: dict = Depends(require_roles("admin"))):
    """Toplu tarihsel izin aktarımı — çok sayıda geçmiş izin kaydını tek seferde işler.
    Kural: Excel'de girilen gün varsa onu kullan (admin yetkisiyle), aksi halde sistem hesaplar.
    Sicil eşleşmeyen, tarih hatalı veya çakışma olan satırlar atlanır. Preview YOK — direkt insert.
    Sonuç audit_log'a `historical_import` action olarak yazılır.
    """
    content = await file.read()
    try:
        entries = await _parse_bulk_leave_excel(content)
    except HTTPException as e:
        raise e

    created: List[dict] = []
    skipped: List[dict] = []
    year_stats: dict = {}

    # Personel eşleştirme için sicil→id cache
    sicil_to_p: dict = {}
    async for p in db.personnel.find({}, {"_id": 0}):
        sicil_to_p[str(p.get("sicil_no", "")).strip()] = p

    # Kolektif çakışma kontrolü için mevcut leaves'ı topla (personnel_id → [(s,e)])
    existing_by_pid: dict = {}
    async for L in db.leaves.find({}, {"_id": 0, "personnel_id": 1, "start_date": 1, "end_date": 1}):
        existing_by_pid.setdefault(L["personnel_id"], []).append((L["start_date"], L["end_date"]))

    for entry in entries:
        row_idx = entry["row"]
        sicil = str(entry["sicil_no"]).strip()
        p = sicil_to_p.get(sicil)
        if not p:
            skipped.append({"row": row_idx, "sicil_no": sicil,
                            "reason": f"Sicil {sicil} bulunamadı"})
            continue
        s = _parse_excel_date(entry["start_raw"])
        e = _parse_excel_date(entry["end_raw"])
        if not s or not e or e < s:
            skipped.append({"row": row_idx, "sicil_no": sicil,
                            "ad_soyad": p.get("ad_soyad", ""),
                            "reason": "Geçersiz tarih"})
            continue
        s_iso, e_iso = s.isoformat(), e.isoformat()
        # Çakışma kontrolü (memory)
        pid = p["id"]
        conflict = False
        for (xs, xe) in existing_by_pid.get(pid, []):
            if s_iso <= xe and xs <= e_iso:
                conflict = True; break
        if conflict:
            skipped.append({"row": row_idx, "sicil_no": sicil,
                            "ad_soyad": p.get("ad_soyad", ""),
                            "reason": f"Çakışma: {s_iso} → {e_iso}"})
            continue
        # Gün: Excel varsa onu kullan, yoksa hesapla
        parsed_days = _parse_excel_days(entry.get("days_excel_raw"))
        if parsed_days == "INVALID" or parsed_days is None or (isinstance(parsed_days, (int, float)) and parsed_days <= 0):
            calc = await calc_leave_days(s, e)
            used_days = float(calc["days"])
            src = "system"
        else:
            used_days = float(parsed_days)
            src = "excel"
        if used_days <= 0:
            skipped.append({"row": row_idx, "sicil_no": sicil,
                            "ad_soyad": p.get("ad_soyad", ""),
                            "reason": "0 gün (hafta sonu/tatil)"})
            continue
        rec = LeaveRecord(
            personnel_id=pid,
            start_date=s_iso, end_date=e_iso,
            days=used_days,
            izin_turu=entry.get("izin_turu") or "Yıllık İzin",
            aciklama=entry.get("aciklama_row") or "Tarihsel içe aktarım",
            created_by=user["id"],
        )
        await db.leaves.insert_one(rec.model_dump())
        await _invalidate_cetvel_for(pid)  # Iter 37
        existing_by_pid.setdefault(pid, []).append((s_iso, e_iso))  # sonraki satırlar için
        yr = s.year
        year_stats[yr] = year_stats.get(yr, 0) + 1
        created.append({"row": row_idx, "sicil_no": sicil,
                        "ad_soyad": p.get("ad_soyad", ""),
                        "start_date": s_iso, "end_date": e_iso,
                        "days": used_days, "source": src})

    # History kaydı
    history_id = str(uuid.uuid4())
    await db.bulk_upload_history.insert_one({
        "id": history_id, "uploaded_by": user["id"],
        "uploaded_by_name": user.get("name") or user.get("email") or "",
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "filename": file.filename or "historical.xlsx",
        "aciklama": "Tarihsel toplu içe aktarım",
        "kind": "historical_import",
        "total_rows": len(entries),
        "created_count": len(created), "skipped_count": len(skipped),
        "total_days": round(sum(c["days"] for c in created), 2),
        "created": created[:2000], "skipped": skipped[:1000],
    })
    await _audit(
        action="historical_import", module="leaves", entity_type="leave",
        entity_id=history_id,
        entity_name=f"Tarihsel toplu içe aktarım: {len(created)} kayıt",
        new_values={"filename": file.filename, "created_count": len(created),
                    "skipped_count": len(skipped),
                    "years": sorted(year_stats.keys()),
                    "year_counts": year_stats},
        description=f"Tarihsel toplu içe aktarım: {len(created)} oluşturuldu, "
                    f"{len(skipped)} atlandı, yıllar: {sorted(year_stats.keys())}",
        request=request, user=user,
    )
    return {"created_count": len(created), "skipped_count": len(skipped),
            "total_rows": len(entries), "year_counts": year_stats,
            "history_id": history_id, "created_sample": created[:5],
            "skipped_sample": skipped[:20]}


@api.post("/leaves/bulk/excel-preview")
async def leaves_bulk_excel_preview(file: UploadFile = File(...),
                                     _: dict = Depends(require_roles("admin", "hr"))):
    content = await file.read()
    entries = await _parse_bulk_leave_excel(content)
    rows = []
    for entry in entries:
        rows.append(await _validate_bulk_excel_row(entry))
    # Dosya-içi çakışma tespiti
    for i, r in enumerate(rows):
        if not r.get("personnel_id"):
            continue
        s_i = _parse_date(r["start_date"]); e_i = _parse_date(r["end_date"])
        if not s_i or not e_i:
            continue
        for j in range(i):
            r2 = rows[j]
            if r2.get("personnel_id") != r["personnel_id"]:
                continue
            s_j = _parse_date(r2["start_date"]); e_j = _parse_date(r2["end_date"])
            if s_j and e_j and s_i <= e_j and s_j <= e_i:
                r["warnings"].append({"level": "error", "code": "file_dup",
                                       "message": f"Dosya içinde {r2['row']}. satırla çakışma"})
                r["can_apply"] = False
                break
    summary = {
        "total": len(rows),
        "applicable": sum(1 for r in rows if r["can_apply"]),
        "blocked": sum(1 for r in rows if not r["can_apply"]),
        "total_days": round(sum(r["computed_days"] for r in rows if r["can_apply"]), 2),
    }
    return {"rows": rows, "summary": summary}


class ExcelConfirmRow(BaseModel):
    sicil_no: str
    izin_turu: str = "Yıllık İzin"
    start_date: str
    end_date: str
    days_choice: str = "system"   # "system" | "excel" | "skip"
    excel_days: Optional[float] = None
    override_reason: str = ""
    aciklama: str = ""


class ExcelConfirmIn(BaseModel):
    rows: List[ExcelConfirmRow]
    aciklama: str = ""
    filename: str = ""

@api.post("/leaves/bulk/excel-confirm")
async def leaves_bulk_excel_confirm(body: ExcelConfirmIn, background_tasks: BackgroundTasks,
                                     request: Request,
                                     user: dict = Depends(require_roles("admin", "hr"))):
    if not body.rows:
        raise HTTPException(status_code=400, detail="Satır yok")
    created: List[dict] = []
    skipped: List[dict] = []
    for idx, r in enumerate(body.rows, start=1):
        if (r.days_choice or "system").lower() == "skip":
            skipped.append({"sicil_no": r.sicil_no, "ad_soyad": "",
                            "reason": "Kullanıcı tarafından işlem dışı bırakıldı"})
            continue
        entry = {
            "row": idx, "sicil_no": r.sicil_no, "ad_soyad_excel": "",
            "departman_excel": "", "izin_turu": r.izin_turu or "Yıllık İzin",
            "start_raw": r.start_date, "end_raw": r.end_date,
            "days_excel_raw": r.excel_days, "aciklama_row": r.aciklama,
        }
        v = await _validate_bulk_excel_row(entry)
        if not v["can_apply"]:
            skipped.append({
                "sicil_no": r.sicil_no,
                "ad_soyad": v.get("matched_ad_soyad") or "",
                "reason": "; ".join(w["message"] for w in v["warnings"] if w["level"] == "error"),
            })
            continue
        s = _parse_date(v["start_date"]); e = _parse_date(v["end_date"])
        calc = await calc_leave_days(s, e)
        # Gün seçimi
        choice = (r.days_choice or "system").lower()
        used_days = float(calc["days"])
        override_note = ""
        if choice == "excel":
            if user.get("role") != "admin":
                skipped.append({"sicil_no": r.sicil_no,
                                "ad_soyad": v.get("matched_ad_soyad") or "",
                                "reason": "Excel değerini kullanma yetkisi yok (yalnızca Yönetici)"})
                continue
            if not r.override_reason or not r.override_reason.strip():
                skipped.append({"sicil_no": r.sicil_no,
                                "ad_soyad": v.get("matched_ad_soyad") or "",
                                "reason": "Excel değeri kullanılırken gerekçe zorunludur"})
                continue
            if r.excel_days is None or r.excel_days <= 0:
                skipped.append({"sicil_no": r.sicil_no,
                                "ad_soyad": v.get("matched_ad_soyad") or "",
                                "reason": "Excel gün değeri geçersiz"})
                continue
            used_days = float(r.excel_days)
            override_note = r.override_reason.strip()
            await _audit(
                action="excel_days_override", module="leaves", entity_type="leave",
                entity_id=None,
                entity_name=f"{v.get('matched_ad_soyad')} — Excel gün üzerine yazma",
                new_values={"sicil_no": r.sicil_no, "start_date": v["start_date"],
                            "end_date": v["end_date"], "system_days": calc["days"],
                            "excel_days": r.excel_days, "reason": override_note},
                description=f"Excel değeri sistem hesabının yerine kullanıldı. "
                            f"Sistem: {calc['days']}, Excel: {r.excel_days}. Gerekçe: {override_note}",
                request=request, user=user,
            )
        rec = LeaveRecord(
            personnel_id=v["personnel_id"],
            start_date=v["start_date"], end_date=v["end_date"],
            days=used_days, izin_turu=v["izin_turu"],
            aciklama=(r.aciklama or body.aciklama or "").strip(),
            created_by=user["id"],
        )
        await db.leaves.insert_one(rec.model_dump())
        await _invalidate_cetvel_for(v["personnel_id"])  # Iter 37
        created.append({
            "leave_id": rec.id, "sicil_no": r.sicil_no,
            "ad_soyad": v["matched_ad_soyad"], "days": used_days,
            "start_date": v["start_date"], "end_date": v["end_date"],
            "izin_turu": v["izin_turu"],
            "source": "excel" if choice == "excel" else "system",
            "override_reason": override_note or None,
        })
        p = await db.personnel.find_one({"id": v["personnel_id"]}, {"_id": 0})
        if p and p.get("email"):
            subject = f"Yıllık İzin Bildirimi — {p['ad_soyad']}"
            html = (f"""<div style="font-family:Arial;padding:20px"><h3>Merkoteks — Yıllık İzin Kaydı</h3>"""
                    f"""<p><b>{p['ad_soyad']}</b> için {v['start_date']} → {v['end_date']} arası """
                    f"""{used_days} günlük {v['izin_turu']} kaydı Excel toplu yükleme ile oluşturuldu.</p></div>""")
            background_tasks.add_task(_send_email, p["email"], subject, html)
    # Upload history kaydı
    history_id = str(uuid.uuid4())
    filename = getattr(body, "filename", "") or "toplu_izin.xlsx"
    total_days = sum(float(c.get("days") or 0) for c in created)
    hist_doc = {
        "id": history_id,
        "uploaded_by": user["id"],
        "uploaded_by_name": user.get("name") or user.get("email") or "",
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "filename": filename,
        "aciklama": body.aciklama or "",
        "total_rows": len(body.rows),
        "created_count": len(created),
        "skipped_count": len(skipped),
        "total_days": round(total_days, 2),
        "created": created,
        "skipped": skipped,
    }
    await db.bulk_upload_history.insert_one(hist_doc)
    await _audit(action="bulk_excel_create", module="leaves", entity_type="leave",
                 entity_id=history_id, entity_name=f"Excel toplu izin: {len(created)} kayıt",
                 new_values={"created_count": len(created), "skipped_count": len(skipped),
                              "aciklama": body.aciklama, "filename": filename},
                 description=f"Excel'den toplu izin: {len(created)} oluşturuldu, {len(skipped)} atlandı",
                 request=request, user=user)
    return {"created": created, "skipped": skipped, "total": len(body.rows),
            "history_id": history_id}


@api.get("/bulk-uploads/history")
async def bulk_upload_history_list(limit: int = 100,
                                    _: dict = Depends(require_roles("admin", "hr"))):
    docs = []
    async for d in db.bulk_upload_history.find({}, {"_id": 0}).sort("uploaded_at", -1).limit(int(limit)):
        docs.append(d)
    return docs


class ErrorReportRow(BaseModel):
    row: Optional[int] = None
    sicil_no: str = ""
    ad_soyad: str = ""
    status_label: str = ""
    reason: str = ""


class ErrorReportIn(BaseModel):
    rows: List[ErrorReportRow]
    filename: str = ""


@api.post("/leaves/bulk/excel-error-report")
async def leaves_bulk_excel_error_report(body: ErrorReportIn,
                                           _: dict = Depends(require_roles("admin", "hr"))):
    """Ön izlemede engellenen veya atlanan satırları xlsx olarak indir."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Atlanan Kayıtlar"
    headers = ["Excel Satır No", "Sicil Numarası", "Adı Soyadı", "Durum", "Hata Nedeni / Açıklama"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="B91C1C")
        c.alignment = Alignment(horizontal="center")
    r = 2
    for row in body.rows:
        ws.cell(row=r, column=1, value=row.row or "")
        ws.cell(row=r, column=2, value=row.sicil_no)
        ws.cell(row=r, column=3, value=row.ad_soyad)
        ws.cell(row=r, column=4, value=row.status_label)
        ws.cell(row=r, column=5, value=row.reason)
        r += 1
    for col, w in enumerate([14, 18, 32, 26, 60], 1):
        ws.column_dimensions[chr(64 + col)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    stem = (body.filename or "toplu_izin.xlsx").rsplit(".", 1)[0]
    fn = f"{stem}_atlanan_kayitlar.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@api.get("/bulk-uploads/history/{hid}")
async def bulk_upload_history_detail(hid: str,
                                      _: dict = Depends(require_roles("admin", "hr"))):
    d = await db.bulk_upload_history.find_one({"id": hid}, {"_id": 0})
    if not d:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    return d


@api.post("/admin/process-advance-offsets")
async def process_advance_offsets(request: Request,
                                   user: dict = Depends(require_roles("admin"))):
    """Bekleyen avans izinleri için mahsup taraması: her personel için FIFO allocation'ları
    çalıştırır ve leave.start_date < entitlement.date olan alokasyonları 'mahsup' olarak
    audit_log'a yazar. Idempotent (hash ile tekrar yazmaz)."""
    offsets_logged = 0
    async for p in db.personnel.find({}, {"_id": 0}):
        bal = await _compute_entitlements(p)
        ents = bal.get("entitlements", [])
        leaves = []
        async for L in db.leaves.find({"personnel_id": p["id"]}, {"_id": 0}).sort("start_date", 1):
            leaves.append(L)
        allocs = await allocate_leaves_fifo(ents, leaves)
        for a in allocs:
            ent_date = a.get("entitlement_date") or ""
            leave_start = a.get("start_date") or ""
            if not ent_date or not leave_start:
                continue
            # Avans: iznin başlangıcı hak ediş tarihinden ÖNCE ise
            if leave_start < ent_date and not a.get("is_advance"):
                key = f"{p['id']}|{a.get('leave_id')}|{ent_date}"
                exists = await db.mahsup_log.find_one({"key": key}, {"_id": 0})
                if exists:
                    continue
                await db.mahsup_log.insert_one({
                    "id": str(uuid.uuid4()), "key": key,
                    "personnel_id": p["id"], "personnel_ad_soyad": p.get("ad_soyad"),
                    "leave_id": a.get("leave_id"),
                    "entitlement_date": ent_date,
                    "entitlement_year": a.get("entitlement_year"),
                    "days": a.get("days"),
                    "leave_start_date": leave_start,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                })
                await _audit(
                    action="advance_offset", module="leaves", entity_type="leave",
                    entity_id=a.get("leave_id"),
                    entity_name=f"{p.get('ad_soyad')} — Avans Mahsup",
                    new_values={"days": a.get("days"), "entitlement_date": ent_date,
                                 "entitlement_year": a.get("entitlement_year"),
                                 "leave_start_date": leave_start},
                    description=f"{p.get('ad_soyad')} için {a.get('days')} gün avans izin "
                                f"{ent_date} hak edişinden mahsup edildi",
                    request=request, user=user,
                )
                offsets_logged += 1
    return {"offsets_logged": offsets_logged}



@api.post("/leaves")
async def create_leave(body: LeaveIn, request: Request, background_tasks: BackgroundTasks, user: dict = Depends(require_roles("admin", "hr"))):
    res = await _upsert_leave(body, background_tasks, user, exclude_id=None, request=request)
    # Iter 37: yeni izin → cetvel bayrağı düşer
    await _invalidate_cetvel_for(body.personnel_id)
    return res

@api.put("/leaves/{lid}")
async def update_leave(lid: str, body: LeaveIn, background_tasks: BackgroundTasks, request: Request, user: dict = Depends(require_roles("admin", "hr"))):
    existing = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="İzin kaydı bulunamadı")
    if_match = request.headers.get("If-Match")
    current_updated_at = existing.get("updated_at") or existing.get("created_at")
    if if_match and current_updated_at and if_match != current_updated_at:
        raise HTTPException(status_code=412, detail={
            "message": "Bu kayıt başka bir kullanıcı tarafından değiştirildi. Lütfen sayfayı yenileyip tekrar deneyin.",
            "current_updated_at": current_updated_at,
        })
    return await _upsert_leave(body, background_tasks, user, exclude_id=lid, existing=existing, request=request)

@api.get("/leaves/single/{lid}")
async def get_leave(lid: str, _: dict = Depends(get_current_user)):
    L = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not L:
        raise HTTPException(status_code=404, detail="İzin kaydı bulunamadı")
    return L

async def _upsert_leave(body: LeaveIn, background_tasks: BackgroundTasks, user: dict,
                         exclude_id: Optional[str], existing: Optional[dict] = None,
                         request: Optional[Request] = None):
    p = await db.personnel.find_one({"id": body.personnel_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    # Iter 47: İşten ayrılmış personele yeni izin kaydı oluşturulamaz (create only)
    if exclude_id is None and p.get("aktif") is False:
        raise HTTPException(status_code=400, detail="İşten ayrılmış personele yeni izin kaydı oluşturulamaz.")
    s = _parse_date(body.start_date); e = _parse_date(body.end_date)
    if not s or not e or e < s:
        raise HTTPException(status_code=400, detail="Geçersiz tarih aralığı")

    # Sadece AYNI personel için, kendi ID hariç, tarih aralığı kesişen kayıt aranır
    q: dict = {
        "personnel_id": body.personnel_id,
        "start_date": {"$lte": body.end_date},
        "end_date": {"$gte": body.start_date},
    }
    if exclude_id:
        q["id"] = {"$ne": exclude_id}
    overlap = await db.leaves.find_one(q, {"_id": 0})
    if overlap:
        # Kesişen tarihleri hesapla
        os_ = max(_parse_date(overlap["start_date"]), s)
        oe_ = min(_parse_date(overlap["end_date"]), e)
        overlap_days = []
        cur = os_
        while cur <= oe_:
            overlap_days.append(cur.isoformat())
            cur += timedelta(days=1)
        raise HTTPException(status_code=409, detail={
            "message": "Bu personelin seçilen tarih aralığında mevcut bir izin kaydı bulunmaktadır.",
            "personnel_ad_soyad": p.get("ad_soyad", ""),
            "personnel_id": p["id"],
            "existing_id": overlap["id"],
            "existing_start": overlap["start_date"],
            "existing_end": overlap["end_date"],
            "existing_izin_turu": overlap.get("izin_turu", ""),
            "new_start": body.start_date,
            "new_end": body.end_date,
            "overlap_dates": overlap_days,
        })

    calc = await calc_leave_days(s, e)
    if existing:
        await db.leaves.update_one({"id": exclude_id}, {"$set": {
            "start_date": body.start_date, "end_date": body.end_date,
            "days": calc["days"], "izin_turu": body.izin_turu, "aciklama": body.aciklama,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": user["id"],
        }})
        rec = {**existing, "start_date": body.start_date, "end_date": body.end_date,
               "days": calc["days"], "izin_turu": body.izin_turu, "aciklama": body.aciklama}
        await _audit(action="update", module="leaves", entity_type="leave",
                     entity_id=exclude_id, entity_name=p.get("ad_soyad"),
                     old_values={"start_date": existing.get("start_date"), "end_date": existing.get("end_date"),
                                  "days": existing.get("days"), "izin_turu": existing.get("izin_turu"),
                                  "aciklama": existing.get("aciklama")},
                     new_values={"start_date": body.start_date, "end_date": body.end_date,
                                  "days": calc["days"], "izin_turu": body.izin_turu,
                                  "aciklama": body.aciklama},
                     description=f"İzin güncellendi: {p.get('ad_soyad')} — {body.start_date} → {body.end_date}",
                     request=request, user=user)
        return {**rec, "breakdown": calc["breakdown"], "notified": []}

    rec = LeaveRecord(
        personnel_id=body.personnel_id,
        start_date=body.start_date, end_date=body.end_date,
        days=calc["days"], izin_turu=body.izin_turu, aciklama=body.aciklama,
        created_by=user["id"],
    )
    await db.leaves.insert_one(rec.model_dump())
    recipients = []
    if p.get("email"): recipients.append(p["email"])
    admin_email = os.environ.get("ADMIN_EMAIL")
    if admin_email: recipients.append(admin_email.lower())
    if recipients:
        subject = f"Yıllık İzin Bildirimi — {p['ad_soyad']}"
        html = f"""<div style="font-family:Arial;padding:20px"><h3>Merkoteks — Yıllık İzin Kaydı</h3><p><b>{p['ad_soyad']}</b> için {body.start_date} → {body.end_date} arası {calc['days']} günlük {body.izin_turu} kaydı oluşturuldu.</p></div>"""
        for r in recipients:
            background_tasks.add_task(_send_email, r, subject, html)
    await _audit(action="create", module="leaves", entity_type="leave",
                 entity_id=rec.id, entity_name=p.get("ad_soyad"),
                 new_values={"start_date": body.start_date, "end_date": body.end_date,
                              "days": calc["days"], "izin_turu": body.izin_turu,
                              "aciklama": body.aciklama, "personnel_id": p["id"]},
                 description=f"İzin eklendi: {p.get('ad_soyad')} — {body.start_date} → {body.end_date} ({calc['days']} gün)",
                 request=request, user=user)
    return {**rec.model_dump(), "breakdown": calc["breakdown"], "notified": recipients}
    rec = LeaveRecord(
        personnel_id=body.personnel_id,
        start_date=body.start_date, end_date=body.end_date,
        days=calc["days"], izin_turu=body.izin_turu, aciklama=body.aciklama,
        created_by=user["id"],
    )
    await db.leaves.insert_one(rec.model_dump())
    # E-posta bildirimleri (arka planda)
    recipients = []
    if p.get("email"):
        recipients.append(p["email"])
    admin_email = os.environ.get("ADMIN_EMAIL")
    if admin_email:
        recipients.append(admin_email.lower())
    if recipients:
        subject = f"Yıllık İzin Bildirimi — {p['ad_soyad']}"
        html = f"""
        <table role="presentation" style="font-family: Arial, sans-serif; max-width:600px; width:100%; background:#f8fafc; padding:24px;">
          <tr><td>
            <div style="background:#1d4ed8;color:#fff;padding:16px 20px;border-radius:8px 8px 0 0;">
              <h2 style="margin:0;font-size:18px;">Merkoteks Personel ve İzin Sistemi</h2>
            </div>
            <div style="background:#fff;padding:20px;border:1px solid #e2e8f0;border-top:0;border-radius:0 0 8px 8px;">
              <p style="color:#0f172a;">Yeni bir yıllık izin kaydı oluşturuldu.</p>
              <table style="width:100%;border-collapse:collapse;margin-top:12px;">
                <tr><td style="padding:6px 8px;background:#f1f5f9;font-weight:600;">Personel</td><td style="padding:6px 8px;">{p['ad_soyad']} (Sicil: {p.get('sicil_no','')})</td></tr>
                <tr><td style="padding:6px 8px;background:#f1f5f9;font-weight:600;">Departman</td><td style="padding:6px 8px;">{p.get('departman','-')}</td></tr>
                <tr><td style="padding:6px 8px;background:#f1f5f9;font-weight:600;">İzin Türü</td><td style="padding:6px 8px;">{body.izin_turu}</td></tr>
                <tr><td style="padding:6px 8px;background:#f1f5f9;font-weight:600;">Başlangıç</td><td style="padding:6px 8px;">{body.start_date}</td></tr>
                <tr><td style="padding:6px 8px;background:#f1f5f9;font-weight:600;">Bitiş</td><td style="padding:6px 8px;">{body.end_date}</td></tr>
                <tr><td style="padding:6px 8px;background:#f1f5f9;font-weight:600;">Kullanılacak Gün</td><td style="padding:6px 8px;"><strong>{calc['days']} gün</strong></td></tr>
              </table>
              <p style="color:#64748b;font-size:12px;margin-top:16px;">Bu e-posta {user.get('name','İK')} tarafından oluşturuldu.</p>
            </div>
          </td></tr>
        </table>
        """
        for r in recipients:
            background_tasks.add_task(_send_email, r, subject, html)
    return {**rec.model_dump(), "breakdown": calc["breakdown"], "notified": recipients}

@api.delete("/leaves/{lid}")
async def delete_leave(lid: str, request: Request, reason: str = "",
                        user: dict = Depends(require_roles("admin", "hr"))):
    L = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not L:
        raise HTTPException(status_code=404, detail="İzin kaydı bulunamadı")
    p = await db.personnel.find_one({"id": L["personnel_id"]}, {"_id": 0})
    bal_before = await _balance_for(p) if p else {"remaining": 0}
    await db.leaves.delete_one({"id": lid})
    await db.leave_allocation.delete_many({"lid": lid})
    bal_after = await _balance_for(p) if p else {"remaining": 0}
    await _audit(action="delete", module="leaves", entity_type="leave",
                 entity_id=lid, entity_name=(p.get("ad_soyad") if p else L["personnel_id"]),
                 old_values={"personnel_id": L["personnel_id"], "sicil_no": p.get("sicil_no") if p else None,
                             "start_date": L["start_date"], "end_date": L["end_date"],
                             "days": L["days"], "izin_turu": L.get("izin_turu"),
                             "balance_before": bal_before.get("remaining"),
                             "balance_after": bal_after.get("remaining"),
                             "reason": reason or "(gerekçe yok)"},
                 description=(f"İzin silindi: {p.get('ad_soyad') if p else '?'} — "
                              f"{L['start_date']} → {L['end_date']} ({L['days']} gün). "
                              f"Bakiye {bal_before.get('remaining')} → {bal_after.get('remaining')}. "
                              f"Gerekçe: {reason or '—'}"),
                 request=request, user=user)
    return {"ok": True, "restored_days": L["days"],
            "balance_before": bal_before.get("remaining"),
            "balance_after": bal_after.get("remaining")}


class BulkDeleteLeavesIn(BaseModel):
    ids: List[str]
    reason: str
    admin_password: Optional[str] = None  # 20+ kayıt için zorunlu


@api.post("/leaves/bulk-delete")
async def bulk_delete_leaves(body: BulkDeleteLeavesIn, request: Request,
                              user: dict = Depends(require_roles("admin", "hr"))):
    ids = list({i for i in body.ids if i})
    if not ids:
        raise HTTPException(status_code=400, detail="Kayıt seçilmedi")
    if not (body.reason or "").strip():
        raise HTTPException(status_code=400, detail="Silme gerekçesi zorunlu")
    # 20+ kayıt için admin şifresi zorunlu (İK dahil)
    if len(ids) >= 20:
        if not body.admin_password:
            raise HTTPException(status_code=403, detail="20+ kayıt için yönetici şifresi zorunlu")
        admin = await db.users.find_one({"role": "admin"})
        if not admin or not verify_password(body.admin_password, admin.get("password_hash", "")):
            raise HTTPException(status_code=403, detail="Yönetici şifresi hatalı")

    docs = []
    async for L in db.leaves.find({"id": {"$in": ids}}, {"_id": 0}):
        docs.append(L)
    if not docs:
        return {"success": 0, "skipped": 0, "errors": [], "total_days_restored": 0,
                 "affected_personnel": 0, "annual_count": 0, "advance_count": 0, "other_count": 0}

    # Personel bakiyeleri toplam takibi için, personel bazında topla
    per_person: dict = {}
    for L in docs:
        per_person.setdefault(L["personnel_id"], []).append(L)

    balances_before: dict = {}
    for pid in per_person:
        p = await db.personnel.find_one({"id": pid}, {"_id": 0})
        b = await _balance_for(p) if p else {"remaining": 0}
        balances_before[pid] = b.get("remaining", 0)

    # Silme uygula
    delres = await db.leaves.delete_many({"id": {"$in": ids}})
    await db.leave_allocation.delete_many({"lid": {"$in": ids}})

    total_days = round(sum(float(L.get("days", 0)) for L in docs), 2)
    annual = sum(1 for L in docs if (L.get("izin_turu") or "").lower().startswith(("yıl", "yil")))
    advance = sum(1 for L in docs if "avans" in (L.get("izin_turu") or "").lower())
    other = len(docs) - annual - advance

    balances_after: dict = {}
    for pid in per_person:
        p = await db.personnel.find_one({"id": pid}, {"_id": 0})
        b = await _balance_for(p) if p else {"remaining": 0}
        balances_after[pid] = b.get("remaining", 0)

    batch_id = str(uuid.uuid4())
    await _audit(action="bulk_delete", module="leaves", entity_type="leave",
                 entity_id=batch_id, entity_name=f"Toplu izin silme ({len(docs)})",
                 old_values={"batch_id": batch_id, "count": len(docs),
                             "annual": annual, "advance": advance, "other": other,
                             "total_days_restored": total_days,
                             "affected_personnel": len(per_person),
                             "reason": body.reason.strip(),
                             "balances_before": balances_before,
                             "balances_after": balances_after,
                             "deleted_leaves": [
                                {"id": L["id"], "personnel_id": L["personnel_id"],
                                 "start": L["start_date"], "end": L["end_date"],
                                 "days": L["days"], "turu": L.get("izin_turu")}
                                for L in docs
                             ]},
                 description=(f"Toplu izin silme: {len(docs)} kayıt, "
                              f"{len(per_person)} personel, {total_days} gün iade. "
                              f"Gerekçe: {body.reason.strip()[:120]}"),
                 request=request, user=user)

    return {
        "success": delres.deleted_count,
        "skipped": len(ids) - delres.deleted_count,
        "total_days_restored": total_days,
        "affected_personnel": len(per_person),
        "annual_count": annual,
        "advance_count": advance,
        "other_count": other,
        "batch_id": batch_id,
    }


class LeavesDeletePreviewIn(BaseModel):
    ids: List[str]


@api.post("/leaves/delete-preview")
async def leaves_delete_preview(body: LeavesDeletePreviewIn,
                                  _: dict = Depends(require_roles("admin", "hr"))):
    """Toplu silme öncesi özet: personel sayısı, toplam gün, tür dağılımı."""
    ids = list({i for i in body.ids if i})
    if not ids:
        return {"total": 0, "items": [], "affected_personnel": 0,
                "total_days_to_restore": 0, "annual": 0, "advance": 0, "other": 0}
    docs = []
    async for L in db.leaves.find({"id": {"$in": ids}}, {"_id": 0}):
        docs.append(L)
    pids = list({L["personnel_id"] for L in docs})
    pmap: dict = {}
    async for p in db.personnel.find({"id": {"$in": pids}}, {"_id": 0, "id": 1, "sicil_no": 1, "ad_soyad": 1}):
        pmap[p["id"]] = p
    items = []
    for L in docs:
        p = pmap.get(L["personnel_id"], {})
        items.append({
            "id": L["id"], "sicil_no": p.get("sicil_no"), "ad_soyad": p.get("ad_soyad"),
            "start_date": L["start_date"], "end_date": L["end_date"],
            "days": L["days"], "izin_turu": L.get("izin_turu"),
        })
    total_days = round(sum(float(L.get("days", 0)) for L in docs), 2)
    annual = sum(1 for L in docs if (L.get("izin_turu") or "").lower().startswith(("yıl", "yil")))
    advance = sum(1 for L in docs if "avans" in (L.get("izin_turu") or "").lower())
    other = len(docs) - annual - advance
    return {
        "total": len(docs),
        "affected_personnel": len(pids),
        "total_days_to_restore": total_days,
        "annual": annual, "advance": advance, "other": other,
        "items": items,
    }

@api.get("/leaves/{lid}/print")
async def print_form(lid: str, _: dict = Depends(get_current_user)):
    L = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not L:
        raise HTTPException(status_code=404, detail="İzin kaydı bulunamadı")
    p = await db.personnel.find_one({"id": L["personnel_id"]}, {"_id": 0})
    # Muvafakatname gerekli mi? — Bakiye negatifse veya avans (hak edişten önce kullanım) ise
    # Iter 60: 0 günlük yer tutucu izinde de bakiye 0 ise muvafakatname gerekli sayılır.
    bal = await _compute_entitlements(p) if p else {"remaining": 0}
    days = float(L.get("days") or 0)
    zero_day_no_balance = (days == 0) and (bal.get("remaining", 0) <= 0)
    consent_required = (
        bal.get("remaining", 0) < 0
        or L["start_date"] < (bal.get("last_entitlement_date") or "0000-00-00")
        or zero_day_no_balance
    )
    return {"leave": L, "personnel": p, "balance": bal, "consent_required": consent_required,
            "form_meta": {"dokuman_no": "İK.FR.07", "duzenleme_tarihi": "02.01.2023", "revizyon_no": "01"}}

# -----------------------------------------------------------------------------
# Holidays
# -----------------------------------------------------------------------------
@api.get("/holidays")
async def list_holidays(_: dict = Depends(get_current_user)):
    m = await get_all_holidays()
    items = [{"date": d, **v} for d, v in m.items()]
    items.sort(key=lambda x: x["date"])
    return items

@api.post("/holidays")
async def add_holiday(body: HolidayIn, request: Request, current: dict = Depends(require_roles("admin", "hr"))):
    doc = body.model_dump()
    existing = await db.holidays.find_one({"date": doc["date"]}, {"_id": 0})
    await db.holidays.update_one({"date": doc["date"]}, {"$set": doc}, upsert=True)
    await _audit(action=("update" if existing else "create"), module="holidays", entity_type="holiday",
                 entity_id=doc["date"], entity_name=doc["name"],
                 old_values=existing, new_values=doc,
                 description=f"Tatil {'güncellendi' if existing else 'eklendi'}: {doc['date']} — {doc['name']}",
                 request=request, user=current)
    return {"ok": True}

@api.delete("/holidays/{date_iso}")
async def del_holiday(date_iso: str, request: Request, current: dict = Depends(require_roles("admin", "hr"))):
    existing = await db.holidays.find_one({"date": date_iso}, {"_id": 0})
    r = await db.holidays.delete_one({"date": date_iso})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Özel tatil bulunamadı (sistem tatilleri silinemez)")
    await _audit(action="delete", module="holidays", entity_type="holiday",
                 entity_id=date_iso, entity_name=(existing.get("name") if existing else date_iso),
                 old_values=existing,
                 description=f"Özel tatil silindi: {date_iso}",
                 request=request, user=current)
    return {"ok": True}


def _parse_import_date(s: str) -> Optional[date]:
    """Yüklenen dosyadaki tarih formatlarını tanır.
    - '1.01.2026' / '01.01.2026' → D.M.Y (TR)
    - '3.15.2000' / '10.28.2024' → M.D.Y (US) — çünkü ikinci parça > 12
    Kural: ilk parça > 12 → D.M; ikinci parça > 12 → M.D; aksi halde D.M (TR default).
    """
    parts = s.strip().replace("/", ".").replace("-", ".").split(".")
    if len(parts) != 3:
        return None
    try:
        a, b, y = int(parts[0]), int(parts[1]), int(parts[2])
    except Exception:
        return None
    if y < 100: y += 2000
    if a > 12:
        d_, m_ = a, b
    elif b > 12:
        m_, d_ = a, b
    else:
        d_, m_ = a, b  # TR default
    try:
        return date(y, m_, d_)
    except Exception:
        return None


def _classify_holiday(name: str) -> str:
    n = (name or "").strip().lower().replace("i̇", "i")
    if not n: return "Diğer"
    if "aref" in n or "arife" in n: return "Arife"
    if "ramazan" in n or "kurban" in n: return "Dinî Bayram"
    if "cumhuriyet" in n or "ulusal egem" in n or "atatürk" in n or "zafer" in n or "demokrasi" in n: return "Ulusal Bayram"
    if "yılbaş" in n or "emek" in n: return "Resmî Tatil"
    return "Diğer"


@api.post("/holidays/bulk-import-text")
async def holidays_bulk_import_text(payload: dict, request: Request,
                                     user: dict = Depends(require_roles("admin", "hr"))):
    """Yapıştırılan metin (sekme veya tab-boşluk ayraçlı) tatil listesini içe aktarır.
    Beklenen format: 'TARIH<TAB>TATIL_TANIMI<TAB>GUN_DEGERI' her satırda.
    Upsert: (date, name) unique. day_value: 1 → 'full', 0.5 / 0,5 → 'half'.
    """
    text = (payload or {}).get("text", "")
    filename = (payload or {}).get("filename", "yapistirilan_metin.txt")
    if not text or not text.strip():
        raise HTTPException(status_code=400, detail="Boş metin")

    total = 0; added = 0; updated = 0; empty_name = 0
    invalid_date = 0; errors: List[dict] = []
    affected_years: set = set()
    seen_pairs: set = set()  # (date_iso, name_lower) — dosya-içi tekrar filtreleme

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line: continue
        # Başlık satırlarını atla
        low = line.lower()
        if ("tarih" in low and "tatil" in low) or low.startswith("sütun"):
            continue
        # Tab / birden fazla boşluk ile split
        import re as _re
        parts = _re.split(r"\t+|\s{2,}", line)
        if len(parts) < 2:
            continue
        total += 1
        d_str = parts[0].strip()
        name = parts[1].strip() if len(parts) >= 2 else ""
        day_val_str = parts[2].strip() if len(parts) >= 3 else "1"

        d = _parse_import_date(d_str)
        if not d:
            invalid_date += 1
            errors.append({"line": line, "reason": "Geçersiz tarih"})
            continue
        iso = d.isoformat()

        try:
            dv = float(day_val_str.replace(",", "."))
        except Exception:
            dv = 1.0
        h_type = "half" if abs(dv - 0.5) < 1e-9 else "full"

        display_name = name if name else "Tatil Tanımı Belirtilmemiş"
        if not name:
            empty_name += 1

        pair = (iso, display_name.strip().lower())
        if pair in seen_pairs:
            continue  # dosya içi mükerrer
        seen_pairs.add(pair)

        cat = _classify_holiday(display_name) if name else "Diğer"
        doc = {
            "date": iso, "year": d.year, "name": display_name,
            "day_value": 0.5 if h_type == "half" else 1.0,
            "type": h_type, "category": cat,
            "source": "Yüklenen Tatil Listesi",
            "active": True, "needs_review": (not name),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        existing = await db.holidays_import.find_one(
            {"date": iso, "name": display_name}, {"_id": 0})
        if existing:
            await db.holidays_import.update_one(
                {"date": iso, "name": display_name}, {"$set": doc})
            updated += 1
        else:
            doc["id"] = str(uuid.uuid4())
            doc["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.holidays_import.insert_one(doc)
            added += 1
        affected_years.add(d.year)

    result = {
        "total_lines": total, "added": added, "updated": updated,
        "duplicates_skipped": total - added - updated - invalid_date,
        "empty_name": empty_name, "invalid_date": invalid_date,
        "error_count": len(errors), "errors": errors[:50],
        "affected_years": sorted(list(affected_years)),
    }
    await _audit(
        action="bulk_import", module="holidays", entity_type="holiday",
        entity_id=None, entity_name=f"Tatil aktarım: {added}+{updated}",
        new_values={"filename": filename, **{k: v for k, v in result.items() if k != "errors"}},
        description=f"Tatil listesi içe aktarıldı: +{added} yeni, ~{updated} güncel, "
                    f"{invalid_date} geçersiz tarih, yıllar: {result['affected_years']}",
        request=request, user=user,
    )
    return result


@api.get("/holidays/records")
async def holidays_records_list(year: Optional[int] = None,
                                 _: dict = Depends(get_current_user)):
    q: dict = {}
    if year: q["year"] = int(year)
    items = []
    async for h in db.holidays_import.find(q, {"_id": 0}).sort([("date", 1), ("name", 1)]):
        items.append(h)
    return items


@api.get("/holidays/years")
async def holidays_years(_: dict = Depends(get_current_user)):
    years = await db.holidays_import.distinct("year")
    return {"years": sorted([y for y in years if y])}


# -----------------------------------------------------------------------------
# Email (Emergent-managed Resend proxy)
# -----------------------------------------------------------------------------
EMAIL_BASE_URL = "https://integrations.emergentagent.com"

@api.post("/settings/notifications/test")
async def send_test_notification(_user: dict = Depends(require_roles("admin"))):
    """Iter 64: Kayıtlı bildirim e-postasına anlık test gönder."""
    doc = await db.app_settings.find_one({"_id": "notifications"})
    recipient = ((doc or {}).get("hr_alert_email") or "").strip()
    if not recipient:
        raise HTTPException(status_code=400, detail="Önce bir e-posta adresi kaydedin")
    html = (
        f"<div style='font-family:Arial;padding:20px'>"
        f"<h2 style='color:#1e293b'>Merkoteks — Test E-postası</h2>"
        f"<p>Bu bir test mesajıdır. Bildirim e-postası ayarınız çalışıyor. "
        f"Her sabah 08:00'de (İstanbul) günlük özel izin uyarı özeti bu adrese gönderilecektir.</p>"
        f"<p style='color:#94a3b8;font-size:11px'>Gönderim zamanı: "
        f"{datetime.now(timezone.utc).astimezone().strftime('%d.%m.%Y %H:%M')}</p>"
        f"</div>"
    )
    ok, reason = await _send_email_with_reason(recipient,
        "Merkoteks — Bildirim E-postası Testi", html)
    if not ok:
        raise HTTPException(status_code=502, detail=reason)
    return {"ok": True, "sent_to": recipient}


@api.get("/settings/notifications")
async def get_notification_settings(_: dict = Depends(get_current_user)):
    """Iter 63: HR uyarı e-postası ayarı okunur."""
    doc = await db.app_settings.find_one({"_id": "notifications"})
    return {"hr_alert_email": ((doc or {}).get("hr_alert_email") or "")}


@api.put("/settings/notifications")
async def set_notification_settings(body: dict, request: Request,
                                     user: dict = Depends(require_roles("admin"))):
    email = (body.get("hr_alert_email") or "").strip()
    if email and ("@" not in email or "." not in email.split("@")[-1]):
        raise HTTPException(status_code=400, detail="Geçerli bir e-posta adresi girin")
    prev = await db.app_settings.find_one({"_id": "notifications"})
    await db.app_settings.update_one(
        {"_id": "notifications"},
        {"$set": {"hr_alert_email": email,
                  "updated_at": datetime.now(timezone.utc).isoformat(),
                  "updated_by": user.get("id")}},
        upsert=True,
    )
    await _audit(action="update", module="settings",
                 entity_type="settings", entity_id="notifications",
                 entity_name="Bildirim E-postası",
                 old_values={"hr_alert_email": (prev or {}).get("hr_alert_email", "")},
                 new_values={"hr_alert_email": email},
                 description="HR uyarı e-postası güncellendi",
                 request=request, user=user)
    return {"ok": True, "hr_alert_email": email}


# -----------------------------------------------------------------------------
# Iter 63: Günlük HR uyarı özeti (cron endpoint)
# -----------------------------------------------------------------------------
def _cron_authed(request: Request) -> bool:
    secret = os.environ.get("WEBHOOK_CRON_SECRET", "")
    if not secret:
        return False
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return False
    return hmac.compare_digest(auth[7:], secret)


@api.post("/cron/hr-daily-alerts")
async def cron_hr_daily_alerts(request: Request, background_tasks: BackgroundTasks):
    # Cron endpoints must ack 2xx immediately; enqueue/background the actual work.
    if not _cron_authed(request):
        raise HTTPException(status_code=401, detail="Unauthorized")
    background_tasks.add_task(_hr_daily_alerts_run)
    return {"ok": True}


async def _hr_daily_alerts_run():
    doc = await db.app_settings.find_one({"_id": "notifications"})
    recipient = ((doc or {}).get("hr_alert_email") or "").strip()
    if not recipient:
        log.info("HR daily alerts skipped — no recipient configured")
        return
    today = date.today()
    today_iso = today.isoformat()
    future_10 = (today + timedelta(days=10)).isoformat()
    past_10 = (today - timedelta(days=10)).isoformat()

    gebelik: list = []
    async for it in db.special_leaves.find(
        {"deleted": {"$ne": True}, "tur": {"$in": ["gebelik", "dogum"]}},
        {"_id": 0},
    ):
        d = it.get("calisamaz_rapor_tarihi") or ""
        if d and today_iso <= d <= future_10:
            gebelik.append(it)

    milk_up: list = []
    milk_end: list = []
    async for it in db.special_leaves.find(
        {"deleted": {"$ne": True}, "tur": "sut_izni"}, {"_id": 0}
    ):
        e = it.get("end_date") or it.get("sut_izni_bitis") or ""
        if not e:
            continue
        if today_iso <= e <= future_10:
            milk_up.append(it)
        elif past_10 <= e < today_iso:
            milk_end.append(it)

    if not gebelik and not milk_up and not milk_end:
        log.info("HR daily alerts — nothing to report today")
        return

    pids = list({x["personnel_id"] for lst in (gebelik, milk_up, milk_end)
                 for x in lst if x.get("personnel_id")})
    pmap: dict = {}
    if pids:
        async for p in db.personnel.find(
            {"id": {"$in": pids}},
            {"_id": 0, "id": 1, "sicil_no": 1, "ad_soyad": 1, "departman": 1},
        ):
            pmap[p["id"]] = p
    for lst in (gebelik, milk_up, milk_end):
        for x in lst:
            p = pmap.get(x["personnel_id"], {})
            x["_sicil"] = p.get("sicil_no") or "—"
            x["_ad"] = p.get("ad_soyad") or "—"
            x["_dep"] = p.get("departman") or "—"

    def _tr_date(iso: str) -> str:
        try:
            y, m, d = iso[:10].split("-")
            return f"{d}.{m}.{y}"
        except Exception:
            return iso or "—"

    def _rows(lst, field):
        if not lst:
            return "<p style='color:#888'>Kayıt yok.</p>"
        body = "".join(
            f"<tr>"
            f"<td style='padding:6px;border:1px solid #ddd'>{x['_sicil']}</td>"
            f"<td style='padding:6px;border:1px solid #ddd'><b>{x['_ad']}</b></td>"
            f"<td style='padding:6px;border:1px solid #ddd'>{x['_dep']}</td>"
            f"<td style='padding:6px;border:1px solid #ddd;font-family:monospace'>{_tr_date(x.get(field, ''))}</td>"
            f"</tr>" for x in lst
        )
        return (
            "<table style='border-collapse:collapse;width:100%;font-size:13px;margin-top:6px'>"
            "<thead><tr style='background:#f4f4f5'>"
            "<th style='padding:6px;border:1px solid #ddd;text-align:left'>Sicil</th>"
            "<th style='padding:6px;border:1px solid #ddd;text-align:left'>Ad Soyad</th>"
            "<th style='padding:6px;border:1px solid #ddd;text-align:left'>Departman</th>"
            "<th style='padding:6px;border:1px solid #ddd;text-align:left'>Tarih</th>"
            "</tr></thead><tbody>" + body + "</tbody></table>"
        )

    html = (
        f"<div style='font-family:Arial,sans-serif;padding:20px;color:#111'>"
        f"<h2 style='color:#1e293b'>Merkoteks — Günlük Özel İzin Uyarıları</h2>"
        f"<p style='color:#64748b'>Tarih: <b>{today.strftime('%d.%m.%Y')}</b></p>"
        f"<h3 style='color:#be185d'>Gebelik — Çalışamaz Raporu Yaklaşan (≤10 gün) — {len(gebelik)}</h3>"
        + _rows(gebelik, "calisamaz_rapor_tarihi") +
        f"<h3 style='color:#0369a1'>Süt İzni — Yaklaşan (≤10 gün) — {len(milk_up)}</h3>"
        + _rows(milk_up, "end_date") +
        f"<h3 style='color:#475569'>Süt İzni — Son 10 Günde Biten — {len(milk_end)}</h3>"
        + _rows(milk_end, "end_date") +
        f"<p style='color:#94a3b8;font-size:11px;margin-top:24px'>Bu e-posta otomatik olarak gönderilmiştir. Ayarlar → Bildirim E-postası bölümünden değiştirebilirsiniz.</p>"
        f"</div>"
    )
    await _send_email(
        recipient,
        f"Günlük Özel İzin Uyarıları — {today.strftime('%d.%m.%Y')}",
        html,
    )


async def _send_email(recipient: str, subject: str, html: str) -> bool:
    key = os.environ.get("EMERGENT_EMAIL_KEY")
    from_name = os.environ.get("EMAIL_FROM_NAME", "Merkoteks")
    if not key or not recipient:
        log.warning("Email skipped (missing key or recipient)")
        return False
    payload = {"to": [recipient], "subject": subject, "html": html, "from_name": from_name}
    try:
        async with httpx.AsyncClient(timeout=20) as c:
            r = await c.post(f"{EMAIL_BASE_URL}/api/v1/email/send",
                             headers={"X-Email-Key": key}, json=payload)
            if r.status_code == 429:
                log.warning("Email rate-limited by provider (429) — try again later")
                return False
            r.raise_for_status()
            log.info("Email sent to %s", recipient)
            return True
    except Exception as e:
        log.error("Email send failed: %s", e)
        return False


async def _send_email_with_reason(recipient: str, subject: str, html: str) -> tuple:
    """Test butonu için: (ok, reason) döner."""
    key = os.environ.get("EMERGENT_EMAIL_KEY")
    if not key:
        return False, "EMERGENT_EMAIL_KEY yapılandırılmamış"
    if not recipient:
        return False, "Alıcı adres boş"
    from_name = os.environ.get("EMAIL_FROM_NAME", "Merkoteks")
    payload = {"to": [recipient], "subject": subject, "html": html, "from_name": from_name}
    try:
        async with httpx.AsyncClient(timeout=20) as c:
            r = await c.post(f"{EMAIL_BASE_URL}/api/v1/email/send",
                             headers={"X-Email-Key": key}, json=payload)
            if r.status_code == 429:
                return False, "E-posta sağlayıcı hız sınırı (429) — lütfen 1-2 dakika sonra tekrar deneyin"
            r.raise_for_status()
            return True, "sent"
    except httpx.HTTPStatusError as e:
        return False, f"E-posta servisi hatası: HTTP {e.response.status_code}"
    except Exception as e:
        return False, f"E-posta gönderilemedi: {e}"

# -----------------------------------------------------------------------------
# Bulk Excel import for personnel
# -----------------------------------------------------------------------------
@api.post("/personnel/import/preview")
async def personnel_import_preview(file: UploadFile = File(...),
                                     _: dict = Depends(require_roles("admin", "hr"))):
    """Excel'i önizler — DB'ye YAZMAZ. Sicil normalize edilir. Duplicate/eksik uyarıları döner."""
    from openpyxl import load_workbook
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı")
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Dosya boş")
    headers_row = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(name):
        try: return headers_row.index(name)
        except ValueError: return -1

    fields = ["sicil_no", "ad_soyad", "tc_no", "ise_giris", "isten_cikis", "dogum_tarihi",
              "departman", "gorev", "sirket", "aktif", "onceki_kidem_yil", "telefon",
              "email", "aciklama"]
    idx = {f: col(f) for f in fields}
    if idx["sicil_no"] < 0 or idx["ad_soyad"] < 0 or idx["ise_giris"] < 0:
        raise HTTPException(status_code=400, detail="Zorunlu kolonlar eksik: sicil_no, ad_soyad, ise_giris")

    def val(row, key, default=""):
        i = idx[key]
        if i < 0 or i >= len(row): return default
        v = row[i]
        if v is None: return default
        if isinstance(v, datetime): return v.date().isoformat()
        if hasattr(v, "isoformat"): return v.isoformat()[:10]
        return str(v).strip()

    existing_sicil = set()
    existing_tc = set()
    async for p in db.personnel.find({}, {"_id": 0, "sicil_no": 1, "tc_no": 1}):
        if p.get("sicil_no"):
            existing_sicil.add(p["sicil_no"])
        if p.get("tc_no"):
            existing_tc.add(p["tc_no"])

    stats = {"total": 0, "valid": 0, "duplicate": 0, "invalid": 0, "existing": 0}
    seen_sicil = {}
    seen_tc = {}
    out_rows = []
    for ri, r in enumerate(rows[1:], start=2):
        sicil_raw = r[idx["sicil_no"]] if idx["sicil_no"] < len(r) else None
        sicil = _normalize_sicil(sicil_raw)
        ad_soyad = val(r, "ad_soyad")
        # Örnek şablon satırlarını atla
        low_ad = (ad_soyad or "").upper()
        if sicil in {"0001", "0002", "0003"} and ("ÖRNEK" in low_ad or "OR" in low_ad[:3]):
            continue
        if not sicil and not ad_soyad:
            continue
        stats["total"] += 1
        tc = val(r, "tc_no")
        ise = val(r, "ise_giris")
        status_code = "valid"
        status_label = "Geçerli — Kayda Hazır"
        reason = ""
        if not sicil:
            status_code = "invalid"; status_label = "Eksik Sicil"; reason = "Sicil no boş"
        elif not ad_soyad:
            status_code = "invalid"; status_label = "Eksik Ad Soyad"; reason = "Ad Soyad boş"
        elif not ise:
            status_code = "invalid"; status_label = "Eksik İşe Giriş"; reason = "İşe giriş tarihi boş"
        elif sicil in seen_sicil:
            status_code = "duplicate"; status_label = "Dosya İçi Mükerrer"
            reason = f"Dosyada {seen_sicil[sicil]}. satırda da var"
        elif sicil in existing_sicil:
            status_code = "existing"; status_label = "Sistemde Mevcut"
            reason = f"Sicil {sicil} sistemde zaten kayıtlı — atlanacak"
        elif tc and tc in seen_tc:
            status_code = "duplicate"; status_label = "Dosya İçi Mükerrer TC"
            reason = f"TC {tc} dosyada {seen_tc[tc]}. satırda da var"
        elif tc and tc in existing_tc:
            status_code = "existing"; status_label = "TC Sistemde Mevcut"
            reason = f"TC {tc} zaten kayıtlı"
        if status_code == "valid":
            seen_sicil[sicil] = ri
            if tc:
                seen_tc[tc] = ri
        stats[status_code] = stats.get(status_code, 0) + 1
        out_rows.append({
            "row": ri, "sicil_no": sicil, "ad_soyad": ad_soyad, "tc_no": tc,
            "ise_giris": ise, "departman": val(r, "departman"),
            "sirket": val(r, "sirket"), "gorev": val(r, "gorev"),
            "isten_cikis": val(r, "isten_cikis"), "dogum_tarihi": val(r, "dogum_tarihi"),
            "aktif_raw": val(r, "aktif", "true"),
            "onceki_kidem_yil": val(r, "onceki_kidem_yil", "0"),
            "telefon": val(r, "telefon"), "email": val(r, "email"),
            "aciklama": val(r, "aciklama"),
            "status_code": status_code, "status_label": status_label, "reason": reason,
            "can_insert": status_code == "valid",
        })
    return {"filename": file.filename or "personel.xlsx", "stats": stats, "rows": out_rows}


@api.post("/personnel/import/confirm")
async def personnel_import_confirm(body: dict, request: Request,
                                     user: dict = Depends(require_roles("admin", "hr"))):
    """Preview'den onaylanan satırları DB'ye yazar. Sadece can_insert=True olanlar işlenir."""
    rows_in = (body or {}).get("rows", [])
    filename = (body or {}).get("filename", "personel.xlsx")
    if not isinstance(rows_in, list) or not rows_in:
        raise HTTPException(status_code=400, detail="Yazılacak kayıt yok")

    # DB double-check (yeniden mükerrer)
    existing_sicil = set()
    existing_tc = set()
    async for p in db.personnel.find({}, {"_id": 0, "sicil_no": 1, "tc_no": 1}):
        if p.get("sicil_no"): existing_sicil.add(p["sicil_no"])
        if p.get("tc_no"): existing_tc.add(p["tc_no"])

    created = 0; skipped = 0; errors = []
    for r in rows_in:
        if not r.get("can_insert"):
            skipped += 1; continue
        sicil = _normalize_sicil(r.get("sicil_no"))
        if not sicil or sicil in existing_sicil:
            skipped += 1; continue
        tc = str(r.get("tc_no") or "").strip()
        if tc and tc in existing_tc:
            skipped += 1; continue
        try:
            aktif_v = str(r.get("aktif_raw") or "true").lower()
            aktif = aktif_v in ("true", "1", "aktif", "evet", "yes", "y", "e")
            try: kidem = int(float(str(r.get("onceki_kidem_yil") or "0")))
            except Exception: kidem = 0
            doc = Personnel(
                sicil_no=sicil, ad_soyad=r.get("ad_soyad", ""),
                tc_no=tc,
                ise_giris=r.get("ise_giris", ""),
                isten_cikis=r.get("isten_cikis") or None,
                dogum_tarihi=r.get("dogum_tarihi") or None,
                departman=r.get("departman", ""), gorev=r.get("gorev", ""),
                sirket=r.get("sirket", ""), aktif=aktif,
                onceki_kidem_yil=kidem, telefon=r.get("telefon", ""),
                email=r.get("email", ""), aciklama=r.get("aciklama", ""),
            ).model_dump()
            await db.personnel.insert_one(doc)
            existing_sicil.add(sicil)
            if tc: existing_tc.add(tc)
            created += 1
        except Exception as e:
            errors.append({"row": r.get("row"), "error": str(e)})

    await _audit(action="bulk_import", module="personnel", entity_type="personnel",
                 entity_id=None, entity_name=f"Toplu personel içeri aktarımı: {created}",
                 new_values={"filename": filename, "created": created, "skipped": skipped,
                             "errors": len(errors)},
                 description=f"{filename} → +{created} personel, {skipped} atlandı",
                 request=request, user=user)
    return {"created": created, "skipped": skipped, "errors": errors}


@api.post("/personnel/import")
async def personnel_import(file: UploadFile = File(...),
                            _: dict = Depends(require_roles("admin", "hr"))):
    """Excel şablonundan personel içe aktar — satır bazlı çakışma/dup kontrolü ile.
    Beklenen kolonlar (1. satır başlık): sicil_no, ad_soyad, tc_no, ise_giris,
      isten_cikis, dogum_tarihi, departman, gorev, sirket, aktif, onceki_kidem_yil,
      telefon, email, aciklama
    """
    from openpyxl import load_workbook
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı")

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Dosya boş")

    headers_row = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(name):
        try: return headers_row.index(name)
        except ValueError: return -1

    fields = ["sicil_no", "ad_soyad", "tc_no", "ise_giris", "isten_cikis", "dogum_tarihi",
              "departman", "gorev", "sirket", "aktif", "onceki_kidem_yil", "telefon",
              "email", "aciklama"]
    idx = {f: col(f) for f in fields}
    if idx["sicil_no"] < 0 or idx["ad_soyad"] < 0 or idx["ise_giris"] < 0:
        raise HTTPException(status_code=400, detail="Zorunlu kolonlar eksik: sicil_no, ad_soyad, ise_giris")

    def val(row, key, default=""):
        i = idx[key]
        if i < 0 or i >= len(row): return default
        v = row[i]
        if v is None: return default
        if isinstance(v, datetime): return v.date().isoformat()
        if hasattr(v, "isoformat"): return v.isoformat()[:10]
        return str(v).strip()

    # Sicil normalize eden yardımcı — Excel'in "1001.0" varyantını "1001" olarak okur
    def norm_sicil_at(row):
        i = idx["sicil_no"]
        if i < 0 or i >= len(row):
            return ""
        return _normalize_sicil(row[i])

    # 1) DB'deki mevcut sicil ve TC'leri topla
    existing_sicil = set()
    existing_tc = set()
    async for p in db.personnel.find({}, {"_id": 0, "sicil_no": 1, "tc_no": 1}):
        if p.get("sicil_no"): existing_sicil.add(p["sicil_no"])
        if p.get("tc_no"): existing_tc.add(p["tc_no"])

    # 2) Dosya içindeki tekrarları tespit et
    seen_sicil: dict = {}
    seen_tc: dict = {}
    file_conflicts = []
    parsed_rows = []
    for ri, r in enumerate(rows[1:], start=2):
        sicil = norm_sicil_at(r)
        if not sicil:
            continue
        tc = val(r, "tc_no")
        parsed_rows.append((ri, r, sicil, tc))
        if sicil in seen_sicil:
            file_conflicts.append({"row": ri, "conflicts_with_row": seen_sicil[sicil],
                                    "field": "sicil_no", "value": sicil,
                                    "message": f"Sicil no {sicil} dosya içinde {seen_sicil[sicil]}. satırda da var"})
        else:
            seen_sicil[sicil] = ri
        if tc and tc in seen_tc:
            file_conflicts.append({"row": ri, "conflicts_with_row": seen_tc[tc],
                                    "field": "tc_no", "value": tc,
                                    "message": f"T.C. No {tc} dosya içinde {seen_tc[tc]}. satırda da var"})
        elif tc:
            seen_tc[tc] = ri

    # 3) DB'deki kayıtlarla çakışmaları tespit et
    db_conflicts = []
    for ri, r, sicil, tc in parsed_rows:
        if sicil in existing_sicil:
            db_conflicts.append({"row": ri, "field": "sicil_no", "value": sicil,
                                  "message": f"Sicil no {sicil} sistemde zaten kayıtlı"})
        if tc and tc in existing_tc:
            db_conflicts.append({"row": ri, "field": "tc_no", "value": tc,
                                  "message": f"T.C. No {tc} sistemde zaten kayıtlı"})

    created, skipped, errors = 0, 0, []
    # Aynı sicil için yalnızca ilk satırı yaz — sonrakileri atla
    written_sicil = set()
    for ri, r, sicil, tc in parsed_rows:
        try:
            if sicil in existing_sicil or sicil in written_sicil:
                skipped += 1
                continue
            if tc and tc in existing_tc:
                skipped += 1
                continue
            aktif_v = val(r, "aktif", "true").lower()
            aktif = aktif_v in ("true", "1", "aktif", "evet", "yes", "y", "e")
            kidem_raw = val(r, "onceki_kidem_yil", "0")
            try: kidem = int(float(kidem_raw))
            except Exception: kidem = 0
            doc = Personnel(
                sicil_no=sicil, ad_soyad=val(r, "ad_soyad"),
                tc_no=tc,
                ise_giris=val(r, "ise_giris"),
                isten_cikis=val(r, "isten_cikis") or None,
                dogum_tarihi=val(r, "dogum_tarihi") or None,
                departman=val(r, "departman"), gorev=val(r, "gorev"),
                sirket=val(r, "sirket"), aktif=aktif,
                onceki_kidem_yil=kidem, telefon=val(r, "telefon"),
                email=val(r, "email"), aciklama=val(r, "aciklama"),
            ).model_dump()
            await db.personnel.insert_one(doc)
            written_sicil.add(sicil)
            if tc: existing_tc.add(tc)
            created += 1
        except Exception as e:
            errors.append({"row": ri, "error": str(e)})
    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "file_conflicts": file_conflicts,   # dosya içi tekrarlar
        "db_conflicts": db_conflicts,       # DB ile çakışan satırlar
    }

@api.get("/personnel/export/list.xlsx")
async def personnel_list_export(_: dict = Depends(require_roles("admin", "hr"))):
    """Sistemdeki personelin Excel formatında dışa aktarılması.
    Sütunlar Personel_Yukleme_Sablonu ile aynı düzendedir; ayrıca aktif/pasif ve son
    işlem tarihleri kolay okumak için eklenmiştir."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Personel Listesi"
    headers = ["sicil_no", "ad_soyad", "tc_no", "departman", "sirket",
               "ise_giris", "dogum_tarihi", "onceki_kidem_yili", "aktif", "isten_cikis"]
    ws.append(headers)

    def _iso_to_tr(iso):
        if not iso: return ""
        s = str(iso)[:10]
        parts = s.split("-")
        if len(parts) == 3:
            return f"{parts[2]}.{parts[1]}.{parts[0]}"
        return s

    row_count = 0
    async for p in db.personnel.find({}, {"_id": 0}).sort([("aktif", -1), ("ad_soyad", 1)]):
        ws.append([
            p.get("sicil_no") or "",
            p.get("ad_soyad") or "",
            p.get("tc_no") or "",
            p.get("departman") or "",
            p.get("sirket") or "",
            _iso_to_tr(p.get("ise_giris")),
            _iso_to_tr(p.get("dogum_tarihi")),
            p.get("onceki_kidem_yil") or 0,
            "Evet" if p.get("aktif", True) else "Hayır",
            _iso_to_tr(p.get("isten_cikis")),
        ])
        row_count += 1

    for i, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for c, w in enumerate([14, 28, 14, 20, 22, 14, 14, 14, 10, 14], 1):
        # Excel kolonu >Z olursa güvenli hesap
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = w
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"Personel_Listesi_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"',
                 "X-Row-Count": str(row_count)},
    )


@api.get("/personnel/import/template")
async def personnel_import_template(_: dict = Depends(require_roles("admin", "hr"))):
    """Toplu personel yükleme şablonu — Iter 45.
    Sütunlar: sicil_no, ad_soyad, tc_no, departman, sirket, ise_giris, dogum_tarihi, onceki_kidem_yili.
    KALDIRILAN alanlar (kullanıcı isteği): isten_cikis, aktif — Excel yalnızca güncel AKTİF listesi.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Personel"
    headers = ["sicil_no", "ad", "soyad", "tc_no", "departman", "sirket",
               "ise_giris", "dogum_tarihi", "onceki_kidem_yili"]
    ws.append(headers)
    # 3 örnek satır
    ws.append(["1001", "AHMET", "YILMAZ", "12345678901", "ÜRETİM", "MERKOTEKS A.Ş.",
               "15.01.2015", "20.05.1988", 0])
    ws.append(["0050", "AYŞE", "DEMİR", "10987654321", "MODELHANE", "MERKOTEKS A.Ş.",
               "01.06.2018", "08.11.1992", 2])
    ws.append(["1002", "MEHMET", "KAYA", "23456789012", "KESİMHANE", "MERKOTEKS A.Ş.",
               "10.03.2012", "14.09.1985", 5])
    for i, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for c, w in enumerate([14, 18, 18, 14, 20, 22, 14, 14, 14], 1):
        ws.column_dimensions[chr(64 + c)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Personel_Yukleme_Sablonu.xlsx"'},
    )


# ---- Senkronizasyon (Personel Listesi Karşılaştır / Güncelle) ----
def _sicil_normalized_map(rows_from_excel, sicil_key):
    """Excel satırlarını normalize sicil ile map'ler; boşları ayrı listeye alır; mükerrer tespit eder."""
    by_sicil = {}
    dupes = []
    empties = []
    for r in rows_from_excel:
        sic = _normalize_sicil(r.get(sicil_key, ""))
        if not sic:
            empties.append(r); continue
        if sic in by_sicil:
            dupes.append((sic, by_sicil[sic], r))
        else:
            by_sicil[sic] = r
    return by_sicil, dupes, empties


@api.post("/personnel/sync/compare")
async def personnel_sync_compare(
    file: UploadFile = File(...),
    mapping: str = "",  # JSON: {"sicil_no": "SICIL", "ad_soyad": "AD SOYAD", ...}
    _: dict = Depends(require_roles("admin", "hr")),
):
    """Yüklenen Excel'i sistemle karşılaştırır. DB'YE YAZMAZ.
    mapping boş → auto-detect. Response 4 grup: matched, new_personnel, missing_from_file, changed.
    """
    import json as _json
    from openpyxl import load_workbook
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Excel okunamadı")
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(status_code=400, detail="Dosya boş")
    headers = [str(h).strip() if h else "" for h in all_rows[0]]

    # Otomatik eşleştirme sözlüğü — Iter 53: ad + soyad ayrı, ad_soyad geriye uyum
    auto_map = {
        "sicil_no": ["sicil", "sicil no", "sicil numarası", "sicil_no", "id"],
        "ad": ["ad", "adı", "isim", "first name", "firstname"],
        "soyad": ["soyad", "soyadı", "soy ad", "surname", "last name", "lastname"],
        "ad_soyad": ["ad soyad", "adı soyadı", "isim soyisim", "ad_soyad", "ad-soyad", "personel", "ad-soyadı"],
        "tc_no": ["tc", "t.c.", "t.c. kimlik", "tc kimlik", "tc_no", "kimlik"],
        "departman": ["departman", "bölüm", "birim"],
        "sirket": ["şirket", "sirket", "firma"],
        "ise_giris": ["işe giriş", "ise giris", "işe başlama", "başlama tarihi", "ise_giris"],
        "isten_cikis": ["işten çıkış", "işten ayrılış", "isten_cikis", "çıkış tarihi", "ayrılış"],
        "dogum_tarihi": ["doğum", "dogum", "doğum tarihi", "dogum_tarihi"],
        "onceki_kidem_yil": ["önceki kıdem", "onceki kidem", "kıdem yılı", "onceki_kidem_yil", "onceki_kidem_yili"],
    }

    def auto_detect(field):
        cands = auto_map[field]
        for i, h in enumerate(headers):
            hl = h.lower().strip()
            # Iter 53: 'ad' & 'soyad' tam eşleşme öncelikli; "ad soyad" başlığını yakalamamalı
            if field in ("ad", "soyad"):
                if hl == field or hl in cands:
                    return i
                # 'ad' başlıklısı olmayan ama 'adı' vs. içeren kısa başlıklar
                if field == "ad" and hl in ("ad", "adı", "isim") and "soyad" not in hl:
                    return i
                if field == "soyad" and ("soyad" in hl):
                    return i
                continue
            if hl in cands or any(c in hl for c in cands):
                return i
        return -1

    user_map: dict = {}
    if mapping:
        try: user_map = _json.loads(mapping)
        except Exception: user_map = {}

    resolved = {}
    for f in auto_map.keys():
        col = user_map.get(f)
        if col and col in headers:
            resolved[f] = headers.index(col)
        else:
            resolved[f] = auto_detect(f)
    if resolved["sicil_no"] < 0:
        return {"headers": headers, "resolved": resolved,
                "error": "Sicil Numarası sütunu tespit edilemedi. Lütfen eşleştirme yapın.",
                "needs_mapping": True}

    def val(row, key):
        i = resolved.get(key, -1)
        if i < 0 or i >= len(row): return ""
        v = row[i]
        if v is None: return ""
        if isinstance(v, datetime): return v.date().isoformat()
        if hasattr(v, "isoformat"): return v.isoformat()[:10]
        return str(v).strip()

    excel_rows = []
    for r_idx, r in enumerate(all_rows[1:], start=2):
        rec = {"_row": r_idx}
        for k in resolved:
            if k == "sicil_no":
                rec[k] = _normalize_sicil(r[resolved["sicil_no"]] if resolved["sicil_no"] < len(r) else "")
            else:
                rec[k] = val(r, k)
        # Iter 53: ad + soyad ayrı geldiyse ad_soyad'ı birleştir (tek alan üstün)
        if not rec.get("ad_soyad"):
            ad = (rec.get("ad") or "").strip()
            soyad = (rec.get("soyad") or "").strip()
            merged = f"{ad} {soyad}".strip()
            if merged:
                rec["ad_soyad"] = merged
        excel_rows.append(rec)

    by_sicil = {}; dupes = []; empties = []
    for r in excel_rows:
        s = r["sicil_no"]
        if not s: empties.append(r); continue
        if s in by_sicil: dupes.append((s, by_sicil[s]["_row"], r["_row"]))
        else: by_sicil[s] = r

    db_by_sicil = {}
    async for p in db.personnel.find({}, {"_id": 0}):
        if p.get("sicil_no"): db_by_sicil[p["sicil_no"]] = p

    matched = []; changed = []; new_personnel = []; missing_from_file = []; rehire = []
    for sic, row in by_sicil.items():
        p = db_by_sicil.get(sic)
        if not p:
            new_personnel.append(row); continue
        # Iter 44: sicil DB'de var ama aktif=False ise → rehire adayı (yeni oluşturma!)
        if p.get("aktif") is False:
            rehire.append({
                "sicil_no": sic, "row": row["_row"],
                "ad_soyad": p.get("ad_soyad"), "personnel_id": p["id"],
                "old_isten_cikis": p.get("isten_cikis"),
                "old_departman": p.get("departman"),
                "excel_ad_soyad": row.get("ad_soyad"),
                "excel_departman": row.get("departman"),
                "excel_ise_giris": row.get("ise_giris"),
            })
            continue
        diffs = {}
        # Iter 45: isten_cikis Excel'de artık YOK — sadece diğer bilgi alanları karşılaştırılır
        for f in ["ad_soyad", "tc_no", "departman", "sirket", "ise_giris", "dogum_tarihi"]:
            excel_v = row.get(f, "") or ""
            db_v = p.get(f) or ""
            if excel_v and str(excel_v).strip().casefold() != str(db_v).strip().casefold():
                diffs[f] = {"old": db_v, "new": excel_v}
        if diffs:
            changed.append({"sicil_no": sic, "row": row["_row"],
                            "ad_soyad": p.get("ad_soyad"), "personnel_id": p["id"],
                            "departman": p.get("departman"),
                            "diffs": diffs})
        else:
            matched.append({"sicil_no": sic, "ad_soyad": p.get("ad_soyad"),
                            "personnel_id": p["id"], "row": row["_row"],
                            "departman": p.get("departman"),
                            "ise_giris": p.get("ise_giris")})

    # Sistemde AKTİF olup Excel'de olmayanlar
    for sic, p in db_by_sicil.items():
        if sic not in by_sicil and p.get("aktif", True):
            missing_from_file.append({
                "sicil_no": sic, "ad_soyad": p.get("ad_soyad"),
                "departman": p.get("departman"), "ise_giris": p.get("ise_giris"),
                "personnel_id": p["id"],
            })

    db_active_count = sum(1 for p in db_by_sicil.values() if p.get("aktif", True))
    term_ratio = (len(missing_from_file) / db_active_count) if db_active_count > 0 else 0.0
    return {
        "headers": headers, "resolved": resolved, "auto_detected": mapping == "",
        "summary": {
            "excel_total": len(by_sicil), "db_active": db_active_count,
            "matched": len(matched), "changed": len(changed),
            "new": len(new_personnel), "missing": len(missing_from_file),
            "rehire": len(rehire),
            "duplicates": len(dupes), "empty_sicil": len(empties),
            "term_ratio": round(term_ratio, 4),
        },
        "matched": matched[:500], "changed": changed[:500],
        "new_personnel": new_personnel[:500],
        "missing_from_file": missing_from_file[:500],
        "rehire": rehire[:200],
        "duplicates": [{"sicil_no": s, "rows": [r1, r2]} for s, r1, r2 in dupes[:100]],
        "empty_sicil": empties[:100],
    }


class SyncApplyIn(BaseModel):
    new_rows: List[dict] = []           # onaylanan yeni personel satırları
    update_rows: List[dict] = []        # onaylanan güncellemeler {personnel_id, diffs}
    terminate_ids: List[str] = []       # işten çıkış yapılacak personel_id'leri
    terminate_date: str = ""            # YYYY-MM-DD; boşsa bugün
    terminate_reason: str = "Güncel personel listesinde bulunmuyor"
    # Iter 44: satır bazlı özel işten çıkış tarihleri {personnel_id: "YYYY-MM-DD"}
    terminate_overrides: Dict[str, str] = {}


@api.post("/personnel/sync/apply")
async def personnel_sync_apply(body: SyncApplyIn, request: Request,
                                 user: dict = Depends(require_roles("admin", "hr"))):
    created = 0; updated = 0; terminated = 0; errors = []
    # NEW PERSONNEL
    for r in body.new_rows:
        sic = _normalize_sicil(r.get("sicil_no"))
        if not sic: errors.append({"row": r.get("_row"), "reason": "Sicil eksik"}); continue
        if await db.personnel.find_one({"sicil_no": sic}):
            errors.append({"row": r.get("_row"), "sicil": sic, "reason": "Sistemde zaten var"}); continue
        try:
            def parse_dt(s):
                if not s: return None
                s = str(s).strip()
                for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
                    try: return datetime.strptime(s[:10], fmt).date().isoformat()
                    except Exception: pass
                return None
            kidem = 0
            try: kidem = int(float(str(r.get("onceki_kidem_yil") or "0")))
            except Exception: pass
            doc = Personnel(
                sicil_no=sic, ad_soyad=r.get("ad_soyad", ""), tc_no=r.get("tc_no", ""),
                ise_giris=parse_dt(r.get("ise_giris", "")) or "",
                isten_cikis=parse_dt(r.get("isten_cikis", "")),
                dogum_tarihi=parse_dt(r.get("dogum_tarihi", "")),
                departman=r.get("departman", ""), gorev="",
                sirket=r.get("sirket", ""), aktif=True,
                onceki_kidem_yil=kidem, telefon="", email="", aciklama="",
            ).model_dump()
            await db.personnel.insert_one(doc)
            created += 1
        except Exception as e:
            errors.append({"row": r.get("_row"), "sicil": sic, "reason": str(e)})
    # UPDATES
    for u in body.update_rows:
        pid = u.get("personnel_id")
        diffs = u.get("diffs", {})
        if not pid or not diffs: continue
        set_ = {k: v.get("new") for k, v in diffs.items() if v.get("new") is not None}
        if set_:
            set_["updated_at"] = datetime.now(timezone.utc).isoformat()
            await db.personnel.update_one({"id": pid}, {"$set": set_})
            await _audit(action="update", module="personnel", entity_type="personnel",
                         entity_id=pid, entity_name=u.get("ad_soyad") or pid,
                         old_values={k: v["old"] for k, v in diffs.items()},
                         new_values=set_,
                         description=f"Senkronizasyon güncellemesi: {list(diffs.keys())}",
                         request=request, user=user)
            updated += 1
    # TERMINATE
    term_date = (body.terminate_date or date.today().isoformat())[:10]
    overrides = body.terminate_overrides or {}
    for pid in body.terminate_ids:
        existing = await db.personnel.find_one({"id": pid}, {"_id": 0})
        if not existing: continue
        this_date = (overrides.get(pid) or term_date)[:10]
        await db.personnel.update_one({"id": pid},
            {"$set": {"aktif": False, "isten_cikis": this_date,
                      "aciklama": (existing.get("aciklama", "") + " · " + (body.terminate_reason or "Senkronizasyon: listede yok")).strip(" ·"),
                      "updated_at": datetime.now(timezone.utc).isoformat()}})
        await _audit(action="terminate", module="personnel", entity_type="personnel",
                     entity_id=pid, entity_name=existing.get("ad_soyad"),
                     old_values={"aktif": True, "isten_cikis": existing.get("isten_cikis")},
                     new_values={"aktif": False, "isten_cikis": this_date,
                                 "reason": body.terminate_reason},
                     description=f"Senkronizasyon işten çıkış: {existing.get('ad_soyad')} → {this_date}",
                     request=request, user=user)
        terminated += 1
    await _audit(action="sync", module="personnel", entity_type="personnel",
                 entity_id=None, entity_name="Personel senkronizasyonu",
                 new_values={"created": created, "updated": updated,
                             "terminated": terminated, "errors": len(errors)},
                 description=f"Senkronizasyon: +{created} yeni · {updated} güncelleme · {terminated} işten çıkış",
                 request=request, user=user)
    return {"created": created, "updated": updated, "terminated": terminated,
            "errors": errors, "term_date": term_date}


async def _build_deleted_leaves_filter(from_date, to_date, user_email, action_type):
    filt = {"module": "leaves", "action": {"$in": ["delete", "bulk_delete"]}}
    if action_type and action_type in ("delete", "bulk_delete"):
        filt["action"] = action_type
    if user_email:
        filt["user_email"] = {"$regex": re.escape(user_email), "$options": "i"}
    if from_date or to_date:
        rng = {}
        if from_date: rng["$gte"] = str(from_date)
        if to_date: rng["$lte"] = str(to_date) + "T23:59:59"
        filt["timestamp"] = rng
    return filt


@api.get("/reports/deleted-leaves")
async def reports_deleted_leaves(limit: int = 100, skip: int = 0,
                                    from_date: Optional[str] = None,
                                    to_date: Optional[str] = None,
                                    user_email: Optional[str] = None,
                                    action_type: Optional[str] = None,
                                    _: dict = Depends(require_roles("admin", "hr"))):
    """Silinen izin kayıtları (audit_log). Tekli + toplu silme kayıtları birlikte."""
    filt = await _build_deleted_leaves_filter(from_date, to_date, user_email, action_type)
    total = await db.audit_log.count_documents(filt)
    limit = max(1, min(int(limit), 500))
    skip = max(0, int(skip))
    items = []
    async for a in db.audit_log.find(filt, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit):
        ov = a.get("old_values") or {}
        row = {
            "id": a.get("id"),
            "timestamp": a.get("timestamp"),
            "user_email": a.get("user_email"),
            "user_name": a.get("user_name"),
            "action": a.get("action"),
            "action_label": "Toplu Silme" if a.get("action") == "bulk_delete" else "Tekli Silme",
            "sicil_no": ov.get("sicil_no"),
            "ad_soyad": ov.get("ad_soyad") or a.get("entity_name"),
            "start_date": ov.get("start_date"),
            "end_date": ov.get("end_date"),
            "days": ov.get("days"),
            "izin_turu": ov.get("izin_turu"),
            "reason": ov.get("reason"),
            "balance_before": ov.get("balance_before"),
            "balance_after": ov.get("balance_after"),
            "bulk_count": ov.get("count"),
            "bulk_total_days": ov.get("total_days_restored"),
            "affected_personnel": ov.get("affected_personnel"),
            "description": a.get("description"),
            "audit": a,  # detay modal için full JSON
        }
        items.append(row)
    return {"total": total, "items": items}


class RestoreDeletedIn(BaseModel):
    audit_id: str
    force: bool = False  # gelecek — çakışmaları görmezden gelme kullanılmıyor
    leave_index: Optional[int] = None  # bulk delete için tek bir kaydı geri yükleme


@api.post("/reports/deleted-leaves/restore")
async def restore_deleted_leave(body: RestoreDeletedIn, request: Request,
                                  user: dict = Depends(require_roles("admin", "hr"))):
    """Audit kaydından silinen izin(ler)i DB'ye geri yükler.
    - Tekli silme: eski leave dokümanını yeniden insert eder.
    - Toplu silme: old_values.deleted_leaves[] içindeki her kaydı ayrı ayrı denenir.
    - Ön kontrol: aynı personel + tarih aralığında ÇAKIŞMA veya kayıt zaten var mı.
    """
    a = await db.audit_log.find_one({"id": body.audit_id, "module": "leaves"}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Silme kaydı bulunamadı")
    if a.get("action") not in ("delete", "bulk_delete"):
        raise HTTPException(status_code=400, detail="Bu audit kaydı silme işlemi değil")
    ov = a.get("old_values") or {}

    def build_leaf_from_audit(personnel_id, start_date, end_date, days, izin_turu, aciklama=""):
        return {
            "id": str(uuid.uuid4()),
            "personnel_id": personnel_id,
            "start_date": start_date, "end_date": end_date,
            "days": float(days), "izin_turu": izin_turu or "Yıllık İzin",
            "aciklama": aciklama or "Geri yüklenen kayıt",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }

    async def try_restore(leaf: dict) -> dict:
        """Tek bir izin dokümanını geri yüklemeyi dener. Sonuç: {status, message, leave_id}."""
        pid = leaf["personnel_id"]
        s = leaf["start_date"]; e = leaf["end_date"]
        p = await db.personnel.find_one({"id": pid}, {"_id": 0})
        if not p:
            return {"status": "error", "message": f"Personel bulunamadı: {pid}"}
        if not p.get("aktif", True):
            return {"status": "warning", "message": f"Personel pasif: {p.get('ad_soyad')}",
                    "personnel": p.get("ad_soyad")}
        overlap = await db.leaves.find_one({
            "personnel_id": pid,
            "start_date": {"$lte": e}, "end_date": {"$gte": s},
        }, {"_id": 0})
        if overlap:
            return {"status": "conflict",
                    "message": f"Çakışma: {overlap['start_date']} → {overlap['end_date']} ({overlap.get('izin_turu')})",
                    "personnel": p.get("ad_soyad"), "conflict_with": overlap["id"]}
        await db.leaves.insert_one(leaf)
        bal = await _balance_for(p)
        return {"status": "restored", "leave_id": leaf["id"],
                "personnel": p.get("ad_soyad"), "sicil_no": p.get("sicil_no"),
                "days": leaf["days"], "izin_turu": leaf["izin_turu"],
                "start_date": s, "end_date": e,
                "balance_after": bal.get("remaining")}

    results = []
    if a["action"] == "delete":
        # Tekli silme audit'inde personnel_id + start/end/days + izin_turu var
        pid = ov.get("personnel_id")
        if not pid:
            # eski format — entity_id'den personel bulunmuyor; hata dön
            raise HTTPException(status_code=400,
                detail="Bu silme kaydı geri yüklemeye uygun bilgi içermiyor (eski format)")
        leaf = build_leaf_from_audit(pid, ov.get("start_date"), ov.get("end_date"),
                                       ov.get("days", 0), ov.get("izin_turu"))
        r = await try_restore(leaf)
        results.append(r)
    else:
        # Toplu silme — old_values.deleted_leaves[] listesinden
        deleted_leaves = ov.get("deleted_leaves") or []
        if body.leave_index is not None:
            if body.leave_index < 0 or body.leave_index >= len(deleted_leaves):
                raise HTTPException(status_code=400, detail="Geçersiz kayıt indexi")
            deleted_leaves = [deleted_leaves[body.leave_index]]
        if not deleted_leaves:
            raise HTTPException(status_code=400, detail="Geri yüklenecek kayıt bulunamadı")
        for L in deleted_leaves:
            leaf = build_leaf_from_audit(L["personnel_id"], L["start"], L["end"],
                                           L["days"], L.get("turu"))
            r = await try_restore(leaf)
            results.append(r)

    restored = sum(1 for r in results if r["status"] == "restored")
    conflicts = sum(1 for r in results if r["status"] == "conflict")
    errors = sum(1 for r in results if r["status"] == "error")
    warnings = sum(1 for r in results if r["status"] == "warning")

    await _audit(action="restore", module="leaves", entity_type="leave",
                 entity_id=body.audit_id, entity_name=f"Silme geri alma: {restored} kayıt",
                 old_values={"source_audit_id": body.audit_id, "action_source": a["action"]},
                 new_values={"restored": restored, "conflicts": conflicts,
                             "errors": errors, "warnings": warnings, "results": results},
                 description=f"Silme geri alma: +{restored} geri yüklendi, {conflicts} çakışma, {errors} hata",
                 request=request, user=user)

    return {"restored": restored, "conflicts": conflicts, "errors": errors,
            "warnings": warnings, "results": results,
            "audit_id": body.audit_id, "action_source": a["action"]}


@api.post("/reports/deleted-leaves/restore-preview")
async def restore_deleted_leave_preview(body: RestoreDeletedIn,
                                         _: dict = Depends(require_roles("admin", "hr"))):
    """Dry-run: silme audit kaydından geri yükleme öncesi çakışma/personel/bakiye kontrolü.
    Hiçbir yazma yapmaz — sadece analiz döner."""
    a = await db.audit_log.find_one({"id": body.audit_id, "module": "leaves"}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Silme kaydı bulunamadı")
    if a.get("action") not in ("delete", "bulk_delete"):
        raise HTTPException(status_code=400, detail="Bu audit kaydı silme işlemi değil")
    ov = a.get("old_values") or {}

    async def preview_one(pid, s, e, days, izin_turu, idx=None) -> dict:
        p = await db.personnel.find_one({"id": pid}, {"_id": 0})
        out = {
            "leave_index": idx,
            "personnel_id": pid,
            "start_date": s, "end_date": e,
            "days": float(days or 0),
            "izin_turu": izin_turu,
            "sicil_no": (p or {}).get("sicil_no"),
            "personnel": (p or {}).get("ad_soyad"),
            "personnel_active": bool((p or {}).get("aktif", False)) if p else False,
        }
        if not p:
            out["status"] = "error"; out["message"] = "Personel bulunamadı"
            return out
        overlap = await db.leaves.find_one({
            "personnel_id": pid,
            "start_date": {"$lte": e}, "end_date": {"$gte": s},
        }, {"_id": 0})
        bal = await _balance_for(p)
        out["balance_current"] = bal.get("remaining")
        out["balance_after_restore"] = round(float(bal.get("remaining", 0)) - float(days or 0), 2)
        if overlap:
            out["status"] = "conflict"
            out["message"] = (f"Çakışma: {overlap['start_date']} → {overlap['end_date']}"
                              f" ({overlap.get('izin_turu') or '-'})")
            out["conflict_with"] = {"id": overlap["id"], "start": overlap["start_date"],
                                     "end": overlap["end_date"], "turu": overlap.get("izin_turu"),
                                     "days": overlap.get("days")}
            return out
        if not p.get("aktif", True):
            out["status"] = "warning"
            out["message"] = "Personel pasif — geri yükleme uyarı verir"
            return out
        out["status"] = "ok"; out["message"] = "Geri yüklemeye uygun"
        return out

    results = []
    if a["action"] == "delete":
        pid = ov.get("personnel_id")
        if not pid:
            raise HTTPException(status_code=400,
                detail="Bu silme kaydı geri yüklemeye uygun bilgi içermiyor (eski format)")
        r = await preview_one(pid, ov.get("start_date"), ov.get("end_date"),
                                ov.get("days", 0), ov.get("izin_turu"))
        results.append(r)
    else:
        deleted_leaves = ov.get("deleted_leaves") or []
        if body.leave_index is not None:
            if body.leave_index < 0 or body.leave_index >= len(deleted_leaves):
                raise HTTPException(status_code=400, detail="Geçersiz kayıt indexi")
            r = await preview_one(deleted_leaves[body.leave_index]["personnel_id"],
                                    deleted_leaves[body.leave_index]["start"],
                                    deleted_leaves[body.leave_index]["end"],
                                    deleted_leaves[body.leave_index]["days"],
                                    deleted_leaves[body.leave_index].get("turu"),
                                    idx=body.leave_index)
            results.append(r)
        else:
            for i, L in enumerate(deleted_leaves):
                r = await preview_one(L["personnel_id"], L["start"], L["end"],
                                        L["days"], L.get("turu"), idx=i)
                results.append(r)

    ok = sum(1 for r in results if r["status"] == "ok")
    conflicts = sum(1 for r in results if r["status"] == "conflict")
    warnings = sum(1 for r in results if r["status"] == "warning")
    errors = sum(1 for r in results if r["status"] == "error")
    return {"ok": ok, "conflicts": conflicts, "warnings": warnings, "errors": errors,
             "total": len(results), "results": results,
             "audit_id": body.audit_id, "action_source": a["action"]}


@api.get("/reports/deleted-leaves/users")
async def deleted_leaves_users(_: dict = Depends(require_roles("admin", "hr"))):
    """Silme yapmış kullanıcıların listesi (filtre dropdown için)."""
    emails = await db.audit_log.distinct(
        "user_email", {"module": "leaves", "action": {"$in": ["delete", "bulk_delete"]}})
    return {"users": sorted([e for e in emails if e])}


@api.get("/reports/deleted-leaves/export.xlsx")
async def deleted_leaves_export(from_date: Optional[str] = None,
                                  to_date: Optional[str] = None,
                                  user_email: Optional[str] = None,
                                  action_type: Optional[str] = None,
                                  _: dict = Depends(require_roles("admin", "hr"))):
    """Silinen izinler filtreli xlsx dışa aktarma."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    filt = await _build_deleted_leaves_filter(from_date, to_date, user_email, action_type)
    wb = Workbook(); ws = wb.active; ws.title = "Silinen İzinler"
    headers = ["Tarih", "Kullanıcı", "Tür", "Sicil", "Ad Soyad", "İzin Türü",
               "Başlangıç", "Bitiş", "Gün", "Bakiye Önce", "Bakiye Sonra", "Gerekçe", "Açıklama"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="B91C1C")
        c.alignment = Alignment(horizontal="center")
    r = 2
    async for a in db.audit_log.find(filt, {"_id": 0}).sort("timestamp", -1):
        ov = a.get("old_values") or {}
        is_bulk = a.get("action") == "bulk_delete"
        ws.cell(row=r, column=1, value=(a.get("timestamp") or "")[:19].replace("T", " "))
        ws.cell(row=r, column=2, value=f"{a.get('user_name', '')} <{a.get('user_email', '')}>")
        ws.cell(row=r, column=3, value=f"Toplu ({ov.get('count', '')})" if is_bulk else "Tekli")
        ws.cell(row=r, column=4, value=ov.get("sicil_no") or "")
        ws.cell(row=r, column=5, value=ov.get("ad_soyad") or a.get("entity_name") or "")
        ws.cell(row=r, column=6, value=ov.get("izin_turu") or "")
        ws.cell(row=r, column=7, value=ov.get("start_date") or "")
        ws.cell(row=r, column=8, value=ov.get("end_date") or "")
        ws.cell(row=r, column=9, value=ov.get("total_days_restored") if is_bulk else ov.get("days"))
        ws.cell(row=r, column=10, value=ov.get("balance_before") if not is_bulk else "")
        ws.cell(row=r, column=11, value=ov.get("balance_after") if not is_bulk else "")
        ws.cell(row=r, column=12, value=ov.get("reason") or "")
        ws.cell(row=r, column=13, value=a.get("description") or "")
        r += 1
    for c, w in enumerate([18, 30, 14, 12, 26, 14, 12, 12, 8, 12, 12, 30, 50], 1):
        ws.column_dimensions[chr(64 + c)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"silinen_izinler_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@api.post("/personnel/sync/report.xlsx")
async def personnel_sync_report_xlsx(body: dict,
                                       _: dict = Depends(require_roles("admin", "hr"))):
    """Sync sonuç raporunu xlsx olarak indir. body = sync sonucu JSON'ı."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Özet"
    ws.append(["Kategori", "Adet"])
    s = body.get("summary", {})
    for k, v in [("Excel Toplam", s.get("excel_total")), ("DB Aktif", s.get("db_active")),
                  ("Eşleşen", s.get("matched")), ("Bilgisi Değişen", s.get("changed")),
                  ("Yeni Personel", s.get("new")), ("Listede Yok (Aktif)", s.get("missing")),
                  ("Mükerrer Sicil", s.get("duplicates")), ("Eksik Sicil", s.get("empty_sicil"))]:
        ws.append([k, v or 0])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1D4ED8")

    def add_sheet(title, rows, header, key_fn):
        w = wb.create_sheet(title)
        w.append(header)
        for c in w[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1D4ED8")
            c.alignment = Alignment(horizontal="center")
        for r in rows or []:
            w.append(key_fn(r))
        w.freeze_panes = "A2"

    add_sheet("Yeni Personel", body.get("new_personnel"),
              ["Sicil", "Ad Soyad", "TC", "Departman", "Şirket", "İşe Giriş"],
              lambda r: [r.get("sicil_no"), r.get("ad_soyad"), r.get("tc_no"),
                          r.get("departman"), r.get("sirket"), r.get("ise_giris")])
    add_sheet("Bilgisi Değişen", body.get("changed"),
              ["Sicil", "Ad Soyad", "Alan", "Eski", "Yeni"],
              lambda r: [r.get("sicil_no"), r.get("ad_soyad"), "|".join(r.get("diffs", {}).keys()),
                          "; ".join(f"{k}={v.get('old')}" for k, v in r.get("diffs", {}).items()),
                          "; ".join(f"{k}={v.get('new')}" for k, v in r.get("diffs", {}).items())])
    add_sheet("Listede Yok (İşten Ayrılış Adayı)", body.get("missing_from_file"),
              ["Sicil", "Ad Soyad", "Departman", "İşe Giriş"],
              lambda r: [r.get("sicil_no"), r.get("ad_soyad"), r.get("departman"), r.get("ise_giris")])
    add_sheet("Eşleşen", body.get("matched"),
              ["Sicil", "Ad Soyad"],
              lambda r: [r.get("sicil_no"), r.get("ad_soyad")])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"personel_senkronizasyon_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# -----------------------------------------------------------------------------
# Charts data
# -----------------------------------------------------------------------------
@api.get("/reports/charts")
async def charts_data(_: dict = Depends(get_current_user)):
    # Departman dağılımı
    dept_map = {}
    async for p in db.personnel.find({"aktif": True}, {"_id": 0, "departman": 1}):
        d = p.get("departman") or "Belirsiz"
        dept_map[d] = dept_map.get(d, 0) + 1
    dept = [{"name": k, "value": v} for k, v in sorted(dept_map.items(), key=lambda x: -x[1])]

    # Aylık izin trendi (son 12 ay)
    today = date.today()
    months = []
    for i in range(11, -1, -1):
        y = today.year; m = today.month - i
        while m <= 0: m += 12; y -= 1
        months.append((y, m))
    trend_map = {f"{y:04d}-{m:02d}": 0.0 for y, m in months}
    trend_people: dict = {f"{y:04d}-{m:02d}": set() for y, m in months}
    first = f"{months[0][0]:04d}-{months[0][1]:02d}-01"
    async for L in db.leaves.find({"start_date": {"$gte": first}}, {"_id": 0, "start_date": 1, "days": 1, "personnel_id": 1}):
        key = L["start_date"][:7]
        if key in trend_map:
            trend_map[key] += float(L.get("days", 0))
            trend_people[key].add(L.get("personnel_id"))
    labels_tr = ["Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
    trend = [{"month": f"{labels_tr[m - 1]} {y}",
              "days": round(trend_map[f"{y:04d}-{m:02d}"], 2),
              "people": len(trend_people[f"{y:04d}-{m:02d}"])}
             for y, m in months]

    # Şirket dağılımı
    comp_map = {}
    async for p in db.personnel.find({"aktif": True}, {"_id": 0, "sirket": 1}):
        s = p.get("sirket") or "Belirsiz"
        comp_map[s] = comp_map.get(s, 0) + 1
    company = [{"name": k, "value": v} for k, v in sorted(comp_map.items(), key=lambda x: -x[1])]

    # Kalan izin dağılımı (aktif personel)
    buckets = {"0-5": 0, "6-10": 0, "11-15": 0, "16-20": 0, "20+": 0}
    async for p in db.personnel.find({"aktif": True}, {"_id": 0}):
        b = await _balance_for(p)
        r = b["remaining"]
        if r <= 5: buckets["0-5"] += 1
        elif r <= 10: buckets["6-10"] += 1
        elif r <= 15: buckets["11-15"] += 1
        elif r <= 20: buckets["16-20"] += 1
        else: buckets["20+"] += 1
    remaining = [{"name": k, "value": v} for k, v in buckets.items()]

    return {"departments": dept, "monthly_trend": trend, "companies": company, "remaining_dist": remaining}

# -----------------------------------------------------------------------------
# Reports (Dashboard stats + exports)
# -----------------------------------------------------------------------------
@api.get("/dashboard/summary")
async def dashboard_summary(_: dict = Depends(get_current_user)):
    """Panel özet kartları için tek endpoint. Alan sözleşmesi:
    - total_active_personnel: aktif personel sayısı
    - today_on_leave: bugün izinli benzersiz aktif personel sayısı
    - total_remaining_leave: aktif personellerin kalan yıllık izin toplamı (0.5'ler korunur)
    - over_20_leave_count: kalanı > 20 olan aktif personel sayısı
    """
    today_d = date.today()
    today_iso = today_d.isoformat()

    # Aktif personel + id set (izin taramasında filtreleme için)
    active_ids: set = set()
    total_active = 0
    async for p in db.personnel.find({"aktif": True}, {"_id": 0, "id": 1}):
        active_ids.add(p["id"])
        total_active += 1

    # Bugün izinli (benzersiz personel, sadece aktifler)
    today_pids: set = set()
    async for L in db.leaves.find(
        {"start_date": {"$lte": today_iso}, "end_date": {"$gte": today_iso}},
        {"_id": 0, "personnel_id": 1},
    ):
        pid = L.get("personnel_id")
        if pid in active_ids:
            today_pids.add(pid)

    # Toplam kalan izin + over-20 count
    total_remaining = 0.0
    over_20 = 0
    async for p in db.personnel.find({"aktif": True}, {"_id": 0}):
        bal = await _compute_entitlements(p)
        rem = float(bal.get("remaining") or 0)
        total_remaining += rem
        if rem >= 20:
            over_20 += 1

    return {
        "total_active_personnel": total_active,
        "today_on_leave": len(today_pids),
        "total_remaining_leave": round(total_remaining, 2),
        "over_20_leave_count": over_20,
    }


@api.get("/dashboard/over-20")
async def dashboard_over_20(limit: int = 100, skip: int = 0,
                             _: dict = Depends(get_current_user)):
    """Kalan izni 20 gün ve üzerinde olan aktif personel listesi (paginated).
    Iter 61: kriter `>= 20` olarak güncellendi (20, 20.5, 21+ dahil; 19.5 dahil değil).
    Response: {total, items[{id, sicil_no, ad_soyad, departman, remaining}]}
    """
    rows: List[dict] = []
    async for p in db.personnel.find({"aktif": True}, {"_id": 0}):
        bal = await _compute_entitlements(p)
        rem = float(bal.get("remaining") or 0)
        if rem >= 20:
            rows.append({
                "id": p["id"], "sicil_no": p.get("sicil_no"),
                "ad_soyad": p.get("ad_soyad"), "departman": p.get("departman"),
                "remaining": rem,
            })
    rows.sort(key=lambda r: -r["remaining"])
    total = len(rows)
    limit = max(1, min(int(limit), 500))
    skip = max(0, int(skip))
    return {"total": total, "items": rows[skip:skip + limit]}


@api.get("/dashboard")
async def dashboard(_: dict = Depends(get_current_user)):
    total = await db.personnel.count_documents({})
    active = await db.personnel.count_documents({"aktif": True})
    left = await db.personnel.count_documents({"aktif": False})
    today = date.today().isoformat()
    on_leave = await db.leaves.count_documents({"start_date": {"$lte": today}, "end_date": {"$gte": today}})
    # new hires this month
    this_month = date.today().replace(day=1).isoformat()
    new_this_month = await db.personnel.count_documents({"ise_giris": {"$gte": this_month}})
    # low balance list + upcoming entitlements (within 7 days)
    low = []
    upcoming: list = []
    today_d = date.today()
    week_ahead = today_d + timedelta(days=7)
    async for p in db.personnel.find({"aktif": True}, {"_id": 0}):
        bal = await _compute_entitlements(p)
        if bal["entitled_total"] > 0 and bal["remaining"] < 10:
            low.append({"personnel": p, "remaining": bal["remaining"], "entitled": bal["entitled_total"], "used": bal["used_total"]})
        ne = bal.get("next_entitlement")
        if ne:
            ne_date = _parse_date(ne["date"])
            if ne_date and today_d <= ne_date <= week_ahead:
                upcoming.append({
                    "personnel": p,
                    "date": ne["date"],
                    "days_until": (ne_date - today_d).days,
                    "leave_days": ne["days"],
                    "total_seniority": ne["total_seniority"],
                })
    low.sort(key=lambda x: x["remaining"])
    upcoming.sort(key=lambda x: x["days_until"])
    return {
        "total": total, "active": active, "left": left,
        "on_leave_today": on_leave, "new_this_month": new_this_month,
        "low_balance": low[:20],
        "upcoming_entitlements": upcoming,
    }

@api.get("/reports/summary")
async def report_summary(_: dict = Depends(get_current_user)):
    items = []
    async for p in db.personnel.find({}, {"_id": 0}).sort("ad_soyad", 1):
        bal = await _balance_for(p)
        items.append({
            "sicil_no": p["sicil_no"], "ad_soyad": p["ad_soyad"],
            "departman": p.get("departman", ""), "sirket": p.get("sirket", ""),
            "ise_giris": p["ise_giris"], "aktif": p["aktif"],
            "entitled": bal["entitled_total"], "used": bal["used_total"], "remaining": bal["remaining"],
        })
    return items

def _xlsx(headers: list, rows: list, sheet_name: str = "Rapor") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for i, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()

def _pdf(title: str, headers: list, rows: list) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 8),
             Paragraph(f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles["Normal"]),
             Spacer(1, 10)]
    data = [headers] + [[str(c) if c is not None else "" for c in r] for r in rows]
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()

def _stream(data: bytes, filename: str, media: str):
    return StreamingResponse(io.BytesIO(data), media_type=media,
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@api.get("/reports/personnel-summary/export")
async def export_summary(format: str = "xlsx", _: dict = Depends(get_current_user)):
    data = await report_summary(_)
    headers = ["Sicil No", "Ad Soyad", "Departman", "Şirket", "İşe Giriş", "Durum",
               "Hak Edilen", "Kullanılan", "Kalan"]
    rows = [[d["sicil_no"], d["ad_soyad"], d["departman"], d["sirket"], d["ise_giris"],
             "Aktif" if d["aktif"] else "Ayrıldı", d["entitled"], d["used"], d["remaining"]] for d in data]
    if format == "pdf":
        return _stream(_pdf("Personel İzin Özeti", headers, rows), "personel_izin_ozeti.pdf", "application/pdf")
    return _stream(_xlsx(headers, rows, "İzin Özeti"), "personel_izin_ozeti.xlsx",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@api.get("/reports/by-department/export")
async def export_dept(format: str = "xlsx", _: dict = Depends(get_current_user)):
    headers = ["Departman", "Sicil No", "Ad Soyad", "Görev", "Şirket", "İşe Giriş", "Durum"]
    rows = []
    async for p in db.personnel.find({}, {"_id": 0}).sort([("departman", 1), ("ad_soyad", 1)]):
        rows.append([p.get("departman", ""), p["sicil_no"], p["ad_soyad"],
                     p.get("gorev", ""), p.get("sirket", ""), p["ise_giris"],
                     "Aktif" if p["aktif"] else "Ayrıldı"])
    if format == "pdf":
        return _stream(_pdf("Departman Bazlı Personel Listesi", headers, rows), "departman_listesi.pdf", "application/pdf")
    return _stream(_xlsx(headers, rows, "Departman"), "departman_listesi.xlsx",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@api.get("/reports/leavers/export")
async def export_leavers(format: str = "xlsx", _: dict = Depends(get_current_user)):
    headers = ["Sicil No", "Ad Soyad", "Departman", "Şirket", "İşe Giriş", "İşten Çıkış"]
    rows = []
    async for p in db.personnel.find({"aktif": False}, {"_id": 0}).sort("isten_cikis", -1):
        rows.append([p["sicil_no"], p["ad_soyad"], p.get("departman", ""),
                     p.get("sirket", ""), p["ise_giris"], p.get("isten_cikis", "")])
    if format == "pdf":
        return _stream(_pdf("İşten Ayrılanlar", headers, rows), "isten_ayrilanlar.pdf", "application/pdf")
    return _stream(_xlsx(headers, rows, "İşten Ayrılanlar"), "isten_ayrilanlar.xlsx",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@api.get("/reports/new-hires/export")
async def export_new(format: str = "xlsx", days: int = 90, _: dict = Depends(get_current_user)):
    cutoff = (date.today() - timedelta(days=days)).isoformat()
    headers = ["Sicil No", "Ad Soyad", "Departman", "Şirket", "İşe Giriş"]
    rows = []
    async for p in db.personnel.find({"ise_giris": {"$gte": cutoff}}, {"_id": 0}).sort("ise_giris", -1):
        rows.append([p["sicil_no"], p["ad_soyad"], p.get("departman", ""),
                     p.get("sirket", ""), p["ise_giris"]])
    if format == "pdf":
        return _stream(_pdf(f"Son {days} Gün Yeni Başlayanlar", headers, rows), "yeni_baslayanlar.pdf", "application/pdf")
    return _stream(_xlsx(headers, rows, "Yeni Başlayanlar"), "yeni_baslayanlar.xlsx",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@api.get("/reports/leaves-in-range/export")
async def export_leaves_range(start: str, end: str, format: str = "xlsx",
                               _: dict = Depends(get_current_user)):
    headers = ["Sicil No", "Ad Soyad", "Departman", "Başlangıç", "Bitiş", "Gün", "İzin Türü", "Açıklama"]
    rows = []
    async for L in db.leaves.find({"start_date": {"$gte": start, "$lte": end}}, {"_id": 0}).sort("start_date", 1):
        p = await db.personnel.find_one({"id": L["personnel_id"]}, {"_id": 0}) or {}
        rows.append([p.get("sicil_no", ""), p.get("ad_soyad", ""), p.get("departman", ""),
                     L["start_date"], L["end_date"], L["days"],
                     L.get("izin_turu", ""), L.get("aciklama", "")])
    if format == "pdf":
        return _stream(_pdf(f"{start} — {end} İzin Kullanımları", headers, rows), "izinler.pdf", "application/pdf")
    return _stream(_xlsx(headers, rows, "İzinler"), "izinler.xlsx",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# -----------------------------------------------------------------------------
# Excel template-driven document generation (BASI OLARAK Excel'i birebir koruyor)
# -----------------------------------------------------------------------------
TEMPLATE_PATH = ROOT_DIR / "templates" / "izin_template.xlsm"

def _keep_only_sheets(wb, keep_visible: str, keep_ref: list = None):
    """Hedef sayfa dışındakileri workbook'tan sil. Formül referansları için gerekli sayfaları gizli tut."""
    keep_ref = keep_ref or []
    keep_all = {keep_visible} | set(keep_ref)
    for sn in list(wb.sheetnames):
        if sn not in keep_all:
            del wb[sn]
    # Formül referans sayfalarını gizle
    for sn in keep_ref:
        if sn in wb.sheetnames:
            wb[sn].sheet_state = "hidden"
    # Hedefi görünür + aktif yap
    if keep_visible in wb.sheetnames:
        wb[keep_visible].sheet_state = "visible"
        wb.active = wb.sheetnames.index(keep_visible)

async def _fill_talep_formu(personnel: dict, leave: dict, hak_edis_tarihi: Optional[str] = None,
                              talep_tarihi_iso: Optional[str] = None) -> bytes:
    """Excel şablonunu birebir koruyarak sadece değişken hücreleri doldurur.
    Ek olarak: A1'deki #VALUE! kaldırılır, tüm tarih hücreleri GG.AA.YYYY formatında yazılır,
    Excel formüllerinin ürettiği hatalı yer tutucular gerçek metinlerle değiştirilir.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from datetime import datetime as _dt
    wb = load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb["İZİN TALEP FORMU"]

    def _d(iso: Optional[str]):
        if not iso: return None
        try: return _dt.fromisoformat(iso[:10]).date()
        except Exception: return None

    def _tr(iso: Optional[str]) -> str:
        d = _d(iso)
        return d.strftime("%d.%m.%Y") if d else ""

    # A1: şirket adı ile #VALUE! hatasını değiştir
    ws["A1"] = "MERKOTEKS TEKSTİL SAN. VE TİC. A.Ş."
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].font = Font(bold=True, size=9)

    # Doküman meta — Düzenleme Tarihi (H2) GG.AA.YYYY formatı
    ws["H2"] = _dt(2023, 1, 2).date()
    ws["H2"].number_format = "DD.MM.YYYY"

    ise_giris = _d(personnel.get("ise_giris"))
    izin_bas = _d(leave.get("start_date"))
    izin_bit = _d(leave.get("end_date"))
    hak_dt = _d(hak_edis_tarihi)

    # Sadece değişken hücreler
    ws["B8"] = personnel.get("ad_soyad", "")
    ws["B9"] = personnel.get("departman", "") or personnel.get("bolum", "")
    if ise_giris:
        ws["B10"] = ise_giris; ws["B10"].number_format = "DD.MM.YYYY"
    ws["B11"] = personnel.get("sicil_no", "")
    if hak_dt:
        ws["B12"] = hak_dt; ws["B12"].number_format = "DD.MM.YYYY"
        # Iter 28: Yıllık İzni Hak Ettiği Tarih hücresini ortala + 12 punto
        _b12 = ws["B12"]
        _b12.alignment = Alignment(horizontal="center", vertical="center")
        _f = _b12.font
        _b12.font = Font(name=_f.name, size=12, bold=_f.bold, italic=_f.italic,
                          color=_f.color)
    if izin_bas:
        ws["B13"] = izin_bas; ws["B13"].number_format = "DD.MM.YYYY"
    if izin_bit:
        ws["B14"] = izin_bit; ws["B14"].number_format = "DD.MM.YYYY"
    # B15 İzin gün sayısı — Türkçe ondalık virgül (20,5)
    ws["B15"] = _fmt_tr_num(leave.get("days", 0))
    ws["B15"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B17"] = personnel.get("telefon", "")
    ws["B18"] = personnel.get("adres", "") or personnel.get("aciklama", "")

    # E9: İnsan Kaynakları'na dönemim metni — gerçek tarihlerle
    if izin_bas and izin_bit:
        ws["E9"] = (
            f"Yıllık Ücretli İzin kullanma dönemim olan "
            f"{izin_bas.strftime('%d.%m.%Y')} ile {izin_bit.strftime('%d.%m.%Y')} tarihleri arasında "
            f'"Yıllık Ücretli İzin Talep Formu"nda beyan ettiğim günlerde kullanmak istiyorum. '
            f"Gereğini Arz Ederim.\n\nSaygılarımla"
        )
    # E14 / E15: personel adı + izin talep tarihi (leave.created_at'ten alınır — fallback: bugün)
    ws["E14"] = f"Personelin Adı : {personnel.get('ad_soyad', '')}"
    talep_d = _d(talep_tarihi_iso) if talep_tarihi_iso else None
    talep_tr = talep_d.strftime("%d.%m.%Y") if talep_d else _dt.now().strftime("%d.%m.%Y")
    ws["E15"] = f"İzin Talep Tarihi : {talep_tr}"
    ws["E15"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # B20 DÜŞÜNCE VE ONAY — gerçek tarihler
    holidays_map = await get_all_holidays()
    isbasi = _next_working_day(izin_bit, holidays_map) if izin_bit else None
    if izin_bas and izin_bit and isbasi:
        ws["B20"] = (
            f"        Kimliği yukarıda yer alan personelimizin, yıllık ücretli izin hakkını "
            f"{izin_bas.strftime('%d.%m.%Y')} tarihinde ayrılmak ve "
            f"{isbasi.strftime('%d.%m.%Y')} tarihinde göreve başlamak kaydıyla kullanması uygundur."
        )
    # A28 taahhüt
    if izin_bas and izin_bit:
        ws["A28"] = (
            f"        Ücretli izin hakkımı {izin_bas.strftime('%d.%m.%Y')} Tarihi ile "
            f"{izin_bit.strftime('%d.%m.%Y')} Tarihleri arasında kullandım."
        )
    # B29 İşbaşı Tarihi — formülü kaldır, gerçek tarih yaz
    if isbasi:
        ws["B29"] = isbasi
        ws["B29"].number_format = "DD.MM.YYYY"

    # ONAY tarih alanları A26, C26, F26 — boş bırak (…../……./20..)  — imza sırasında elle yazılacak
    # Şablon aynen korunuyor

    _keep_only_sheets(wb, "İZİN TALEP FORMU", keep_ref=["Tatiller"])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()

async def _fill_izin_cetveli(personnel: dict, allocations: list) -> bytes:
    """İZİN CETVELİ — FIFO allocations kullanarak 15 kolonlu şablonu doldurur.

    Iter 23 kuralları:
      • Sade beyaz arka plan (renk YOK). Yıl bantlaması kaldırıldı.
      • Kolon B ("Bir yıl önceki izin hakkını kullandığı tarih") veri satırlarında BOŞ.
      • Aynı izin gerçekten hak ediş anniversary geçtiği için birden çok slice'a
        bölünmüşse, sınır satırların Düşünceler'e "-" yazılır. Normal tek-parça
        izinler için Düşünceler boş.
      • 12 satırdan fazla allocation varsa yeni satırlar row 20'nin stilini
        kopyalayarak eklenir. ws.print_title_rows="1:8" ile üst başlık her
        yazdırılan sayfada tekrarlanır.
    """
    from openpyxl import load_workbook
    from copy import copy
    from datetime import datetime as _dt
    wb = load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb["İZİN CETVELİ"]

    def _d(iso: Optional[str]):
        if not iso: return None
        try: return _dt.fromisoformat(iso[:10]).date()
        except Exception: return None

    ws["B3"] = personnel.get("ad_soyad", "")
    ws["H3"] = personnel.get("sicil_no", "")
    ise = _d(personnel.get("ise_giris"))
    if ise:
        ws["O3"] = ise
        ws["O3"].number_format = "DD.MM.YYYY"

    # Iter 27: "Bir yıllık çalışma süresi bakımından kesilmeler" başlık hücresi
    # (C5:H5) TEK büyük hücre gibi görünsün — iç dikey çizgileri kaldır.
    from openpyxl.styles import Side, Border, Alignment, Font
    no_side = Side(style=None)
    for col_idx in range(3, 9):  # C..H
        cell = ws.cell(row=5, column=col_idx)
        old = cell.border
        cell.border = Border(
            top=old.top if col_idx == 3 else old.top,
            bottom=old.bottom,
            left=old.left if col_idx == 3 else no_side,
            right=old.right if col_idx == 8 else no_side,
        )
    # Ayrıca birleştir (C5:H5) — tek hücre görünümü
    try:
        ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=8)
    except Exception:
        pass
    # Iter 28: birleştirilmiş hücrenin metnini yatay + dikey ortala
    c5 = ws.cell(row=5, column=3)
    # Fazla boşluk temizle
    if isinstance(c5.value, str):
        c5.value = c5.value.strip()
    c5.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Şablon veri satırları A9:P20 (12 satır). Önce hepsini boşalt (renk kaldır).
    from openpyxl.styles import PatternFill
    default_fill = PatternFill(fill_type=None)
    template_row_r = 20  # stilin referans olarak alınacağı satır
    for r in range(9, 21):
        for col_idx in range(1, 17):
            try:
                cell = ws.cell(row=r, column=col_idx)
                cell.value = None
                cell.fill = default_fill
            except Exception:
                pass

    sorted_allocs = sorted(allocations, key=lambda x: (x.get("slice_start", ""), x.get("entitlement_date", "")))
    # Aynı leave birden fazla slice'a bölünmüşse Düşünceler'e "-" işareti
    leave_slice_count: dict = {}
    for a in sorted_allocs:
        lid = a.get("leave_id")
        leave_slice_count[lid] = leave_slice_count.get(lid, 0) + 1

    total = len(sorted_allocs)
    template_slots = 12

    # 12'den fazlaysa ek satırları row 20'nin stilini kopyalayarak ekle
    if total > template_slots:
        extra = total - template_slots
        # row 20 sonrasına satır ekle ve stilleri kopyala
        for i in range(extra):
            new_r = 21 + i
            # stil kopyalama (kenarlık, font, alignment) row 20'den
            for col_idx in range(1, 17):
                src = ws.cell(row=template_row_r, column=col_idx)
                dst = ws.cell(row=new_r, column=col_idx)
                dst.value = None
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.border = copy(src.border)
                    dst.alignment = copy(src.alignment)
                    dst.number_format = src.number_format
                    dst.fill = default_fill
        # Satır yüksekliğini row 20'den al
        h20 = ws.row_dimensions[template_row_r].height
        for i in range(extra):
            new_r = 21 + i
            if h20:
                ws.row_dimensions[new_r].height = h20

    for i, a in enumerate(sorted_allocs):
        r = 9 + i
        yr = a.get("entitlement_year")
        # A: YIL
        ws.cell(row=r, column=1, value=yr)
        # B: BOŞ (Iter 23 — kolon B veri satırlarında doldurulmaz)
        # I: İzne Hak Kazandığı Tarih
        ed = _d(a.get("entitlement_date", ""))
        if ed:
            c = ws.cell(row=r, column=9, value=ed); c.number_format = "DD.MM.YYYY"
        # J: İşyerindeki Kıdemi
        ws.cell(row=r, column=10, value=f"{a.get('seniority_at', 0)} Yıl")
        # K: İzin Süresi
        ws.cell(row=r, column=11, value=_fmt_tr_num(a.get("days", 0)))
        # M: İzne Başladığı Tarih
        sd = _d(a.get("slice_start", ""))
        if sd:
            c = ws.cell(row=r, column=13, value=sd); c.number_format = "DD.MM.YYYY"
        # N: İzinden Dönüş Tarihi
        rd = _d(a.get("return_date", ""))
        if rd:
            c = ws.cell(row=r, column=14, value=rd); c.number_format = "DD.MM.YYYY"
        # O: Düşünceler — sadece gerçek anniversary bölünmelerinde "-"
        lid = a.get("leave_id")
        remark = "-" if leave_slice_count.get(lid, 1) > 1 else (a.get("aciklama") or "")
        ws.cell(row=r, column=15, value=remark)

    # Her yazdırılan sayfada üst başlık (1-8. satırlar) tekrarlansın
    ws.print_title_rows = "1:8"
    # Iter 30: NOT metnini şablonun sabit A22 satırından kaldırıp SAYFA ALT BİLGİSİNE
    # (page footer) taşı — böylece çok sayfalı cetvellerde her sayfanın altında görünür.
    NOT_TEXT = ("NOT: 1- İlk defa yıllık ücretli izin hakkından faydalandırılanlar için "
                "(Bir yıl önceki izin hakkını kullandığı tarih) sütunu boş bırakılacaktır.")
    # Şablondaki sabit not satırlarını temizle (aksi halde son sayfada içerik olarak da çıkar)
    for r in range(21, 25):
        for col_idx in range(1, 17):
            try:
                cell = ws.cell(row=r, column=col_idx)
                if isinstance(cell.value, str) and cell.value.strip().startswith("NOT"):
                    cell.value = None
            except Exception:
                pass
    # Page footer — LibreOffice PDF üretiminde her sayfanın altında görünür
    try:
        # Iter 31: NOT daha net görünsün — 10 punto + BOLD (&B prefix ile)
        ws.oddFooter.left.text = "&B" + NOT_TEXT
        ws.oddFooter.left.size = 10
        ws.oddFooter.left.color = "000000"
        ws.evenFooter.left.text = "&B" + NOT_TEXT
        ws.evenFooter.left.size = 10
        ws.evenFooter.left.color = "000000"
        # Iter 36: Sayfa numarası sadece "&P" — kullanıcı isteği ("1", "2", "3")
        ws.oddFooter.right.text = "&P"
        ws.oddFooter.right.size = 10
        ws.oddFooter.right.color = "000000"
        ws.evenFooter.right.text = "&P"
        ws.evenFooter.right.size = 10
        ws.evenFooter.right.color = "000000"
    except Exception:
        pass
    # Iter 31: Sayfa kenar boşluklarını dengele — footer için alt boşluk artırıldı
    from openpyxl.worksheet.page import PageMargins
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.9,
                                    header=0.3, footer=0.4)
    # A4 Landscape + sığdırma
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_options.horizontalCentered = True
    # Iter 36: Son sayfa (satırları az olan) dikey ortalanmasın — üstten hizalı kalsın.
    ws.print_options.verticalCentered = False
    try:
        ws.page_setup.verticalCentered = False
    except Exception:
        pass
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 = kaç sayfa gerekiyorsa

    _keep_only_sheets(wb, "İZİN CETVELİ")
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()

def _xlsx_to_pdf(xlsx_bytes: bytes, sheet_name: str) -> Optional[bytes]:
    """LibreOffice headless ile XLSX → PDF. Sadece hedef (gizli olmayan) sayfa render edilir."""
    import subprocess, tempfile, shutil
    soffice = shutil.which("libreoffice") or shutil.which("soffice")
    if not soffice:
        return None
    # Filtre: ExportHiddenSheets=false → gizli sayfalar PDF'e alınmaz
    pdf_filter = (
        'pdf:calc_pdf_Export:{'
        '"ExportHiddenSheets":{"type":"boolean","value":"false"},'
        '"SinglePageSheets":{"type":"boolean","value":"false"}'
        '}'
    )
    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "doc.xlsx"
        src.write_bytes(xlsx_bytes)
        try:
            subprocess.run(
                [soffice, "--headless", "--calc",
                 "--convert-to", pdf_filter, "--outdir", tmp, str(src)],
                check=True, capture_output=True, timeout=90,
            )
        except subprocess.CalledProcessError as e:
            log.error("LibreOffice PDF fail: %s / %s", e.stderr, e.stdout)
            return None
        pdf = Path(tmp) / "doc.pdf"
        if not pdf.exists():
            return None
        return pdf.read_bytes()

async def _resolve_hak_edis_for_leave(p: dict, L: dict) -> Optional[str]:
    """İzin Talep Formunun 'Yıllık İzin Hak Ettiği Tarih' alanı.

    Iter 23:
      A) Personelin gerçek allocation'ı varsa (kendi leave_id'sine ait, avans değil)
         → allocation'ın entitlement_date'i (birden fazlasa en eski).
      B) Personel henüz hak ediş kazanmamış veya sadece avans kullanıyorsa
         → gelecek ilk hak ediş tarihi (`bal.next_entitlement.date`).
      C) Hiç entitlement yoksa ama past ent varsa (recompute sonrası) → en yakın
         geçmiş entitlement (start_date <= ent.date).
    """
    bal = await _compute_entitlements(p)
    ents = sorted(bal.get("entitlements", []), key=lambda x: x["date"])
    all_leaves: list = []
    async for L2 in db.leaves.find({"personnel_id": p["id"]}, {"_id": 0}).sort("start_date", 1):
        all_leaves.append(L2)
    allocs = await allocate_leaves_fifo(ents, all_leaves, personnel=p)
    dates = [a.get("entitlement_date") for a in allocs
             if a.get("leave_id") == L.get("id") and not a.get("is_advance") and a.get("entitlement_date")]
    if dates:
        return min(dates)
    # Avans veya henüz hak ediş oluşmadıysa: gelecek ilk hak ediş tarihi
    nxt = bal.get("next_entitlement")
    if nxt and nxt.get("date"):
        return nxt["date"]
    # Fallback: leave.start_date <= entitlement en yenisi
    for e in ents:
        if e["date"] <= L["start_date"]:
            return e["date"]
    return None


@api.get("/leaves/{lid}/talep-formu.xlsx")
async def leave_form_xlsx(lid: str, _: dict = Depends(get_current_user)):
    L = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not L: raise HTTPException(status_code=404, detail="İzin bulunamadı")
    p = await db.personnel.find_one({"id": L["personnel_id"]}, {"_id": 0})
    if not p: raise HTTPException(status_code=404, detail="Personel bulunamadı")
    hak_edis = await _resolve_hak_edis_for_leave(p, L)
    data = await _fill_talep_formu(p, L, hak_edis_tarihi=hak_edis,
                                     talep_tarihi_iso=L.get("created_at"))
    return _stream(data, f"izin_talep_{p.get('sicil_no','')}.xlsx",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@api.get("/leaves/{lid}/talep-formu.pdf")
async def leave_form_pdf(lid: str, _: dict = Depends(get_current_user)):
    L = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not L: raise HTTPException(status_code=404, detail="İzin bulunamadı")
    p = await db.personnel.find_one({"id": L["personnel_id"]}, {"_id": 0})
    if not p: raise HTTPException(status_code=404, detail="Personel bulunamadı")
    hak_edis = await _resolve_hak_edis_for_leave(p, L)
    xlsx = await _fill_talep_formu(p, L, hak_edis_tarihi=hak_edis,
                                     talep_tarihi_iso=L.get("created_at"))
    pdf = _xlsx_to_pdf(xlsx, "İZİN TALEP FORMU")
    if pdf is None:
        raise HTTPException(status_code=503, detail="PDF dönüştürücü (LibreOffice) yüklenmedi. Şimdilik Excel çıktısını kullanın.")
    return _stream(pdf, f"izin_talep_{p.get('sicil_no','')}.pdf", "application/pdf")

@api.post("/admin/recompute-leave-days")
async def recompute_leave_days(_: dict = Depends(require_roles("admin"))):
    """Tüm izin kayıtlarının 'days' değerini güncel tatil takvimine göre yeniden hesapla.
    Örn. 28.10 Cumhuriyet Arifesi eklendi → mevcut 30.10.2026 biten kayıtlar 21 → 20,5 olur.
    """
    updated = []
    async for L in db.leaves.find({}, {"_id": 0}):
        s = _parse_date(L["start_date"]); e = _parse_date(L["end_date"])
        if not s or not e:
            continue
        calc = await calc_leave_days(s, e)
        new_days = float(calc["days"])
        old_days = float(L.get("days", 0))
        if abs(new_days - old_days) > 1e-9:
            await db.leaves.update_one(
                {"id": L["id"]},
                {"$set": {"days": new_days,
                          "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            updated.append({"leave_id": L["id"], "personnel_id": L["personnel_id"],
                            "start_date": L["start_date"], "end_date": L["end_date"],
                            "old_days": old_days, "new_days": new_days})
    return {"updated_count": len(updated), "updated": updated}


@api.get("/personnel/{pid}/cetveli.xlsx")
async def cetveli_xlsx(pid: str, _: dict = Depends(get_current_user)):
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p: raise HTTPException(status_code=404, detail="Personel bulunamadı")
    bal = await _compute_entitlements(p)
    leaves = []
    async for L in db.leaves.find({"personnel_id": pid}, {"_id": 0}).sort("start_date", 1):
        leaves.append(L)
    allocations = await allocate_leaves_fifo(bal.get("entitlements", []), leaves, personnel=p)
    data = await _fill_izin_cetveli(p, allocations)
    return _stream(data, f"izin_cetveli_{p.get('sicil_no','')}.xlsx",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@api.get("/personnel/{pid}/cetveli.pdf")
async def cetveli_pdf(pid: str, _: dict = Depends(get_current_user)):
    p = await db.personnel.find_one({"id": pid}, {"_id": 0})
    if not p: raise HTTPException(status_code=404, detail="Personel bulunamadı")
    bal = await _compute_entitlements(p)
    leaves = []
    async for L in db.leaves.find({"personnel_id": pid}, {"_id": 0}).sort("start_date", 1):
        leaves.append(L)
    allocations = await allocate_leaves_fifo(bal.get("entitlements", []), leaves, personnel=p)
    xlsx = await _fill_izin_cetveli(p, allocations)
    pdf = _xlsx_to_pdf(xlsx, "İZİN CETVELİ")
    if pdf is None:
        raise HTTPException(status_code=503, detail="PDF dönüştürücü (LibreOffice) yüklenmedi. Şimdilik Excel çıktısını kullanın.")
    return _stream(pdf, f"izin_cetveli_{p.get('sicil_no','')}.pdf", "application/pdf")

# -----------------------------------------------------------------------------
# Mount & CORS
# -----------------------------------------------------------------------------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    client.close()
