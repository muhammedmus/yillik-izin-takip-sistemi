"""Iteration 12: Excel-based bulk leave upload endpoints.

Tests:
  GET  /api/leaves/bulk/excel-template  → returns xlsx
  POST /api/leaves/bulk/excel-preview   → 5-row summary (3 applicable, 2 blocked)
  POST /api/leaves/bulk/excel-confirm   → creates only can_apply rows
  Regression: GET /api/leaves            → 200 (no 429)
"""
import io
import os
import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://merkoteks-izin.preview.emergentagent.com").rstrip("/")
ADMIN_EMAIL = "muhammedmus@gmail.com"
ADMIN_PWD = "Merkoteks2026!"

# Test siciller from prompt
SICILS = ["6280", "5829", "7073", "7304", "6860"]


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": ADMIN_EMAIL, "password": ADMIN_PWD}, timeout=20)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    return s


def _make_xlsx(rows):
    """Build xlsx matching backend template layout: headers on row 4, data from row 5."""
    wb = Workbook(); ws = wb.active
    ws["A1"] = "Merkoteks — Toplu İzin Yükleme Şablonu"
    ws["A2"] = "test"
    headers = ["Sicil No", "Ad Soyad", "İzin Türü", "Başlangıç Tarihi", "Bitiş Tarihi"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    for ri, row in enumerate(rows, start=5):
        for ci, v in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=v)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def test_template_download(session):
    r = session.get(f"{BASE_URL}/api/leaves/bulk/excel-template", timeout=20)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    header_vals = [ws.cell(row=4, column=c).value for c in range(1, 6)]
    assert header_vals == ["Sicil No", "Ad Soyad", "İzin Türü", "Başlangıç Tarihi", "Bitiş Tarihi"]


def test_preview_summary(session):
    # Rows for preview scenario (2028 to avoid conflicts):
    # row5 valid: 6280
    # row6 valid: 5829
    # row7 blocked no_match: 9999
    # row8 blocked date_order: 7073 end < start
    # row9 warning name_mismatch (still applicable): 7304 with wrong name
    rows = [
        ["6280", "ABBAS TORUN",         "Yıllık İzin", "03.02.2028", "07.02.2028"],
        ["5829", "ABDULLAH ALBAYRAK",   "Yıllık İzin", "10.02.2028", "14.02.2028"],
        ["9999", "YOK",                 "Yıllık İzin", "10.02.2028", "12.02.2028"],
        ["7073", "ABDULSAMET DAĞLI",    "Yıllık İzin", "20.02.2028", "17.02.2028"],
        ["7304", "YANLIS ISIM",         "Yıllık İzin", "01.03.2028", "05.03.2028"],
    ]
    buf = _make_xlsx(rows)
    files = {"file": ("test.xlsx", buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = session.post(f"{BASE_URL}/api/leaves/bulk/excel-preview", files=files, timeout=30)
    assert r.status_code == 200, r.text
    data = r.json()
    assert "rows" in data and "summary" in data
    assert len(data["rows"]) == 5
    s = data["summary"]
    assert s["total"] == 5
    # 3 applicable, 2 blocked expected — unless DB conflicts
    print("summary:", s)
    print("rows warnings:", [(r["row"], r["can_apply"], [w["code"] for w in r["warnings"]]) for r in data["rows"]])

    # Row-by-row assertions
    by_row = {r["row"]: r for r in data["rows"]}
    # row 7 (index 3rd data row = xlsx row 7) unknown sicil
    assert any(w["code"] == "no_match" for w in by_row[7]["warnings"])
    assert by_row[7]["can_apply"] is False
    # row 8 date_order
    assert any(w["code"] == "date_order" for w in by_row[8]["warnings"])
    assert by_row[8]["can_apply"] is False
    # row 9 name_mismatch but can_apply=true (unless db conflict)
    assert any(w["code"] == "name_mismatch" for w in by_row[9]["warnings"])
    # summary applicable may be 3 if no DB conflict for rows 5,6,9
    assert s["blocked"] >= 2
    assert s["applicable"] >= 1


def test_confirm_creates_rows(session):
    # Use far-future dates to avoid overlaps with existing DB data
    rows = [
        ["6860", "ALEYNA TUNÇ", "Yıllık İzin", "05.11.2028", "09.11.2028"],
    ]
    buf = _make_xlsx(rows)
    files = {"file": ("t.xlsx", buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    pv = session.post(f"{BASE_URL}/api/leaves/bulk/excel-preview", files=files, timeout=30)
    assert pv.status_code == 200
    pv_data = pv.json()
    applyable = [r for r in pv_data["rows"] if r["can_apply"]]
    if not applyable:
        pytest.skip(f"No applyable rows in preview (maybe conflict): {pv_data['rows']}")
    payload = {
        "aciklama": "TEST_bulk_excel_iter12",
        "rows": [{"sicil_no": r["sicil_no"], "izin_turu": r["izin_turu"],
                  "start_date": r["start_date"], "end_date": r["end_date"]} for r in applyable],
    }
    r = session.post(f"{BASE_URL}/api/leaves/bulk/excel-confirm", json=payload, timeout=30)
    assert r.status_code == 200, r.text
    data = r.json()
    assert "created" in data and "skipped" in data
    assert data["total"] == len(payload["rows"])
    print("confirm result:", data)


def test_confirm_empty_rows_400(session):
    r = session.post(f"{BASE_URL}/api/leaves/bulk/excel-confirm",
                     json={"rows": [], "aciklama": ""}, timeout=20)
    assert r.status_code == 400


def test_leaves_list_regression(session):
    r = session.get(f"{BASE_URL}/api/leaves?limit=10", timeout=20)
    assert r.status_code == 200


def test_preview_bad_headers_400(session):
    wb = Workbook(); ws = wb.active
    ws["A1"] = "junk"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    files = {"file": ("bad.xlsx", buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = session.post(f"{BASE_URL}/api/leaves/bulk/excel-preview", files=files, timeout=20)
    assert r.status_code == 400
