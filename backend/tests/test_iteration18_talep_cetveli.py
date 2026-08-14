"""Iteration 18 — İzin Talep Formu (FIFO fix) + İzin Cetveli exports."""
import io
import os
import datetime as dt
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = (os.environ.get("REACT_APP_BACKEND_URL")
            or open("/app/frontend/.env").read().split("REACT_APP_BACKEND_URL=")[1].split()[0]).rstrip("/")
ADMIN = {"email": "muhammedmus@gmail.com", "password": "Merkoteks2026!"}

LEAVE_ID = "74af3f90-4a59-4174-b368-9dbbbb8bbcd2"
PID = "c3f9dcc8-f55e-4a41-a79b-bac1c6ad0d32"


def _d(v):
    """Normalize datetime -> date."""
    if isinstance(v, dt.datetime):
        return v.date()
    return v


@pytest.fixture(scope="module")
def s():
    sess = requests.Session()
    r = sess.post(f"{BASE_URL}/api/auth/login", json=ADMIN, timeout=30)
    assert r.status_code == 200, r.text
    return sess


class TestTalepFormuXlsx:
    def test_xlsx_fields(self, s):
        r = s.get(f"{BASE_URL}/api/leaves/{LEAVE_ID}/talep-formu.xlsx", timeout=60)
        assert r.status_code == 200, r.text
        assert "spreadsheet" in r.headers.get("content-type", "").lower() or \
               "excel" in r.headers.get("content-type", "").lower() or \
               "octet-stream" in r.headers.get("content-type", "").lower()
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["İZİN TALEP FORMU"] if "İZİN TALEP FORMU" in wb.sheetnames else wb.active
        # Log for debugging
        print(f"B8={ws['B8'].value!r}")
        print(f"B10={ws['B10'].value!r} fmt={ws['B10'].number_format}")
        print(f"B11={ws['B11'].value!r}")
        print(f"B12={ws['B12'].value!r} fmt={ws['B12'].number_format}")
        print(f"B13={ws['B13'].value!r} fmt={ws['B13'].number_format}")
        print(f"B14={ws['B14'].value!r} fmt={ws['B14'].number_format}")
        print(f"B15={ws['B15'].value!r}")
        print(f"E15={ws['E15'].value!r}")
        print(f"B29={ws['B29'].value!r} fmt={ws['B29'].number_format}")

        assert ws["B8"].value == "MUHAMMED MUSLU"
        # ise_giris 2021-10-02
        b10 = ws["B10"].value
        assert (_d(b10) == dt.date(2021, 10, 2)) or str(b10).startswith("2021-10-02"), b10
        assert str(ws["B11"].value) == "5650"
        # KEY FIX: earliest FIFO entitlement 2022-10-02 (not 2025-10-02)
        b12 = ws["B12"].value
        assert _d(b12) == dt.date(2022, 10, 2), f"expected 2022-10-02, got {b12}"
        assert "DD.MM.YYYY" in (ws["B12"].number_format or "").upper() or "dd.mm.yyyy" in (ws["B12"].number_format or "")
        # Leave range
        b13 = ws["B13"].value
        b14 = ws["B14"].value
        assert _d(b13) == dt.date(2026, 8, 17), b13
        assert _d(b14) == dt.date(2026, 8, 28), b14
        assert str(ws["B15"].value) in ("10", "10.0", "10 gün", "10 Gün")
        e15 = str(ws["E15"].value or "")
        assert "10.08.2026" in e15 and "Talep" in e15, e15
        # İşbaşı 2026-08-31 (Mon)
        b29 = ws["B29"].value
        assert _d(b29) == dt.date(2026, 8, 31), b29


class TestTalepFormuPdf:
    def test_pdf(self, s):
        r = s.get(f"{BASE_URL}/api/leaves/{LEAVE_ID}/talep-formu.pdf", timeout=90)
        assert r.status_code == 200, r.text[:400]
        assert "application/pdf" in r.headers.get("content-type", "").lower()
        assert len(r.content) > 500
        assert r.content[:5] == b"%PDF-"


class TestCetveliXlsx:
    def test_cetveli_xlsx(self, s):
        r = s.get(f"{BASE_URL}/api/personnel/{PID}/cetveli.xlsx", timeout=60)
        assert r.status_code == 200, r.text
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["İZİN CETVELİ"] if "İZİN CETVELİ" in wb.sheetnames else wb.active
        # find our leave row: look for start date 2026-08-17 in any column
        found = None
        found_col = None
        for row in range(3, 80):
            for c in range(1, 20):
                v = ws.cell(row=row, column=c).value
                if isinstance(v, dt.datetime): v = v.date()
                if v == dt.date(2026, 8, 17):
                    found, found_col = row, c
                    break
            if found: break
        print(f"Found leave row = {found} at col {found_col}")
        # Dump full rows around found
        for row in range(max(3, (found or 8) - 2), (found or 8) + 3):
            vals = [ws.cell(row=row, column=c).value for c in range(1, 16)]
            print(f"row {row}: {vals}")
        assert found is not None, "leave row not found in cetveli"
        yil = ws.cell(row=found, column=1).value
        ent = ws.cell(row=found, column=9).value  # I = entitlement
        sure = ws.cell(row=found, column=11).value  # K = Süre
        bas = ws.cell(row=found, column=13).value  # M = Baş
        don = ws.cell(row=found, column=14).value  # N = Dönüş
        print(f"YIL={yil} ent={ent} sure={sure} bas={bas} don={don}")
        assert _d(ent) == dt.date(2022, 10, 2), f"entitlement expected 2022-10-02, got {ent}"
        assert str(sure) in ("10", "10.0")
        assert _d(bas) == dt.date(2026, 8, 17)
        assert _d(don) == dt.date(2026, 8, 31)
        assert "DD.MM.YYYY" in (ws.cell(row=found, column=9).number_format or "").upper()


class TestCetveliPdf:
    def test_cetveli_pdf(self, s):
        r = s.get(f"{BASE_URL}/api/personnel/{PID}/cetveli.pdf", timeout=90)
        assert r.status_code == 200, r.text[:400]
        assert "application/pdf" in r.headers.get("content-type", "").lower()
        assert len(r.content) > 500
        assert r.content[:5] == b"%PDF-"
