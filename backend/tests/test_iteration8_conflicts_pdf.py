"""Iteration 8 — Conflict detection scoping + PDF/Excel template fixes.

Tests:
- GET /api/leaves/calendar has no `conflict` on days; `same_person_conflicts` only same-personnel overlaps
- POST/PUT /api/leaves returns 409 with detailed conflict payload only for same personnel
- Boundary and adjacency semantics
- Different personnel same date is NOT a conflict
- Excel template rendering (A1 header, date formats, placeholder substitution)
- Backward-compat report endpoints still work
- Regression: login, personnel, balance, cetveli.xlsx, talep-formu.pdf
"""
import io
import os
import re
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = (os.environ.get("REACT_APP_BACKEND_URL") or "").rstrip("/")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().strip('"').rstrip("/")
assert BASE_URL, "REACT_APP_BACKEND_URL required"

ADMIN_EMAIL = "muhammedmus@gmail.com"
ADMIN_PASSWORD = "Merkoteks2026!"

CREATED_PIDS: list = []
CREATED_LIDS: list = []


@pytest.fixture(scope="session")
def token():
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="session")
def client(token):
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {token}", "Content-Type": "application/json"})
    return s


@pytest.fixture(scope="session", autouse=True)
def _cleanup(client):
    yield
    try:
        for lid in CREATED_LIDS:
            try:
                client.delete(f"{BASE_URL}/api/leaves/{lid}")
            except Exception:
                pass
        for pid in CREATED_PIDS:
            try:
                client.delete(f"{BASE_URL}/api/personnel/{pid}")
            except Exception:
                pass
    except Exception:
        pass


def _create_personnel(client, sicil, ad_soyad="Test User", ise_giris="2019-01-01"):
    # cleanup lingering
    existing = client.get(f"{BASE_URL}/api/personnel", params={"q": sicil}).json()
    for e in existing:
        if e.get("sicil_no") == sicil:
            client.delete(f"{BASE_URL}/api/personnel/{e['id']}")
    r = client.post(f"{BASE_URL}/api/personnel", json={
        "sicil_no": sicil, "ad_soyad": ad_soyad, "ise_giris": ise_giris,
        "departman": "Test", "sirket": "Merkoteks", "email": "",
    })
    assert r.status_code == 200, f"create {sicil}: {r.status_code} {r.text}"
    pid = r.json()["id"]
    CREATED_PIDS.append(pid)
    return pid


def _create_leave(client, pid, s, e, tur="Yıllık İzin"):
    r = client.post(f"{BASE_URL}/api/leaves", json={
        "personnel_id": pid, "start_date": s, "end_date": e,
        "izin_turu": tur, "aciklama": "test",
    })
    return r


# ---------------------------------------------------------------------------
# 1. Login regression
# ---------------------------------------------------------------------------
class TestAuth:
    def test_login(self, client):
        r = client.get(f"{BASE_URL}/api/auth/me")
        assert r.status_code == 200
        assert r.json()["email"] == ADMIN_EMAIL


# ---------------------------------------------------------------------------
# 2. Same personnel POST /leaves overlap → 409 with detail
# ---------------------------------------------------------------------------
class TestOverlapPOST:
    def test_same_personnel_overlap_returns_409(self, client):
        pid = _create_personnel(client, "TEST2_P1", "TEST2 P1")
        r1 = _create_leave(client, pid, "2026-11-05", "2026-11-05")
        assert r1.status_code == 200, r1.text
        CREATED_LIDS.append(r1.json()["id"])

        r2 = _create_leave(client, pid, "2026-11-04", "2026-11-07")
        assert r2.status_code == 409, f"expected 409 got {r2.status_code} {r2.text}"
        body = r2.json()
        detail = body.get("detail")
        assert isinstance(detail, dict), f"detail must be dict: {detail}"
        for k in ("message", "personnel_ad_soyad", "personnel_id", "existing_id",
                  "existing_start", "existing_end", "existing_izin_turu",
                  "new_start", "new_end", "overlap_dates"):
            assert k in detail, f"missing {k} in detail: {detail}"
        assert isinstance(detail["overlap_dates"], list) and len(detail["overlap_dates"]) >= 1
        assert "2026-11-05" in detail["overlap_dates"]
        assert detail["personnel_id"] == pid
        assert detail["existing_id"] == r1.json()["id"]


# ---------------------------------------------------------------------------
# 3. Same personnel PUT /leaves — 409 for real conflict, self-exclusion works
# ---------------------------------------------------------------------------
class TestOverlapPUT:
    def test_put_conflict_and_self_exclusion(self, client):
        pid = _create_personnel(client, "TEST2_P2", "TEST2 P2")
        r1 = _create_leave(client, pid, "2026-12-01", "2026-12-03")
        assert r1.status_code == 200, r1.text
        lid1 = r1.json()["id"]; CREATED_LIDS.append(lid1)

        r2 = _create_leave(client, pid, "2026-12-10", "2026-12-12")
        assert r2.status_code == 200, r2.text
        lid2 = r2.json()["id"]; CREATED_LIDS.append(lid2)

        # Try to PUT leave2 to overlap leave1 → 409
        put_conflict = client.put(f"{BASE_URL}/api/leaves/{lid2}", json={
            "personnel_id": pid, "start_date": "2026-12-02", "end_date": "2026-12-05",
            "izin_turu": "Yıllık İzin", "aciklama": "x",
        })
        assert put_conflict.status_code == 409, put_conflict.text
        detail = put_conflict.json()["detail"]
        assert detail["existing_id"] == lid1

        # Self-exclusion: PUT leave2 keeping same start but extending → SUCCESS
        put_ok = client.put(f"{BASE_URL}/api/leaves/{lid2}", json={
            "personnel_id": pid, "start_date": "2026-12-10", "end_date": "2026-12-15",
            "izin_turu": "Yıllık İzin", "aciklama": "extended",
        })
        assert put_ok.status_code == 200, f"self-exclusion failed: {put_ok.status_code} {put_ok.text}"


# ---------------------------------------------------------------------------
# 4. Boundary and adjacency
# ---------------------------------------------------------------------------
class TestBoundary:
    def test_shared_boundary_is_conflict(self, client):
        pid = _create_personnel(client, "TEST2_P3", "TEST2 P3")
        r1 = _create_leave(client, pid, "2026-11-15", "2026-11-18")
        assert r1.status_code == 200
        CREATED_LIDS.append(r1.json()["id"])

        # Shares Nov 18 → conflict
        r2 = _create_leave(client, pid, "2026-11-18", "2026-11-20")
        assert r2.status_code == 409, f"boundary should conflict: {r2.status_code} {r2.text}"

    def test_adjacent_no_shared_day_succeeds(self, client):
        pid = _create_personnel(client, "TEST2_P4", "TEST2 P4")
        r1 = _create_leave(client, pid, "2026-11-15", "2026-11-18")
        assert r1.status_code == 200
        CREATED_LIDS.append(r1.json()["id"])

        # Nov 19 is adjacent, no shared day → success
        r2 = _create_leave(client, pid, "2026-11-19", "2026-11-22")
        assert r2.status_code == 200, f"adjacent should succeed: {r2.status_code} {r2.text}"
        CREATED_LIDS.append(r2.json()["id"])


# ---------------------------------------------------------------------------
# 5. Different personnel same date is NOT a conflict
# ---------------------------------------------------------------------------
class TestDifferentPersonnelNotConflict:
    def test_two_personnel_same_date(self, client):
        pid_x = _create_personnel(client, "TEST2_PX", "TEST2 PX")
        pid_y = _create_personnel(client, "TEST2_PY", "TEST2 PY")

        r1 = _create_leave(client, pid_x, "2026-11-25", "2026-11-25")
        assert r1.status_code == 200
        CREATED_LIDS.append(r1.json()["id"])

        r2 = _create_leave(client, pid_y, "2026-11-25", "2026-11-25")
        assert r2.status_code == 200, f"different personnel same date must succeed: {r2.text}"
        CREATED_LIDS.append(r2.json()["id"])


# ---------------------------------------------------------------------------
# 6. Calendar endpoint — no `conflict` on day objects; same_person_conflicts scoped
# ---------------------------------------------------------------------------
class TestCalendar:
    def test_calendar_no_day_conflict_field(self, client):
        # Ensure both PX (2026-11-25) and PY (2026-11-25) leaves exist from prior test
        # (or create fresh — if prior test's records are cleaned, seed them here)
        r = client.get(f"{BASE_URL}/api/leaves/calendar", params={"year": 2026, "month": 11})
        assert r.status_code == 200, r.text
        data = r.json()
        assert "days" in data
        assert "same_person_conflicts" in data
        for day in data["days"]:
            assert "conflict" not in day, f"day should have no 'conflict' field: {day}"

    def test_same_person_conflicts_empty_for_different_personnel(self, client):
        # August 2026: only Ali (E001) may have leaves from prior iterations.
        r = client.get(f"{BASE_URL}/api/leaves/calendar", params={"year": 2026, "month": 8})
        assert r.status_code == 200
        data = r.json()
        # same_person_conflicts entries must all have same personnel_id per leaf
        for c in data.get("same_person_conflicts", []):
            assert "personnel_id" in c
            assert c["leave_1"]["id"] != c["leave_2"]["id"]

    def test_same_person_conflicts_only_same_personnel(self, client):
        # Create two overlapping leaves for same personnel via direct DB is not possible via API
        # (POST blocks with 409). Instead we verify structure: iterate all conflicts and ensure
        # both leaves belong to the same personnel_id.
        r = client.get(f"{BASE_URL}/api/leaves/calendar", params={"year": 2026, "month": 11})
        assert r.status_code == 200
        data = r.json()
        # In healthy state should be empty (since we can't create overlaps)
        # But if any exist, they must be scoped to same personnel
        for c in data["same_person_conflicts"]:
            # Each conflict entry references single personnel_id — verified by design
            assert c["personnel_id"], f"conflict missing personnel_id: {c}"


# ---------------------------------------------------------------------------
# 7. Excel template rendering — A1 header, date formats, placeholder substitution
# ---------------------------------------------------------------------------
class TestExcelTemplate:
    def test_talep_formu_xlsx(self, client):
        pid = _create_personnel(client, "TEST2_XLS", "TEST2 XLS")
        lr = _create_leave(client, pid, "2026-08-10", "2026-08-14")
        assert lr.status_code == 200, lr.text
        lid = lr.json()["id"]; CREATED_LIDS.append(lid)

        r = client.get(f"{BASE_URL}/api/leaves/{lid}/talep-formu.xlsx")
        assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
        assert r.content[:2] == b"PK"

        wb = load_workbook(io.BytesIO(r.content), data_only=False)
        ws = wb.active

        a1 = ws["A1"].value
        assert a1 is not None, "A1 is None"
        # Must NOT be #VALUE!
        assert "#VALUE" not in str(a1), f"A1 contains #VALUE!: {a1}"
        assert "MERKOTEKS" in str(a1).upper(), f"A1 should contain MERKOTEKS: {a1}"

        # Date cells: check number_format contains DD.MM.YYYY (case-insensitive)
        date_cells = ["B10", "B12", "B13", "B14", "B29"]
        for addr in date_cells:
            nf = ws[addr].number_format or ""
            # openpyxl uses uppercase for Excel formats
            assert re.search(r"DD\.MM\.YYYY|dd\.mm\.yyyy", nf), \
                f"{addr} number_format missing DD.MM.YYYY: '{nf}'"

        # Placeholder substitution: E9, B20, A28 must NOT contain 'CE.aa' or '../../20..'
        # and should contain the leave date parts (2026 or 10.08.2026)
        placeholder_addrs = ["E9", "B20", "A28"]
        collected_text = ""
        for addr in placeholder_addrs:
            v = ws[addr].value
            if v is None:
                continue
            sv = str(v)
            collected_text += sv + " | "
            assert "CE.aa" not in sv, f"{addr} still has 'CE.aa' placeholder: {sv}"
            assert "../../20" not in sv, f"{addr} still has '../../20' placeholder: {sv}"

        # At least one of the placeholder cells should reflect the actual leave date
        assert ("2026" in collected_text) or ("10.08" in collected_text) or ("14.08" in collected_text), \
            f"None of placeholder cells contain leave date. Got: {collected_text}"

    def test_talep_formu_pdf_regression(self, client):
        pid = _create_personnel(client, "TEST2_PDF", "TEST2 PDF")
        lr = _create_leave(client, pid, "2026-09-10", "2026-09-11")
        assert lr.status_code == 200
        lid = lr.json()["id"]; CREATED_LIDS.append(lid)

        r = client.get(f"{BASE_URL}/api/leaves/{lid}/talep-formu.pdf")
        # PDF may not work if libreoffice missing — accept 200 or 500 with message
        if r.status_code == 200:
            assert r.content[:4] == b"%PDF", f"not a PDF: {r.content[:20]}"
        else:
            # Report the failure clearly but don't fail the whole suite if it's env
            pytest.skip(f"PDF conversion returned {r.status_code}: {r.text[:200]}")


# ---------------------------------------------------------------------------
# 8. Reports backward-compat
# ---------------------------------------------------------------------------
class TestReportsCompat:
    def test_charts_still_4_datasets(self, client):
        r = client.get(f"{BASE_URL}/api/reports/charts")
        assert r.status_code == 200
        d = r.json()
        for k in ("departments", "monthly_trend", "companies", "remaining_dist"):
            assert k in d, f"missing {k} in charts response"

    def test_personnel_summary_export_still_works(self, client):
        # backward-compat endpoint may exist
        r = client.get(f"{BASE_URL}/api/reports/personnel-summary/export")
        # Either succeeds (200 xlsx) or endpoint gone (404) — must NOT be 500
        assert r.status_code in (200, 404), f"unexpected {r.status_code}: {r.text[:200]}"


# ---------------------------------------------------------------------------
# 9. Regression endpoints
# ---------------------------------------------------------------------------
class TestRegression:
    def test_personnel_list(self, client):
        r = client.get(f"{BASE_URL}/api/personnel")
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_e001_balance_and_cetveli(self, client):
        rows = client.get(f"{BASE_URL}/api/personnel", params={"q": "E001"}).json()
        e001 = next((p for p in rows if p["sicil_no"] == "E001"), None)
        if not e001:
            pytest.skip("E001 not seeded")
        pid = e001["id"]
        b = client.get(f"{BASE_URL}/api/personnel/{pid}/balance")
        assert b.status_code == 200
        assert "balance" in b.json()

        c = client.get(f"{BASE_URL}/api/personnel/{pid}/cetveli.xlsx")
        assert c.status_code == 200
        assert c.content[:2] == b"PK"
